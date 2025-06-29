#!/usr/bin/env python
"""
Скрипт для анализа структуры Excel-документа с изображениями
"""
import os
import sys
import django

# Добавляем путь к проекту
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from project.settings import MEDIA_ROOT

# Настройка Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'project.settings')
django.setup()

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

def analyze_template():
    file_path = f'{MEDIA_ROOT}/documents/Счет на оплату № НФ-168 от 17.04.2025.xlsx'
    
    print(f"Анализируем файл: {file_path}")
    print(f"Файл существует: {os.path.exists(file_path)}")
    
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        print(f"Лист: {ws.title}")
        print(f"Строк: {ws.max_row}, Столбцов: {ws.max_column}")
        
        # Анализ структуры
        print("\n=== СТРУКТУРА ДОКУМЕНТА ===")
        
        # Находим основные секции
        sections = {}
        for row in range(1, min(50, ws.max_row + 1)):
            for col in range(1, min(10, ws.max_column + 1)):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str):
                    if 'счет' in cell_value.lower() or 'оплату' in cell_value.lower():
                        sections['header'] = (row, col, cell_value)
                    elif 'поставщик' in cell_value.lower():
                        sections['supplier'] = (row, col, cell_value)
                    elif 'покупатель' in cell_value.lower():
                        sections['buyer'] = (row, col, cell_value)
                    elif '№' in cell_value and 'товар' in cell_value.lower():
                        sections['table_header'] = (row, col, cell_value)
                    elif 'итого' in cell_value.lower():
                        sections['total'] = (row, col, cell_value)
                    elif 'руководитель' in cell_value.lower():
                        sections['signature'] = (row, col, cell_value)
        
        print("Найденные секции:")
        for section, (row, col, value) in sections.items():
            print(f"  {section}: строка {row}, столбец {col} - {value}")
        
        # Показываем первые 30 строк с данными
        print("\n=== СОДЕРЖИМОЕ ДОКУМЕНТА ===")
        for row in range(1, min(31, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(12, ws.max_column + 1)):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    # Обрезаем длинные значения
                    if isinstance(cell_value, str) and len(cell_value) > 40:
                        cell_value = cell_value[:40] + "..."
                    row_data.append(f"{get_column_letter(col)}:{cell_value}")
            
            if row_data:
                print(f"Строка {row:2d}: {row_data}")
        
        # Анализ объединенных ячеек
        print("\n=== ОБЪЕДИНЕННЫЕ ЯЧЕЙКИ ===")
        for merged_range in ws.merged_cells.ranges:
            top_left = ws.cell(merged_range.min_row, merged_range.min_col)
            print(f"{merged_range}: {top_left.value}")
        
        # Анализ стилей
        print("\n=== СТИЛИ ЯЧЕЕК ===")
        for row in range(1, min(15, ws.max_row + 1)):
            for col in range(1, min(8, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    print(f"{get_column_letter(col)}{row}:")
                    print(f"  Значение: {cell.value}")
                    print(f"  Шрифт: {cell.font.name}, размер: {cell.font.size}, жирный: {cell.font.bold}")
                    print(f"  Выравнивание: {cell.alignment.horizontal}, {cell.alignment.vertical}")
                    print()
        
        # Размеры столбцов
        print("=== РАЗМЕРЫ СТОЛБЦОВ ===")
        for col in range(1, ws.max_column + 1):
            width = ws.column_dimensions[get_column_letter(col)].width
            if width:
                print(f"Столбец {get_column_letter(col)}: {width}")
        
        wb.close()
        
    except Exception as e:
        print(f"Ошибка: {e}")
        import traceback
        traceback.print_exc()

def analyze_excel_template_with_images(file_path):
    """Анализирует структуру Excel-документа с изображениями"""
    try:
        print("=== АНАЛИЗ СТРУКТУРЫ EXCEL ДОКУМЕНТА С ИЗОБРАЖЕНИЯМИ ===")
        
        # Загружаем рабочую книгу
        wb = load_workbook(file_path)
        ws = wb.active
        
        print(f"Активный лист: {ws.title}")
        print(f"Размеры: {ws.max_row} строк, {ws.max_column} столбцов")
        print()
        
        # Анализ изображений
        print("=== ИЗОБРАЖЕНИЯ В ДОКУМЕНТЕ ===")
        if hasattr(ws, '_images'):
            for i, img in enumerate(ws._images, 1):
                print(f"Изображение {i}:")
                print(f"  Тип: {type(img).__name__}")
                print(f"  Размеры: {img.width} x {img.height} пикселей")
                print(f"  Позиция: {img.anchor}")
                print(f"  Координаты: {img.anchor._from} -> {img.anchor._to}")
                print()
        else:
            print("Изображения не найдены")
        
        # Анализ объединенных ячеек
        print("=== ОБЪЕДИНЕННЫЕ ЯЧЕЙКИ ===")
        for merged_range in ws.merged_cells.ranges:
            print(f"Объединенный диапазон: {merged_range}")
            # Получаем значение из верхней левой ячейки
            top_left_cell = ws[merged_range.start_cell.coordinate]
            if top_left_cell.value:
                print(f"  Значение: '{top_left_cell.value}'")
            print()
        
        # Анализ стилей ячеек
        print("=== СТИЛИ ЯЧЕЕК ===")
        for row in range(1, min(50, ws.max_row + 1)):  # Анализируем первые 50 строк
            for col in range(1, min(20, ws.max_column + 1)):  # Анализируем первые 20 столбцов
                cell = ws.cell(row=row, column=col)
                if cell.value and str(cell.value).strip():
                    print(f"Ячейка {cell.coordinate}:")
                    print(f"  Значение: '{cell.value}'")
                    if cell.font:
                        print(f"  Шрифт: {cell.font.name}, размер: {cell.font.size}, жирный: {cell.font.bold}")
                    if cell.alignment:
                        print(f"  Выравнивание: {cell.alignment.horizontal}, {cell.alignment.vertical}")
                    if cell.border:
                        print(f"  Границы: {cell.border}")
                    print()
        
        # Анализ размеров столбцов и строк
        print("=== РАЗМЕРЫ СТОЛБЦОВ ===")
        for col in range(1, min(30, ws.max_column + 1)):
            col_letter = ws.cell(row=1, column=col).column_letter
            width = ws.column_dimensions[col_letter].width
            if width:
                print(f"Столбец {col_letter}: ширина = {width}")
        
        print("\n=== РАЗМЕРЫ СТРОК ===")
        for row in range(1, min(50, ws.max_row + 1)):
            height = ws.row_dimensions[row].height
            if height:
                print(f"Строка {row}: высота = {height}")
        
        wb.close()
        
    except Exception as e:
        print(f"Ошибка при анализе файла: {e}")
        import traceback
        traceback.print_exc()

def analyze_images_in_excel(file_path):
    """Специально анализирует изображения в Excel-документе"""
    try:
        print("=== ДЕТАЛЬНЫЙ АНАЛИЗ ИЗОБРАЖЕНИЙ ===")
        
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Проверяем наличие изображений
        if hasattr(ws, '_images') and ws._images:
            print(f"Найдено изображений: {len(ws._images)}")
            
            for i, img in enumerate(ws._images, 1):
                print(f"\nИзображение {i}:")
                print(f"  Тип объекта: {type(img).__name__}")
                
                # Информация о размерах
                if hasattr(img, 'width') and hasattr(img, 'height'):
                    print(f"  Размеры: {img.width} x {img.height} пикселей")
                
                # Информация о позиции
                if hasattr(img, 'anchor'):
                    print(f"  Позиция: {img.anchor}")
                    if hasattr(img.anchor, '_from') and hasattr(img.anchor, '_to'):
                        print(f"  От: {img.anchor._from}")
                        print(f"  До: {img.anchor._to}")
                
                # Информация о файле изображения
                if hasattr(img, 'path'):
                    print(f"  Путь к файлу: {img.path}")
                
                # Дополнительные атрибуты
                for attr in ['name', 'format', 'type']:
                    if hasattr(img, attr):
                        value = getattr(img, attr)
                        if value:
                            print(f"  {attr.capitalize()}: {value}")
        else:
            print("Изображения не найдены в документе")
        
        wb.close()
        
    except Exception as e:
        print(f"Ошибка при анализе изображений: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    # Путь к файлу для анализа
    file_path = "media/documents/Счет на оплату № НФ-168 от 17.04.2025.xlsx"
    
    if os.path.exists(file_path):
        print(f"Анализируем файл: {file_path}")
        print("=" * 60)
        
        # Общий анализ
        analyze_excel_template_with_images(file_path)
        
        print("\n" + "=" * 60)
        
        # Детальный анализ изображений
        analyze_images_in_excel(file_path)
        
    else:
        print(f"Файл не найден: {file_path}")
        print("Убедитесь, что файл существует по указанному пути") 