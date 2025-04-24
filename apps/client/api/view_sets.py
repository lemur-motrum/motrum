import datetime
import email
import json
import math
from operator import itemgetter
import os
import random
import re
from itertools import chain
import traceback
import unicodedata
from xmlrpc.client import boolean
from django.conf import settings
from django.db.models import Prefetch
from django.db import IntegrityError, transaction
from rest_framework.views import APIView
# from regex import F
from django.http import HttpResponseRedirect
from django.shortcuts import redirect, render
from django.template import loader
from django.template.loader import render_to_string
from pathspec import PathSpec
from rest_framework import routers, serializers, viewsets, mixins, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.core.cache import cache
from django.contrib.auth import authenticate, login
from django.db.models import Q, F, OrderBy, Count
from django.db.models import Case, When, Value, IntegerField
from apps import specification
from apps.client.utils import crete_pdf_bill
from apps.core.bitrix_api import (
    add_info_order,
    add_new_order_web,
    add_new_order_web_not_info,
    get_info_for_order_bitrix,
    save_new_doc_bx,
    save_payment_order_bx,
    save_shipment_order_bx,
)
from apps.core.models import BaseInfoAccountRequisites
from apps.logs.utils import error_alert
from apps.notifications.models import Notification

from apps.client.api.serializers import (
    AccountRequisitesSerializer,
    AllAccountRequisitesSerializer,
    ClientRequisitesSerializer,
    ClientSerializer,
    DocumentSerializer,
    EmailsAllWebSerializer,
    EmailsCallBackSerializer,
    LkOrderDocumentSerializer,
    LkOrderSerializer,
    OrderOktSerializer,
    OrderSaveCartSerializer,
    OrderSerializer,
    PhoneClientSerializer,
    RequisitesAddressSerializer,
    RequisitesSerializer,
    RequisitesToOktOrderSerializer,
    RequisitesV2Serializer,
)
from django.contrib.sessions.models import Session
from apps.client.models import (
    STATUS_ORDER,
    STATUS_ORDER_BITRIX,
    AccountRequisites,
    Client,
    ClientRequisites,
    DocumentShipment,
    EmailsCallBack,
    Order,
    PaymentTransaction,
    PhoneClient,
    Requisites,
    RequisitesAddress,
    RequisitesOtherKpp,
)
from apps.core.utils import (
    after_save_order_products,
    check_delite_product_cart_in_upd_spes,
    client_info_bitrix,
    create_info_request_order_1c,
    create_info_request_order_bitrix,
    create_time_stop_specification,
    get_presale_discount,
    json_serial,
    loc_mem_cache,
    save_new_product_okt,
    save_order_web,
    save_specification,
    save_spesif_web,
    send_requests,
)
from apps.core.utils_web import (
    _get_pin,
    _verify_pin,
    get_product_item_data,
    send_email_message,
    send_email_message_html,
    send_pin,
    send_pin_smsru,
)
from apps.product.api.serializers import ProductCartSerializer
from apps.product.models import Cart, CurrencyRate, Lot, Price, Product, ProductCart
from apps.specification.api.serializers import (
    ProductSpecification1cSerializer,
    ProductSpecificationSerializer,
    ProductSpecificationToAddBillSerializer,
    SpecificationSerializer,
)
from apps.specification.models import ProductSpecification, Specification
from apps.specification.utils import crete_pdf_specification, save_shipment_doc
from apps.user.models import AdminUser
from openpyxl import load_workbook

from project.settings import IS_PROD, IS_TESTING, IS_WEB, DADATA_TOKEN, DADATA_SECRET
from dadata import Dadata
from rest_framework.authentication import SessionAuthentication, BasicAuthentication
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.decorators import permission_classes, authentication_classes


class ClientViewSet(viewsets.ModelViewSet):
    queryset = Client.objects.all()
    serializer_class = ClientSerializer

    http_method_names = ["get", "post", "put"]

    # РЕГИСТРАЦИЯ ИЛИ АВТОРИЗАЦИЯ
    @action(detail=False, methods=["post"], url_path=r"login")
    def create_or_login_users(self, request, *args, **kwargs):
        try:
            data = request.data
            # phone = data["phone"].replace(" ", "")
            phone = re.sub(r"[^0-9+]+", r"", data["phone"])
            # if "first_name" in data:
            #     data["is_active"] = first_name
            #     first_name  = data['name']
            # else:
            #     first_name = None
            first_name = data["first_name"]
            pin_user = data["pin"]
            data["is_active"] = True
            data["username"] = phone
            data["first_name"] = first_name
            cart_id = request.COOKIES.get("cart")
            print(data)
            # pin = _get_pin(4)
            # pin = 1111
            # cache.set(phone, pin, 120)
            # первый шаг отправка пин
            if pin_user == "":
                if IS_PROD:
                    pin = _get_pin(4)
                    cache.set(phone, pin, 180)
                    send_pin_smsru(pin, phone)
                else:
                    pin = 1111
                    cache.set(phone, pin, 180)

                return Response(pin_user, status=status.HTTP_200_OK)

            # сравнение пин и логин
            else:
                verify_pin = _verify_pin(phone, int(pin_user))
                print("verify_pin", verify_pin)
                # коды совпадают
                if verify_pin:
                    print("@@@", verify_pin)
                    serializer = self.serializer_class(data=data, many=False)

                    # создание новый юзер
                    if serializer.is_valid():
                        client = serializer.save()
                        if client.manager == None:
                            client.add_manager()

                    # старый юзер логин
                    else:
                        # error = "error"
                        # location = "Логины пользователей2"
                        # info = f" ошибка {serializer.errors}"
                        # e = error_alert(error, location, info)
                        
                        client = Client.objects.get(username=phone)
                        serializer = ClientSerializer(client, many=False)

                    # юзер логин
                    if client.is_active:
                        # логин старого пользователя
                        if client.last_login:
                            login(request, client)

                            cart = Cart.objects.filter(client=client, is_active=False)
                            if cart.count() > 0:
                                cart = cart[0]
                                if cart_id is None:
                                    response = Response()
                                    response.data = serializer.data["id"]
                                    response.status = status.HTTP_200_OK
                                    response.set_cookie("cart", cart.id, max_age=2629800)
                                    return response
                                else:
                                    old_cart_prod = ProductCart.objects.filter(
                                        cart=cart.id
                                    ).values_list("product_id", flat=True)
                                    print(old_cart_prod)
                                    product_cart_no_user = ProductCart.objects.filter(
                                        cart=cart_id
                                    )
                                    for products_no_user in product_cart_no_user:
                                        if products_no_user.product.id not in old_cart_prod:
                                            products_no_user.cart = cart
                                            products_no_user.save()

                                    product_cart_no_user.delete()
                                    response = Response()
                                    response.data = serializer.data["id"]
                                    response.status = status.HTTP_200_OK
                                    response.set_cookie("cart", cart.id, max_age=2629800)
                                    return response
                            else:
                                Cart.objects.filter(id=cart_id).update(client=client)
                                return Response(serializer.data, status=status.HTTP_200_OK)

                            # try:
                            #     cart = Cart.objects.get(client=client, is_active=False)
                            #     if cart_id is None:
                            #         response = Response()
                            #         response.data = serializer.data["id"]
                            #         response.status = status.HTTP_200_OK
                            #         response.set_cookie("cart", cart.id, max_age=2629800)
                            #         return response
                            #     else:

                            #         product_cart_no_user = ProductCart.objects.filter(
                            #             cart=cart_id
                            #         )
                            #         for products_no_user in product_cart_no_user:
                            #             products_no_user.cart = cart
                            #             products_no_user.save()

                            #         product_cart_no_user.delete()
                            #         response = Response()
                            #         response.data = serializer.data["id"]
                            #         response.status = status.HTTP_200_OK
                            #         response.set_cookie("cart", cart.id, max_age=2629800)
                            #         return response
                            #         # return Response(serializer.data, status=status.HTTP_200_OK)

                            # except Cart.DoesNotExist:
                            #     Cart.objects.filter(id=cart_id).update(client=client)
                            #     return Response(serializer.data, status=status.HTTP_200_OK)

                        # логин нового пользоваеля
                        else:

                            login(request, client)
                            if cart_id:
                                Cart.objects.filter(id=cart_id).update(client=client)
                            return Response(serializer.data, status=status.HTTP_201_CREATED)

                    else:
                        return Response(serializer.data, status=status.HTTP_403_FORBIDDEN)

                # коды не совпадают
                else:
                    return Response(pin_user, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            print(e)
            tr = traceback.format_exc()
            print(tr)
            error = "error"
            location = "Логины пользователей"
            info = f" ошибка {e}{tr}"
            e = error_alert(error, location, info)
            return Response(e, status=status.HTTP_400_BAD_REQUEST)
        
    @action(detail=False, methods=["post"], url_path=r"get-client-requisites")
    def get_client_requisites(self, request, *args, **kwargs):
        serializer_class = AccountRequisitesSerializer
        data = request.data
        requisites_serch = data["client"]
        queryset = Requisites.objects.filter(
            Q(legal_entity__icontains=requisites_serch)
            | Q(inn__icontains=requisites_serch)
        )
        serializer = RequisitesToOktOrderSerializer(queryset, many=True)
        print(
            serializer.data,
        )
        return Response(serializer.data, status=status.HTTP_200_OK)

    # # добавление клиента через б24
    @action(detail=False, methods=["post"], url_path=r"add-client-bitrix")
    def add_client_bitrix(self, request, *args, **kwargs):
        # data = request.data

        data = {
            "login": {
                "bitrix_id": 13,
                "token": "pbkdf2_sha256$870000$ICTtR17wFHiGIj2sKT2g7d$OM5H9t4fgyMZl8gZVbAcVUB3+GL92fSVg2da03SyhHk=",
            },
            "company": {
                "ИНФА": 22,
            },
        }
        result = "ok"

        data_admin = AdminUser.login_bitrix(data["login"], None, request)
        print(data_admin)
        if data_admin["status_admin"] == 200:

            if result == "ok":
                return Response(result, status=201)
            elif result == "upd":
                return Response(result, status=status.HTTP_200_OK)
            else:
                return Response(result, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response(data_admin, status=data_admin["status_admin"])

    @action(detail=True, methods=["post"], url_path=r"upd-user-lk")
    def update_user_web(self, request, pk=None, *args, **kwargs):
        client = Client.objects.get(pk=pk)
        data = request.data
        # print(data)
        # data = {
        #     "client": {
        #         "last_name": "last_name",
        #         "first_name": "first_name",
        #         "middle_name": "middle_name",
        #         "position": "position",
        #         "email": "steisysi@gamil.com",
        #     },
        #     "phones": [
        #         8927666666,
        #     ],
        # }

        serializer = ClientSerializer(
            client,
            data["client"],
            partial=True,
        )
        if serializer.is_valid():
            print("serializer")
            client_db = serializer.save()
            client_phones = PhoneClient.objects.filter(client=client)
            if client_phones:
                for phone in client_phones:
                    if phone.phone not in data["phone"]:
                        phone.delete()

            if len(data["phone"]) > 0:
                for phone_new in data["phone"]:
                    PhoneClient.objects.get_or_create(client=client, phone=phone_new)

            return Response(serializer.data, status=status.HTTP_200_OK)
        else:

            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    

class ClientRequisitesAccountViewSet(viewsets.ModelViewSet):
    queryset = Client.objects.all()
    serializer_class = ClientRequisitesSerializer
    http_method_names = [
        "get",
        "post",
    ]


class RequisitesViewSet(viewsets.ModelViewSet):
    queryset = Requisites.objects.all()
    serializer_class = RequisitesV2Serializer
    http_method_names = ["get", "post"]

    def get_serializer(self, *args, **kwargs):
        # add many=True if the data is of type list
        if isinstance(kwargs.get("data", {}), list):
            kwargs["many"] = True

        return super(RequisitesViewSet, self).get_serializer(*args, **kwargs)

    @action(detail=False, methods=["post"], url_path=r"add")
    def add_all_requisites(self, request, *args, **kwargs):
        data = request.data
        print(data)
        requisites = data["requisites"]
        requisitesKpp = data["requisitesKpp"]
        adress = data["adress"]
        account_requisites = data["account_requisites"]
        for k, v in adress["legal_adress"].items():
            if v == "" or v == "null" or v == "None":
                print(adress["legal_adress"][k])
                adress["legal_adress"][k] = None

        for k, v in requisitesKpp.items():
            if v == "" or v == "null" or v == "None":
                requisitesKpp[k] = None

        print(adress)
        # i = -1
        valid_all = True

        serializer_data_new = []
        if requisitesKpp["kpp"]:
            type_client = 1
        else:
            type_client = 3

        first_name = None
        last_name = None
        middle_name = None

        if requisites["name"]:
            first_name = requisites["first_name"]

        if requisites["surname"]:
            first_name = requisites["last_name"]

        if requisites["patronymic"]:
            first_name = requisites["middle_name"]

        req = Requisites.objects.update_or_create(
            inn=requisites["inn"],
            defaults={
                "legal_entity": requisites["legal_entity"],
                "type_client": type_client,
                "first_name": first_name,
                "last_name": last_name,
                "middle_name": middle_name,
            },
        )

        if type_client == 1:
            reqKpp = RequisitesOtherKpp.objects.update_or_create(
                requisites=req[0],
                kpp=requisitesKpp["kpp"],
                defaults={
                    "ogrn": requisitesKpp["ogrn"],
                    "legal_post_code": adress["legal_adress"]["post_code"],
                    "legal_city": adress["legal_adress"]["city"],
                    "legal_address": f"{adress["legal_adress"]["legal_address1"]}{adress["legal_adress"]["legal_address2"]}",
                    "postal_post_code": adress["legal_adress"]["post_code"],
                    "postal_city": adress["legal_adress"]["city"],
                    "postal_address": f"{adress["legal_adress"]["legal_address1"]}{adress["legal_adress"]["legal_address2"]}",
                    "tel": requisitesKpp["phone"],
                    "email": requisitesKpp["email"],
                },
            )
        elif type_client == 2:
            reqKpp = RequisitesOtherKpp.objects.update_or_create(
                requisites=req[0],
                ogrn=requisitesKpp["ogrn"],
                defaults={
                    "legal_post_code": adress["legal_adress"]["post_code"],
                    "legal_city": adress["legal_adress"]["city"],
                    "legal_address": f"{adress["legal_adress"]["legal_address1"]}{adress["legal_adress"]["legal_address2"]}",
                    "postal_post_code": adress["postal_adress"]["post_code"],
                    "postal_city": adress["postal_adress"]["city"],
                    "postal_address": f"{adress["postal_adress"]["legal_address1"]}{adress["postal_adress"]["legal_address2"]}",
                    "tel": requisitesKpp["phone"],
                    "email": requisitesKpp["email"],
                },
            )

        clientReq = ClientRequisites.objects.get_or_create(
            client_id=requisites["client"],
            requisitesotherkpp=reqKpp[0],
        )

        adress_req = RequisitesAddress.objects.update_or_create(
            requisitesKpp=reqKpp[0],
            type_address_bx="web-lk-adress",
            defaults={
                "country":adress["legal_adress"]["country"],
                "region":adress["legal_adress"]["region"],
                "province":adress["legal_adress"]["province"],
                "post_code":adress["legal_adress"]["post_code"],
                "city":adress["legal_adress"]["city"],
                "address1":adress["legal_adress"]["legal_address1"],
                "address2":adress["legal_adress"]["legal_address2"],
            },
        )

        for account_req in account_requisites:
            reqAcc = AccountRequisites.objects.update_or_create(
                requisitesKpp=reqKpp[0],
                account_requisites=account_req["account_requisites"],
                defaults={
                    "bank": account_req["bank"],
                    "kpp": account_req["kpp"],
                    "bic": account_req["bic"],
                },
            )

        return Response(["ok"], status=status.HTTP_200_OK)

    @action(detail=True, methods=["post", "get"], url_path=r"update")
    def update_requisites(self, request, pk=None, *args, **kwargs):

        queryset = Requisites.objects.get(pk=pk)
        data = request.data
        serializer_data_new = []
        valid_all = True

        for data_item in data:
            requisites = data_item.get("requisites")
            account_requisites_data = data_item.get("account_requisites")

            serializer = self.serializer_class(queryset, data=requisites, partial=True)
            if serializer.is_valid():
                serializer.save()
                serializer_data_new.append(serializer.data)
                account_requisites = []

                for account_requisites_data_item in account_requisites_data:

                    serializer_class_new = AccountRequisitesSerializer
                    queryset = AccountRequisites.objects.get(
                        pk=account_requisites_data_item["id"]
                    )
                    serializer = serializer_class_new(
                        queryset,
                        data=account_requisites_data_item,
                        partial=True,
                    )
                    if serializer.is_valid():
                        account_requisites_item = serializer.save()
                        account_requisites.append(serializer.data)

                    else:
                        valid_all = False
                        return Response(
                            serializer.errors, status=status.HTTP_400_BAD_REQUEST
                        )
            else:
                valid_all = False
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        if valid_all:
            serializer_data_new[0]["account_requisites"] = account_requisites
            return Response(serializer_data_new, status=status.HTTP_200_OK)

    @action(detail=False, methods=["post"], url_path="serch-requisites")
    def serch_requisites(self, request, *args, **kwargs):

        data = request.data
        with Dadata(DADATA_TOKEN, DADATA_SECRET) as dadata:
            info = dadata.suggest(name="party", query=data["inn"])

            return Response(info, status=status.HTTP_200_OK)


class RequisitesAddressViewSet(viewsets.ModelViewSet):
    queryset = RequisitesAddress.objects.all()
    serializer_class = RequisitesAddressSerializer
    http_method_names = ["get", "post", "patch", "put"]
    

class AccountRequisitesViewSet(viewsets.ModelViewSet):
    queryset = AccountRequisites.objects.all()
    serializer_class = AccountRequisitesSerializer

    http_method_names = ["get", "post", "put", "patch"]


class OrderViewSet(viewsets.ModelViewSet):
    queryset = Order.objects.all()
    serializer_class = OrderSerializer

    http_method_names = [
        "get",
        "post",
        "put",
    ]

    # сохранение заказа с сайта
    @action(detail=False, methods=["post"], url_path=r"add_order")
    def add_order(self, request, *args, **kwargs):
        data = request.data
        try:
            #order_flag - тип сохранения в битркис - со всеми данными или нет если нулл останется- ошибка сохранения
            order_flag = None 
            with transaction.atomic():
                # такая data приходит
                # data = {
                #     "all_client_info": 0,
                #     "client": 33,
                #     "cart": 376,
                #     "requisitesKpp": None,
                #     "account_requisites": None,
                #     "type_delivery": None,
                # }
                cart = int(data["cart"])
                cart = Cart.objects.get(id=data["cart"])
                client = cart.client
                extra_discount = None

                products_cart = ProductCart.objects.filter(cart_id=cart)
                all_info_requisites = False
                all_info_product = True
                requisites_id = None
                account_requisites_id = None
                motrum_requisites = None
                requisites = None
                type_delivery = data["type_delivery"]
                if data["requisitesKpp"] != None:
                    all_info_requisites = True
                    requisitesKpp = RequisitesOtherKpp.objects.get(
                        id=data["requisitesKpp"]
                    )
                    requisitesKpp_id = requisitesKpp.id
                    requisites_id = requisitesKpp.requisites.id
                    requisites = Requisites.objects.get(id=requisites_id)
                    account_requisites = AccountRequisites.objects.get(
                        id=data["account_requisites"]
                    )
                    account_requisites_id = account_requisites.id

                # не используется но будет - рудимент надо или нет привлекать менеджера в заказ в зависимости от заполненности инфо о товарах заказа
                for product_cart in products_cart:
                    if (
                        product_cart.product.price
                        and product_cart.product.price.rub_price_supplier == 0
                    ):
                        all_info_product = False

                # сохранение спецификации для заказа
                status_save_spes, specification, specification_name = save_spesif_web(
                    cart, products_cart, extra_discount, requisites
                )
                
                if status_save_spes == "ok" and specification_name:
                    status_order = "PRE-PROCESSING"

                try:
                    #заказов  с сайта с готовым ордером не бывает - этот кусок для стабильности работы 
                    order = Order.objects.get(cart_id=cart)
                    if data["requisitesKpp"] != None or data["requisitesKpp"] == None:
                        data_order = {
                            "client": client,
                            "name": None,
                            "specification": specification.id,
                            "requisites": (
                                requisites.id if data["requisitesKpp"] != None else None
                            ),
                            "account_requisites": (
                                account_requisites.id
                                if data["requisitesKpp"] != None
                                else None
                            ),
                            "status": "PRE-PROCESSING",
                            "cart": cart.id,
                            "bill_name": None,
                            "bill_file": None,
                            "bill_date_start": None,
                            "bill_date_stop": None,
                            "bill_sum": None,
                            "prepay_persent": (
                                requisites.prepay_persent
                                if data["requisitesKpp"] != None
                                else None
                            ),
                            "postpay_persent": (
                                requisites.postpay_persent
                                if data["requisitesKpp"] != None
                                else None
                            ),
                            
                            "id_bitrix": None,
                            "type_delivery": (
                                type_delivery if data["requisitesKpp"] != None else None
                            ),
                            
                        }

                        serializer = self.serializer_class(order, data=data_order, partial=True)
                        
                        if serializer.is_valid():
                            print("serializer.is_valid(ORDER):")
                            cart.is_active = True
                            cart.save()
                            serializer.save()
                            print("serializer.data", serializer.data)
                            print("serializer.data", serializer.data["id"])
                            order_id = serializer.data["id"]
                            if IS_TESTING:
                                return Response(
                                    serializer.data, status=status.HTTP_201_CREATED
                                )
                            else:
                                if data["requisitesKpp"] != None:
                                    order_flag = "add_new_order_web"
                                else:
                                    order_flag = "add_new_order_web_not_info"
                                    

                        else:
                            return Response(
                                serializer.errors,
                                status=status.HTTP_400_BAD_REQUEST,
                            )
                except Order.DoesNotExist:
                    if data["requisitesKpp"] != None or data["requisitesKpp"] == None:
                        data_order = {
                            "client": client,
                            "name": None,
                            "specification": specification.id,
                            "requisites": (
                                requisites.id if data["requisitesKpp"] != None else None
                            ),
                            "account_requisites": (
                                account_requisites.id
                                if data["requisitesKpp"] != None
                                else None
                            ),
                            "status": "PRE-PROCESSING",
                            "cart": cart.id,
                            "bill_name": None,
                            "bill_file": None,
                            "bill_date_start": None,
                            "bill_date_stop": None,
                            "bill_sum": None,
                            
                            "prepay_persent": (
                                requisites.prepay_persent
                                if data["requisitesKpp"] != None
                                else None
                            ),
                            "postpay_persent": (
                                requisites.postpay_persent
                                if data["requisitesKpp"] != None
                                else None
                            ),
                            
                            "id_bitrix": None,
                            "type_delivery": (
                                type_delivery if data["requisitesKpp"] != None else None
                            ),
                            
                        }

                        serializer = self.serializer_class(data=data_order, many=False)

                        if serializer.is_valid():
                            cart.is_active = True
                            cart.save()
                            serializer.save()
                            order_id = serializer.data["id"]
                            if IS_TESTING:
                                return Response(
                                    serializer.data, status=status.HTTP_201_CREATED
                                )
                            else:
                                if data["requisitesKpp"] != None:
                                    order_flag = "add_new_order_web"
                                else:
                                    order_flag = "add_new_order_web_not_info"
                                    

                        else:
                            return Response(
                                serializer.errors,
                                status=status.HTTP_400_BAD_REQUEST,
                            )
                    else:
                        # эта часть не отрабатывает - так и надо костыль - все данные в ифе берутся 
                        cart.is_active = True
                        cart.save()
                        return Response(None, status=status.HTTP_201_CREATED)

            if order_flag:
                # сохранение еслие сть все данные для создания сделки в битрикс
                if order_flag == "add_new_order_web":
                        status_operation, info = add_new_order_web(order_id)
                        
                # сохранение если данных не достаточно
                else:
                    status_operation, info = add_new_order_web_not_info(
                        order_id
                    )
                if status_operation == "ok":
                    return Response(
                        serializer.data, status=status.HTTP_201_CREATED
                    )
                    
                else:
                    return Response(
                        info,
                        status=status.HTTP_400_BAD_REQUEST,
                    )
            
        except Exception as e:
            print(e)
            tr = traceback.format_exc()
            print(tr)
            error = "error"
            location = "Сохранение заказа с сайта"
            info = f" ошибка {e}{tr}"
            e = error_alert(error, location, info)
            return Response(e, status=status.HTTP_400_BAD_REQUEST)

    # сохранение рыбы Заказа БИТРИКС
    @action(detail=False, methods=["post", "get"], url_path=r"order-bitrix")
    def order_bitrix(self, request, *args, **kwargs):
        try:
            import ast
            data = request.data
            id_bitrix = request.COOKIES.get("bitrix_id_order")
            s = data["serializer"]
            s = unicodedata.normalize('NFKD', s)
            json_acceptable_string = s.replace('"', "").replace("'", '"')
            d = json.loads(json_acceptable_string)

            serializer_class = OrderSerializer
            id_bitrix_serialazer = d["id_bitrix"]
            order = Order.objects.get(id_bitrix=int(id_bitrix_serialazer))
            serializer = serializer_class(order, data=d, many=False)
            if serializer.is_valid():
                order = serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            else:
                tr = serializer.errors
                error = "error"
                location = "взятие заказа при открытие окна битрикс 2"
                info = f" ошибка {tr}{data}"
                e = error_alert(error, location, info)
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            tr = traceback.format_exc()
            error = "error"
            location = "взятие заказа при открытие окна битрикс"
            info = f" ошибка {e}{tr}{s}"
            e = error_alert(error, location, info)
            return Response(None, status=status.HTTP_400_BAD_REQUEST)

    # сохранение спецификации OKT дмин специф
    @action(detail=False, methods=["post"], url_path=r"add-order-admin")
    def add_order_admin(self, request, *args, **kwargs):
        data = request.data
        cart = Cart.objects.get(id=data["id_cart"])
        account_requisites_data = int(data["client_requisites"])
        motrum_requisites_data = int(data["motrum_requisites"])
        id_bitrix = int(data["id_bitrix"])

        type_delivery = data["type_delivery"]

        id_specification = data["id_specification"]
        type_save = data["type_save"]
        admin_creator_id = int(data["admin_creator_id"])
        post_update = data["post_update"]
        type_save = request.COOKIES.get("type_save")
        account_requisites = AccountRequisites.objects.get(id=account_requisites_data)
        motrum_requisites = BaseInfoAccountRequisites.objects.get(
            id=motrum_requisites_data
        )
        requisites = account_requisites.requisitesKpp.requisites

        if requisites.prepay_persent == 100:
            pre_sale = True
        else:
            pre_sale = False

        if requisites.client:
            client = requisites.client
        else:
            client = None

        try:

            with transaction.atomic():
                specification = save_specification(
                    data,
                    pre_sale,
                    request,
                    motrum_requisites,
                    account_requisites,
                    requisites,
                    id_bitrix,
                    type_delivery,
                    post_update,
                    # specification_name,
                    type_save,
                )

        except Exception as e:
            print(e)
            #   добавление в козину удаленного товара при сохранении спецификации из апдейта
            # if data["id_specification"] != None:
            #     product_cart_list = ProductCart.objects.filter(cart=cart).values_list(
            #         "product__id"
            #     )

            #     product_spes_list = ProductSpecification.objects.filter(
            #         specification_id=data["id_specification"]
            #     ).exclude(product_id__in=product_cart_list)

            #     if product_spes_list:
            #         for product_spes_l in product_spes_list:
            #             if product_spes_l.product:
            #                 new = ProductCart(
            #                     cart=cart,
            #                     product=product_spes_l.product,
            #                     quantity=product_spes_l.quantity,
            #                 )
            #                 new.save()

            #             else:
            #                 pass
            tr = traceback.format_exc()
            error = "error"
            location = "Сохранение спецификации админам окт"
            info = f" ошибка {e}{tr}"
            e = error_alert(error, location, info)

        if specification:
            try:
                product_cart = ProductCart.objects.filter(cart=cart).update(date_delivery=None)
                data_order = {
                    "comment": data["comment"],
                    "name": id_bitrix,
                    "specification": specification.id,
                    "requisites": requisites.id,
                    "account_requisites": account_requisites.id,
                    "comment": data["comment"],
                    "prepay_persent": requisites.prepay_persent,
                    "postpay_persent": requisites.postpay_persent,
                    "motrum_requisites": motrum_requisites.id,
                    "type_delivery": type_delivery,
                }

                order = Order.objects.get(cart_id=cart)
                serializer = self.serializer_class(order, data=data_order, many=False)
                if serializer.is_valid():
                    serializer._change_reason = "Ручное"
                    serializer.save()
                    cart.is_active = True
                    cart.save()

                    return Response(serializer.data, status=status.HTTP_200_OK)
                else:
                    return Response(
                        serializer.errors, status=status.HTTP_400_BAD_REQUEST
                    )
            except Order.DoesNotExist:
                data_order = {
                    "client": client,
                    "name": None,
                    "specification": specification.id,
                    "requisites": requisites.id,
                    "account_requisites": account_requisites.id,
                    "status": "PROCESSING",
                    "cart": cart.id,
                    "bill_name": None,
                    "bill_file": None,
                    "bill_date_start": None,
                    "bill_date_stop": None,
                    "bill_sum": None,
                    "comment": data["comment"],
                    "prepay_persent": requisites.prepay_persent,
                    "postpay_persent": requisites.postpay_persent,
                    "motrum_requisites": motrum_requisites.id,
                    "id_bitrix": id_bitrix,
                    "type_delivery": type_delivery,
                    # "manager": admin_creator_id,
                }

                serializer = self.serializer_class(data=data_order, many=False)

                if serializer.is_valid():
                    print("serializer.is_valid():")
                    cart.is_active = True
                    cart.save()
                    serializer.save()

                    return Response(serializer.data, status=status.HTTP_201_CREATED)
                else:
                    print()
                    return Response(
                        serializer.errors, status=status.HTTP_400_BAD_REQUEST
                    )
        else:
            return Response(None, status=status.HTTP_400_BAD_REQUEST)

    # НЕ ИСПОЛЬЗУЕТСЯ сохранение заказа в корзины отложенные без документов для дальнейшего использования
    @action(detail=False, methods=["post"], url_path=r"add-order-no-spec-admin")
    def add_order_no_spec_admin(self, request, *args, **kwargs):
        data = request.data
        cart = Cart.objects.get(id=data["id_cart"])
        type_save = request.COOKIES.get("type_save")

        if data["id_bitrix"] != None:
            id_bitrix = int(data["id_bitrix"])
        else:
            id_bitrix = None

        if data["client_requisites"] != None:
            account_requisites_data = int(data["client_requisites"])
            account_requisites = AccountRequisites.objects.get(
                id=account_requisites_data
            )
            requisites = account_requisites.requisitesKpp.requisites
            requisites_id = requisites.id
            account_requisites_id = account_requisites.id
            prepay_persent = requisites.prepay_persent
            postpay_persent = requisites.postpay_persent
            if requisites.client:
                client = requisites.client
            else:
                client = None
        else:
            account_requisites_id = None
            requisites_id = None
            client = None
            prepay_persent = None
            postpay_persent = None

        if data["motrum_requisites"] != None:
            motrum_requisites_id = int(data["motrum_requisites"])
        else:
            motrum_requisites_id = None

        if data["type_delivery"] != None:
            type_delivery = data["type_delivery"]
        else:
            type_delivery = None

        data_order = {
            "client": client,
            "name": None,
            "specification": None,
            "requisites": requisites_id,
            "account_requisites": account_requisites_id,
            "status": "",
            "cart": cart.id,
            "bill_name": None,
            "bill_file": None,
            "bill_date_start": None,
            "bill_date_stop": None,
            "bill_sum": None,
            "comment": data["comment"],
            "prepay_persent": prepay_persent,
            "postpay_persent": postpay_persent,
            "motrum_requisites": motrum_requisites_id,
            "id_bitrix": id_bitrix,
            "type_delivery": type_delivery,
        }

        try:
            order = Order.objects.get(cart=cart)
            serializer = self.serializer_class(order, data=data_order, many=False)
            if serializer.is_valid():
                serializer._change_reason = "Ручное"
                order = serializer.save()
                cart.is_active = True
                cart.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        except Order.DoesNotExist:
            serializer = self.serializer_class(data=data_order, many=False)
            if serializer.is_valid():
                cart.is_active = True
                cart.save()
                order = serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            else:

                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    # ОКТ создание счета к заказу - отпарвка 1с после и битркис
    @action(
        detail=True,
        methods=[
            "get",
            "post",
        ],
        url_path=r"create-bill-admin",
    )
    def create_bill_admin(self, request, pk=None, *args, **kwargs):

        try:
            import json

            user = request.user
            data_get = request.data
            type_save = request.COOKIES.get("type_save")
            post_update = data_get["post_update"]
            products = data_get["products"]
            order = Order.objects.get(specification_id=pk)
            # увеломление о наличии документа спецификации для пользователя сайта
            if order.specification.file and order.client:
                Notification.add_notification(
                    order.id, "DOCUMENT_SPECIFICATION", order.specification.file
                )

            for obj in products:
                prod = ProductSpecification.objects.filter(id=obj["id"]).update(
                    text_delivery=obj["text_delivery"]
                )

            if order.requisites.contract:
                is_req = True
            else:
                is_req = False
                
            # сохранение товара в окт нового
            order_products = after_save_order_products(products)

            # создание пдф счетов
            order_pdf = order.create_bill(
                request,
                is_req,
                order,
                # bill_name,
                post_update,
                type_save,
            )

            if order_pdf:
                pdf = request.build_absolute_uri(order.bill_file_no_signature.url)
                pdf_signed = request.build_absolute_uri(order.bill_file.url)
                if order.specification.file:
                    document_specification = request.build_absolute_uri(
                        order.specification.file.url
                    )
                else:
                    document_specification = None
                data = {"pdf": pdf, "name_bill": order.bill_name}

                
                data_for_1c = create_info_request_order_1c(order, order_products)
                
                tr = traceback.format_exc()
                error = "info_error_order"
                location = "ИНФО НЕ ОШИБКА"
                info = f"ИНФО НЕ ОШИБКА данные по заказу для отпарвки в 1с data_for_1c{data_for_1c}"
                e = error_alert(error, location, info)

                type_save = request.COOKIES.get("type_save")

                if IS_TESTING or user.username == "testadmin":
                    pass
                    # json_data = json.dumps(data_for_1c)
                    # print("json_data", json_data)
                    # if user.username == "testadmin":
                    #     print("if IS_TESTING or user.username == testadmin")
                    #     url = "https://dev.bmgspb.ru/grigorev_unf_m/hs/rest/order"
                    #     headers = {"Content-type": "application/json"}
                    #     response = send_requests(url, headers, json_data, "1c")
                    #     print(response)
                    # pass
                else:
                    json_data = json.dumps(data_for_1c)
                    url = "https://dev.bmgspb.ru/grigorev_unf_m/hs/rest/order"
                    headers = {"Content-type": "application/json"}
                    # отправка в 1с
                    response = send_requests(url, headers, json_data, "1c")
                    # отправка в битрикс
                    add_info_order(request, order, type_save)

                return Response(data, status=status.HTTP_200_OK)

            else:
                return Response(None, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:

            tr = traceback.format_exc()
            error = "error"
            location = "Сохранение счета админам окт"
            info = f"Сохранение счета админам окт ошибка {e}{tr}"
            e = error_alert(error, location, info)

            return Response(e, status=status.HTTP_400_BAD_REQUEST)

    @action(
        detail=False,
        methods=[
            "get",
            "post",
        ],
        url_path=r"test1s",
    )
    def test1s(self, request, *args, **kwargs):
        data = request.data
        print("test1s", data)
        return Response(data, status=status.HTTP_200_OK)

    # НЕ ИСПОЛЬЗУЕТСЯ ОКТ изменение спецификации дмин специф
    @action(
        detail=True,
        methods=[
            "get",
            "post",
        ],
        url_path=r"update-order-admin",
    )
    def update_order_admin(self, request, pk=None, *args, **kwargs):
        data = request.data
        cart_id = request.COOKIES.get("cart")
        cart = Cart.objects.filter(id=cart_id).update(is_active=False)
        return Response(cart, status=status.HTTP_200_OK)

    # ОКТ выйти из изменения без сохранения спецификации дмин специф
    @action(
        detail=False,
        methods=[
            "get",
            "post",
        ],
        url_path=r"exit-order-admin",
    )
    def exit_order_admin(self, request, *args, **kwargs):
        cart_id = request.COOKIES.get("cart")
        # specification = request.COOKIES.get("specificationId")
        cart = Cart.objects.filter(id=cart_id).update(is_active=True)
        # if specification and specification != 0 and specification != "":
        #     check_delite_product_cart_in_upd_spes(specification,cart)
        
        return Response(cart, status=status.HTTP_200_OK)

    # ОКТ получить список товаров для создания счета с датами псотавки 
    @action(detail=True, methods=["get"], url_path=r"get-specification-product")
    def get_specification_product(self, request, pk=None, *args, **kwargs):

        product_specification = ProductSpecification.objects.filter(specification=pk)
        if product_specification.exists():
            serializer = ProductSpecificationToAddBillSerializer(
                product_specification, many=True
            )
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            return Response(None, status=status.HTTP_400_BAD_REQUEST)

    # САЙТ страница мои заказов  аякс загрузка
    @action(detail=False, url_path="load-ajax-order-list")
    def load_ajax_order_list(self, request):
        # закоменченное в этой функции не удалаять!
        from django.db.models.functions import Length

        current_count = int(request.query_params.get("count"))
        page_get = request.query_params.get("page")
        count_last = 5

        serializer_class = LkOrderSerializer
        current_user = request.user.id

        client = Client.objects.get(pk=current_user)
        req_user_all = ClientRequisites.objects.filter(client=client)
        print(req_user_all)
        all_inn = []
        all_innkpp =[]
        for req_user in req_user_all:
            print(req_user)
            if req_user.requisitesotherkpp.requisites.id not in all_inn:
                all_inn.append(req_user.requisitesotherkpp.requisites.id)
            
            print(req_user.requisitesotherkpp.accountrequisites_set.all())
            for acc in req_user.requisitesotherkpp.accountrequisites_set.all():
                if acc.requisitesKpp.id not in all_innkpp:
                    all_innkpp.append(acc.requisitesKpp.id)    
      
        print(req_user_all)
        print("all_innkpp",all_innkpp)
        # сортировки
        sorting = "-notification_count"
        # direction = "ASC"
        if request.query_params.get("sort"):
            sorting = request.query_params.get("sort")
        if request.query_params.get("direction"):
            direction = request.query_params.get("direction")

        if sorting == "date_order":
            if direction == "ASC":
                sorting = F("date_order").asc(nulls_last=True)
            else:
                sorting = F("date_order").desc(nulls_last=True)

        if sorting == "bill_sum":
            if direction == "ASC":
                sorting = F("bill_sum").asc(nulls_last=True)
            else:
                sorting = F("bill_sum").desc(nulls_last=True)

        if sorting == "status":
            if direction == "ASC":
                sorting = Case(
                    When(status="PRE-PROCESSING", then=Value(1)),
                    When(status="PROCESSING", then=Value(2)),
                    When(status="PAYMENT", then=Value(3)),
                    When(status="IN_MOTRUM", then=Value(4)),
                    When(status="SHIPMENT_AUTO", then=Value(5)),
                    When(status="SHIPMENT_PICKUP", then=Value(6)),
                    When(status="CANCELED", then=Value(7)),
                    When(status="COMPLETED", then=Value(8)),
                    output_field=IntegerField(),
                )
            else:
                sorting = Case(
                    When(status="PRE-PROCESSING", then=Value(8)),
                    When(status="PROCESSING", then=Value(7)),
                    When(status="PAYMENT", then=Value(6)),
                    When(status="IN_MOTRUM", then=Value(5)),
                    When(status="SHIPMENT_AUTO", then=Value(4)),
                    When(status="SHIPMENT_PICKUP", then=Value(3)),
                    When(status="CANCELED", then=Value(2)),
                    When(status="COMPLETED", then=Value(1)),
                    output_field=IntegerField(),
                )

        # ОСТАТКИ СЛОЖНОЙ СОРТИРОВКИ ПО НЕСКОЛЬКИМ ПОЛЯм ОДНОВРЕМЕННО НЕ УДАЛЯТЬ
        # if request.query_params.get("sortDate"):
        #     date_get = request.query_params.get("sortDate")
        # else:
        #     date_get = "ASC"
        #     # date_get = None

        # if request.query_params.get("sortSum"):
        #     sum_get = request.query_params.get("sortSum")
        # else:
        #     # sum_get = "ASC"
        #     sum_get = None

        # if request.query_params.get("sortStatus"):
        #     status_get = request.query_params.get("sortStatus")
        # else:
        #     status_get = "ASC"
        # status_get = None

        # if date_get:
        #     sort = "date_order"
        #     if date_get == "ASC":
        #         ordering_filter_date = F(sort).asc(nulls_last=True)
        #     else:
        #         ordering_filter_date = F(sort).desc(nulls_last=True)

        # else:
        #     ordering_filter_date = None

        # if sum_get:
        #     sort = "bill_sum"
        #     if sum_get == "ASC":
        #         ordering_filter_summ = F(sort).asc(nulls_last=True)
        #
        #     else:
        #         ordering_filter_summ = F(sort).desc(nulls_last=True)
        # else:
        #     ordering_filter_summ = None

        # if status_get:
        #     sort = "status"
        #     if status_get == "ASC":

        #         custom_order = Case(
        #             When(status="PROCESSING", then=Value(1)),
        #             When(status="PAYMENT", then=Value(2)),
        #             When(status="IN_MOTRUM", then=Value(3)),
        #             When(status="SHIPMENT_AUTO", then=Value(4)),
        #             When(status="SHIPMENT_PICKUP", then=Value(5)),
        #             When(status="CANCELED", then=Value(6)),
        #             When(status="COMPLETED", then=Value(7)),
        #             output_field=IntegerField(),
        #         )

        #     else:
        #         custom_order = Case(
        #             When(status="PROCESSING", then=Value(7)),
        #             When(status="PAYMENT", then=Value(6)),
        #             When(status="IN_MOTRUM", then=Value(5)),
        #             When(status="SHIPMENT_AUTO", then=Value(4)),
        #             When(status="SHIPMENT_PICKUP", then=Value(3)),
        #             When(status="CANCELED", then=Value(2)),
        #             When(status="COMPLETED", then=Value(1)),
        #             output_field=IntegerField(),
        #         )
        # else:
        #     custom_order = None
        # для такой сортировки если нет данных использовать custom_order ordering_filter_summ ordering_filter_date = None

        # ordering = {
        # 'ordering_filter_date': ordering_filter_date,
        # 'ordering_filter_summ': ordering_filter_summ,
        # 'custom_order': custom_order,
        # }
        # ordering = {k:v for k,v in ordering.items() if v is not None}
        # ordering = list(ordering.values())
        # lot = Lot.objects.all()
        
     
        orders = (
            # all_innkpp
            Order.objects.filter( Q(account_requisites__requisitesKpp_id__in=all_innkpp) | Q(client=client))
            # Order.objects.filter( Q(requisites_id__in=all_inn) | Q(client=client))
            # Order.objects.filter( Q(requisites_id__in=all_inn) & Q(client=client))
            # Q(requisites_id__in=all_inn) & Q(client=client)
            # Order.objects.filter(client=client)
            .select_related(
                "specification",
                "cart",
                "requisites",
                "account_requisites",
            )
            # .prefetch_related(
            #     Prefetch(
            #         "notification_set",
            #         queryset=Notification.objects.filter(
            #             type_notification="STATUS_ORDERING", is_viewed=False
            #         ),
            #         to_attr="filtered_notification_items",
            #     )
            # )
            .prefetch_related(
                Prefetch("specification__productspecification_set")
            ).prefetch_related(
                Prefetch("specification__productspecification_set__product")
            )
            # .prefetch_related(
            #     Prefetch('specification__productspecification_set__product__stock'))
            # .prefetch_related(
            #     Prefetch('specification__productspecification_set__product__stock__lot'))
            .annotate(
                notification_count=Count(
                    "notification",
                    filter=Q(
                        notification__is_viewed=False,
                        notification__type_notification="STATUS_ORDERING",
                    ),
                )
            )
        )
        page_count = orders.count()

        if page_count % count_last == 0:
            count = page_count / count_last
        else:
            count = math.trunc(page_count / count_last) + 1

        if page_count <= count_last * 2:
            small = True
        else:
            small = False

        orders_sorting = orders.order_by(sorting, "-id")[
            current_count : current_count + count_last
        ]

        queryset_next = orders[
            current_count + count_last : current_count + count_last + 1
        ].exists()

        # [count : count + count_last]
        serializer = serializer_class(orders_sorting, many=True)
        data = serializer.data
        
        # прочитать уведомления только выведенные ордеры
        for data_order in data:
            if int(data_order["notification_count"]) > 0:
                Notification.objects.filter(
                    order=data_order["id"],
                    client_id=current_user,
                    type_notification="STATUS_ORDERING",
                    is_viewed=False,
                ).update(is_viewed=True)

        # прочитать уведомления всех ордеров
        # notifications = Notification.objects.filter(
        #     client_id=current_user, type_notification="STATUS_ORDERING", is_viewed=False
        # ).update(is_viewed=True)

        # if sorting == "-id":
        #     data = sorted(data, key=lambda x: len(x["notification_set"]), reverse=True)

        data_response = {
            "data": data,
            "page": page_get,
            "count": count,
            "small": small,
            "next": queryset_next,
        }
        return Response(data=data_response, status=status.HTTP_200_OK)

    # САЙТ страница мои документы аякс загрузка
    @action(detail=False, url_path="load-ajax-document-list")
    def load_ajax_document_list(self, request):
        # закоменченное в этой функции не удалаять!
        
        current_count = int(request.query_params.get("count"))
        page_get = request.query_params.get("page")
        count_last = 10
        serializer_class = LkOrderDocumentSerializer
        current_user = request.user.id
        client = Client.objects.get(pk=current_user)
        req_user_all = ClientRequisites.objects.filter(client=client)

        all_inn = []
        all_innkpp =[]
        for req_user in req_user_all:
            if req_user.requisitesotherkpp.requisites.id not in all_inn:
                all_inn.append(req_user.requisitesotherkpp.requisites.id)
            for acc in req_user.requisitesotherkpp.accountrequisites_set.all():
                if acc.requisitesKpp.id not in all_innkpp:
                    all_innkpp.append(acc.requisitesKpp.id)    
        # print(req_user_all)

        # сортировки
        sorting = None
        sorting_directing = None
        # sorting_spesif = "specification__date"
        # sorting_bill = "bill_date_start"
        if request.query_params.get("sort"):
            sorting = request.query_params.get("sort")

        if request.query_params.get("direction"):
            sorting_directing = request.query_params.get("direction")

        # заказы сериализировать
        orders = (
            Order.objects.filter( Q(account_requisites__requisitesKpp_id__in=all_innkpp) | Q(client=client))
            # Order.objects.filter( Q(requisites_id__in=all_inn) & Q(client=client))
            # Order.objects.filter(requisites_id__in=all_inn)
            # Order.objects.filter(client=client)
            .select_related(
                "specification",
                "cart",
                "requisites",
                "account_requisites",
            ).prefetch_related(
                Prefetch(
                    "notification_set",
                    queryset=Notification.objects.filter(is_viewed=False).exclude(
                        type_notification="STATUS_ORDERING"
                    ),
                    to_attr="filtered_notification_items",
                )
            )
        )

        serializer = serializer_class(orders, many=True)

        # формирование отдельных документов из сериализированных заказов
        documents = []
        print(serializer.data)
        for order in serializer.data:

            # if order["specification_list"]["file"]:
            if  order["specification_list"] and "file" in order["specification_list"] and order["specification_list"]["file"]!= None:
                data_spesif = {
                    "type": 1,
                    "name": "спецификация",
                    "order": order["id"],
                    "status": order["status"],
                    "status_full": order["status_full"],
                    "pdf": order["specification_list"]["file"],
                    "requisites_contract": order["requisites_full"]["contract"],
                    "requisites_legal_entity": order["requisites_full"]["legal_entity"],
                    "date_created": order["specification_list"]["date"],
                    "data_stop": order["specification_list"]["date_stop"],
                    "amount": order["specification_list"]["total_amount"],
                    "notification_set": [],
                    "type_notification": "DOCUMENT_SPECIFICATION",
                    "number_document": order["specification_list"]["number"],
                }

                for notification_set in order["notification_set"]:
                    if (
                        notification_set["type_notification"]
                        == "DOCUMENT_SPECIFICATION"
                    ):

                        data_spesif["notification_set"].append(notification_set)

                documents.append(data_spesif)

            if order["bill_file"]:
                data_bill = {
                    "type": 2,
                    "name": "счёт",
                    "order": order["id"],
                    "status": order["status"],
                    "status_full": order["status_full"],
                    "pdf": order["bill_file"],
                    "requisites_contract": order["requisites_full"]["contract"],
                    "requisites_legal_entity": order["requisites_full"]["legal_entity"],
                    "date_created": order["bill_date_start"],
                    "data_stop": order["bill_date_stop"],
                    "amount": order["bill_sum"],
                    "notification_set": [],
                    "type_notification": "DOCUMENT_BILL",
                    "number_document": order["bill_name"],
                }
                print(data_bill)
                for notification_set in order["notification_set"]:
                    if notification_set["type_notification"] == "DOCUMENT_BILL":
                        data_bill["notification_set"].append(notification_set)

                documents.append(data_bill)

            if len(order["documentshipment_set"]) > 0:
                for doc_shipment in order["documentshipment_set"]:
                    data_act = {
                        "type": 3,
                        "name": "акт",
                        "order": order["cart"],
                        "status": order["status"],
                        "status_full": order["status_full"],
                        "pdf": doc_shipment["file"],
                        "requisites_contract": order["requisites_full"]["contract"],
                        "requisites_legal_entity": order["requisites_full"][
                            "legal_entity"
                        ],
                        "date_created": doc_shipment["date"],
                        "data_stop": doc_shipment["date"],
                        "amount": 0,
                        "notification_set": [],
                        "number_document": doc_shipment["id"],
                    }

                    for notification_set in order["notification_set"]:
                        if notification_set["type_notification"] == "DOCUMENT_ACT":
                            if notification_set["file"] == doc_shipment["file"]:
                                data_spesif["notification_set"].append(notification_set)

                    documents.append(data_act)

        # сортировки для документов
        print(documents)
        print(sorting)
        print(sorting_directing)
        if sorting_directing:
            if sorting_directing == "ASC":
                sorting_directing = False
            elif sorting_directing == "DESC":
                sorting_directing = True
        if sorting:
            if sorting == "date":
                print("sort date")
                documents = sorted(
                    documents,
                    key=lambda x: datetime.datetime.strptime(
                        x["date_created"], "%d.%m.%Y"
                    ),
                    reverse=sorting_directing,
                )
                print(documents)
            else:
                documents = sorted(
                    documents, key=itemgetter(sorting), reverse=sorting_directing
                )
        else:
            documents = sorted(
                documents, key=lambda x: len(x["notification_set"]), reverse=True
            )
        # прочитать уведомления только выведенные ордеры
        # for data_order in documents:
        #     if len(data_order["notification_set"]) > 0:
        #         Notification.objects.filter(order=data_order["order"],
        #         client_id=current_user, type_notification=data_order["type_notification"], is_viewed=False
        #     ).update(is_viewed=True)

        notifications = (
            Notification.objects.filter(client_id=current_user, is_viewed=False)
            .exclude(type_notification="STATUS_ORDERING")
            .update(is_viewed=True)
        )
        page_count = len(documents)

        if page_count % count_last == 0:
            count = page_count / count_last
        else:
            count = math.trunc(page_count / count_last) + 1

        if page_count <= count_last * 2:
            small = True
        else:
            small = False

        queryset_next = len(
            documents[current_count + count_last : current_count + count_last + 1]
        )

        if queryset_next == 1:
            next_elem = True
        else:
            next_elem = False

        data_response = {
            "data": documents[current_count : current_count + count_last],
            "page": page_get,
            "count": count,
            "small": small,
            "next": next_elem,
            "len": page_count,
        }

        return Response(data=data_response, status=status.HTTP_200_OK)

    # ОКТ страница все спецификации окт аякс загрузка
    @action(
        detail=False, methods=["post", "get"], url_path=r"load-ajax-specification-list"
    )
    def load_ajax_specification_list(self, request, *args, **kwargs):
        # закоменченное в этой функции не удалаять!
        
        count = int(request.query_params.get("count"))
        count_last = 10
        page_get = request.query_params.get("page")
        q_object = Q()
        if request.user.is_superuser:
            superadmin = True
        else:
            superadmin = False

        user_admin = AdminUser.objects.get(user=request.user)
        user_admin_type = user_admin.admin_type
        # if user_admin_type == "ALL":
        #     q_object &= Q(cart__cart_admin_id__isnull=False)
        # elif user_admin_type == "BASE":
        #     q_object &= Q(cart__cart_admin_id=request.user.id)

        sort_specif = request.query_params.get("specification")

        if sort_specif == "+":
            q_object &= Q(specification__isnull=True)
        else:
            q_object &= Q(specification__isnull=False)

        iframe = request.query_params.get("frame")
        bx_id_order = request.query_params.get("bx_id_order")

        if iframe == "True":
            q_object &= Q(id_bitrix=int(bx_id_order))
        else:
            if user_admin_type == "ALL":
                pass
                # q_object &= Q(cart__cart_admin_id__isnull=False)
            elif user_admin_type == "BASE":
                q_object &= Q(cart__cart_admin_id=request.user.id)
            # if IS_TESTING:
            #     pass
            # else:
            #     if user_admin_type == "ALL":
            #         q_object &= Q(cart__cart_admin_id__isnull=False)
            #     elif user_admin_type == "BASE":
            #         q_object &= Q(cart__cart_admin_id=request.user.id)

        now_date = datetime.datetime.now()
        print(q_object)
        queryset = (
            Order.objects.select_related(
                "specification", "cart", "client", "requisites", "account_requisites"
            )
            .prefetch_related(
                Prefetch("specification__admin_creator"),
                Prefetch("cart__productcart_set"),
                Prefetch("cart__productcart_set__product"),
                Prefetch("cart__cart_admin"),
                # Prefetch("requisites__accountrequisites_set"),
            )
            .filter(q_object)
            .order_by("-id")[count : count + count_last]
        )
        print("queryset", queryset)

        page_count = Order.objects.filter(q_object).count()

        queryset_next = Order.objects.filter(q_object)[
            count + count_last : count + count_last + 1
        ].exists()

        if page_count % 10 == 0:
            count = page_count / 3
        else:
            count = math.trunc(page_count / 10) + 1

        if page_count <= 10:
            small = True
        else:
            small = False

        serializer = OrderOktSerializer(
            queryset, many=True, context={"request": request}
        )
        data_response = {
            "data": serializer.data,
            "superadmin": superadmin,
            "count": math.ceil(page_count / count_last),
            "small": small,
            "page": page_get,
            "next": queryset_next,
        }
        return Response(data=data_response, status=status.HTTP_200_OK)

    # НЕ ИСПОЛЬЗУЕТСЯ ОКТ добавление оплаты открыть получить отстаок суммы
    @action(detail=True, methods=["get"], url_path=r"get-payment")
    def get_payment(self, request, pk=None, *args, **kwargs):
        order = Order.objects.get(id=pk)

        sum_pay = float(order.bill_sum) - float(order.bill_sum_paid)
        data = {
            "sum_pay": sum_pay,
        }
        if sum_pay:
            return Response(data, status=status.HTTP_200_OK)
        else:
            return Response(None, status=status.HTTP_400_BAD_REQUEST)

    # НЕ ИСПОЛЬЗУЕТСЯ ОКТ сохранение суммы оплаты счета
    @action(detail=True, methods=["post"], url_path=r"save-payment")
    def save_payment(self, request, pk=None, *args, **kwargs):
        order = Order.objects.get(id=pk)
        data = request.data
        bill_sum_paid = float(data["bill_sum_paid"])
        date_tarnsaction = datetime.datetime.now()

        old_sum = order.bill_sum_paid
        new_sum = old_sum + bill_sum_paid
        tarnsaction = PaymentTransaction.objects.create(
            order=order, date=date_tarnsaction, amount=bill_sum_paid
        )
        order.bill_sum_paid = new_sum
        order._change_reason = "Ручное"
        order.save()

        if order.bill_sum_paid == order.bill_sum:
            is_all_sum = True
        else:
            is_all_sum = False

        data = {"all_bill_sum_paid": new_sum, "is_all_sum": is_all_sum}
        if new_sum:
            return Response(data, status=status.HTTP_200_OK)
        else:
            return Response(None, status=status.HTTP_400_BAD_REQUEST)

    # НЕ ИСПОЛЬЗУЕТСЯ ОКТ получение суммы уже оплаченной при открытии модалки внесения оплаты
    @action(detail=True, methods=["get"], url_path=r"view-payment")
    def view_payment(self, request, pk=None, *args, **kwargs):
        order = Order.objects.filter(id=pk)
        # data = request.data
        # bill_sum_paid = data["bill_sum_paid"]
        # order.bill_sum_paid = float(bill_sum_paid)

        sum_prepay = order.bill_sum - (
            order.bill_sum / 100 * order.requisites.postpay_persent
        )
        sum_postpay = order.bill_sum - sum_prepay
        if order.bill_sum_paid >= sum_prepay:
            sum_prepay_fact = order.bill_sum_paid
            sum_postpay_fact = order.bill_sum_paid - sum_prepay_fact
        else:
            sum_prepay_fact = sum_prepay - order.bill_sum_paid
            sum_postpay_fact = 0

        data = {
            "sum_prepay": sum_prepay,
            "sum_postpay": sum_postpay,
            "bill_sum": order.bill_sum,
            "bill_sum_paid": order.bill_sum_paid,
            "sum_prepay_fact": sum_prepay_fact,
            "sum_postpay_fact": sum_postpay_fact,
        }
        return Response(data, status=status.HTTP_200_OK)

    # НЕ ИСПОЛЬЗУЕТСЯ добавить дату завершения вручную
    @action(detail=True, methods=["post"], url_path=r"status-order-bitrix")
    def date_completed(self, request, pk=None, *args, **kwargs):
        data = request.data
        date_completed_data = data["date_completed"]

        date_completed = datetime.datetime.strptime(
            date_completed_data, "%Y-%m-%d"
        ).date()
        order = Order.objects.get(pk=pk)
        order.date_completed = date_completed
        order.status = "COMPLETED"
        order._change_reason = "Ручное"
        order.save()
        data = {}
        return Response(data, status=status.HTTP_200_OK)

    # заполнение корзины из фаила кп
    @action(detail=True, methods=["post"], url_path=r"add-file-dowlad")
    def add_file_dowlad(self, request, pk=None, *args, **kwargs):
        pass
        data = request.data
        up_file = data["file"]

        workbook = load_workbook(up_file)
        data_sheet = workbook.active

        for index in range(3, data_sheet.max_row):
            row_level = data_sheet.row_dimensions[index].outline_level + 1
            item_value = data_sheet.cell(row=index, column=2).value

            if row_level == 1:
                vendor_str = data_sheet.cell(row=index, column=1).value


    # ОКТ 1С сроки поставки товаров ОКТ Б24
    @authentication_classes([BasicAuthentication])
    @permission_classes([IsAuthenticated])
    @action(detail=False, methods=["post"],authentication_classes =[BasicAuthentication],permission_classes=[IsAuthenticated], url_path=r"add-info-order-1c")
    def add_info_order_1c(self, request, *args, **kwargs):
        print("add_info_order_1c")
        print('user',request.user,  # `django.contrib.auth.User` instance.
        'auth',str(request.auth),)
        
        data = request.data
        pdf = None
        pdf_signed = None
        try:
            order = Order.objects.get(id_bitrix=int(data["bitrix_id"]))
            product_spesif = ProductSpecification.objects.filter(
                specification=order.specification
            )
            is_need_new_pdf = False
            for order_products_item in data["order_products"]:

                prod = product_spesif.get(
                    product__article=order_products_item["article_motrum"]
                )
                print(prod)
                if order_products_item["date_delivery"]:
                    date_delivery = datetime.datetime.strptime(
                        order_products_item["date_delivery"], "%d-%m-%Y"
                    ).date()
                    if prod.date_delivery_bill != date_delivery:
                        is_need_new_pdf = True

                        prod.date_delivery_bill = date_delivery

                if order_products_item["date_shipment"]:
                    date_shipment = order_products_item["date_shipment"]
                    if date_shipment != "":
                        date_shipment = datetime.datetime.strptime(
                            order_products_item["date_shipment"], "%d-%m-%Y"
                        ).date()

                        prod.date_shipment = date_shipment

                if order_products_item["reserve"] or order_products_item["reserve"] == 0:
                    prod.reserve = int(order_products_item["reserve"])

                if order_products_item["client_shipment"] or order_products_item["client_shipment"] == 0:
                    prod.client_shipment = int(order_products_item["client_shipment"])

                prod.save()

            if is_need_new_pdf:
                if order.requisites.contract:
                    is_req = True
                else:
                    is_req = False

                type_save = request.COOKIES.get("type_save")
                order_pdf = order.create_bill(
                    request,
                    is_req,
                    order,
                    # bill_name,
                    None,
                    None,
                )
                if order_pdf:
                    pdf = request.build_absolute_uri(order.bill_file_no_signature.url)
                    pdf_signed = request.build_absolute_uri(order.bill_file.url)

                    print(order_pdf)
            data_resp = {"result": "ok", "error": None}

            error = "info_error_order"
            location = "OK INFO add_info_order_1c"
            info = f"OK INFO add_info_order_1c{data}"
            e = error_alert(error, location, info)

            return Response(data_resp, status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            tr = traceback.format_exc()
            error = "file_api_error"
            location = "Получение\сохранение данных o товаратах 1с "
            info = f"Получение\сохранение данных o товаратах 1с . Тип ошибки:{e}{tr} DATA из 1с -  {data}"
            e = error_alert(error, location, info)
            data_resp = {"result": "error", "error": f"info-error {info}"}
            return Response(data_resp, status=status.HTTP_400_BAD_REQUEST)

        finally:
            # МЕСТО ДЛЯ ОТПРАВКИ ЭТОЙ ЖЕ ИНФЫ В БИТРИКС
            # если есть изденения даты для переделки счета:

            if pdf:
                if IS_TESTING:
                    pass
                else:
                    is_save_new_doc_bx = save_new_doc_bx(order)

    # ОКТ 1С получение оплат ОКТ Б24
    @authentication_classes([BasicAuthentication])
    @permission_classes([IsAuthenticated])
    @action(detail=False, methods=["post", "put"],authentication_classes =[BasicAuthentication],permission_classes=[IsAuthenticated], url_path=r"add-payment-info-1c")
    def add_payment_order_1c(self, request, *args, **kwargs):
        data = request.data
        

        try:
            data_payment = data["payment"]
            for data_item in data_payment:
                order = Order.objects.get(id_bitrix=int(data_item["bitrix_id"]))
                amount_sum = float(data_item["amount_sum"])
                date_tarnsaction = datetime.datetime.strptime(
                    data_item["date_transaction"], "%d-%m-%Y"
                ).date()
                tarnsaction = PaymentTransaction.objects.create(
                    order=order, date=date_tarnsaction, amount=data_item["amount_sum"]
                )
                order.bill_sum_paid = order.bill_sum_paid + amount_sum
                order.save()

            data_resp = {"result": "ok", "error": None}
            error = "info_error_order"
            location = "add-payment-info-1c"
            info = f"OK INFO add-payment-info-1c  DATA из 1с -  {data}"
            e = error_alert(error, location, info)
            return Response(data_resp, status=status.HTTP_200_OK)

        except Exception as e:
            print(e)
            tr = traceback.format_exc()
            error = "file_api_error"
            location = "Получение\сохранение данных o оплатах 1с "
            info = f"Получение\сохранение данных o оплатах 1с. Тип ошибки:{e}{tr} DATA из 1с -  {data}"
            e = error_alert(error, location, info)
            data_resp = {"result": "error", "error": f"info-error {info}"}

            return Response(data_resp, status=status.HTTP_400_BAD_REQUEST)

        finally:
            if IS_TESTING:
                pass
            else:
                save_payment_order_bx(data)

    # ОКТ 1С получение документов откгрузки ОКТ Б24
    @authentication_classes([BasicAuthentication])
    @permission_classes([IsAuthenticated])
    @action(detail=False, methods=["post"],authentication_classes =[BasicAuthentication],permission_classes=[IsAuthenticated], url_path=r"shipment-info-1c")
    def add_shipment_order_1c(self, request, *args, **kwargs):
        data = request.data
        
        try:
            data_shipment = data["shipment"]
            for data_item in data_shipment:
                order = Order.objects.get(id_bitrix=int(data_item["bitrix_id"]))
                date = datetime.datetime.strptime(data_item["date"], "%d-%m-%Y").date()
                document_shipment = DocumentShipment.objects.create(
                    order=order, date=date
                )
                image_path = save_shipment_doc(data_item["pdf"], document_shipment)

                document_shipment.name = data_item["document_name"]
                document_shipment.file = image_path
                document_shipment.save()
            data_resp = {"result": "ok", "error": None}

            error = "info_error_order"
            location = "shipment-info-1c"
            info = f"OK INFO shipment-info-1c  DATA из 1с -  {data}"
            e = error_alert(error, location, info)

            return Response(data_resp, status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            tr = traceback.format_exc()
            error = "file_api_error"
            location = "ОКТ 1С получение документов откгрузки ОКТ Б24"
            info = f"ОКТ 1С получение документов откгрузки ОКТ Б24 . Тип ошибки:{e}{tr} DATA из 1с -  {data}"
            e = error_alert(error, location, info)
            data_resp = {"result": "error", "error": f"info-error {info}"}
            return Response(data_resp, status=status.HTTP_400_BAD_REQUEST)
        finally:
            if IS_TESTING:
                pass
            else:
                save_shipment_order_bx(data)

# НЕ ИСПОЛЬЗУЮТСЯ
class EmailsViewSet(viewsets.ModelViewSet):
    queryset = EmailsCallBack.objects.none()
    serializer_class = EmailsCallBackSerializer
    http_method_names = [
        "get",
        "post",
        "put",
    ]

    @action(detail=False, methods=["post"], url_path=r"call-back-email")
    def send_email_callback(self, request, *args, **kwargs):
        data = request.data
        serializer = self.serializer_class(data=data, many=False)

        if serializer.is_valid():
            email_obj = serializer.save()
            user_name = serializer.data["name"]
            user_phone = serializer.data["phone"]

            to_manager = os.environ.get("EMAIL_MANAGER_CALLBACK")
            send_code = send_email_message(
                "Обратный звонок",
                f"Имя: {user_name}. Телефон: {user_phone}",
                to_manager,
            )
            if send_code:
                return Response(serializer.data, status=status.HTTP_200_OK)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["post"], url_path=r"manager-email")
    def send_manager_email(self, request, *args, **kwargs):
        data = request.data
        client_id = data["client_id"]
        text_message = data["text_message"]
        client = Client.objects.get(id=int(client_id))
        title_email = "Сообщение с сайта от клиента"
        text_email = f"Клиент: {client.contact_name}Телефон: {client.phone}Сообщение{text_message}"
        to_manager = client.manager.email
        html_message = loader.render_to_string(
            "core/emails/email.html",
            {
                "client_name": client.contact_name,
                "client_phone": client.phone,
                "text": text_message,
            },
        )
        send_code = send_email_message_html(
            title_email, text_email, to_manager, html_message=html_message
        )
        if send_code:
            return Response("ok", status=status.HTTP_200_OK)
        else:
            return Response("error", status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["post"], url_path=r"manager-email")
    def send_email_all(self, request, *args, **kwargs):
        serializer_class = EmailsAllWebSerializer
        data = request.data
        serializer = serializer_class(data=data, many=False)
