import copy
import datetime
from enum import auto
import itertools
from locale import setlocale
import os
from pickle import NONE
from re import T
import re
import traceback

# Импорты для конвертации XLSX в PDF
import pandas as pd
import pdfkit
import tempfile

from apps.client.models import Order, OrderDocumentBill
from apps.core.models import BaseImage, BaseInfo, TypeDelivery
from apps.core.utils import check_spesc_directory_exist, transform_date, rub_words
from PIL import Image
import io
import pathlib
from reportlab.lib.units import mm, cm, inch
import num2words


import reportlab
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from django.conf import settings
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.styles import getSampleStyleSheet

from apps.logs.utils import error_alert
from apps.specification.utils import MyCanvas
from project.settings import IS_TESTING, MEDIA_ROOT, MEDIA_URL, STATIC_ROOT
from django.db.models import Prefetch, OuterRef
from apps.product.models import Product, ProductCart, Stock
from apps.specification.models import ProductSpecification, Specification
from reportlab.lib.fonts import addMapping
from reportlab.lib.enums import TA_JUSTIFY, TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.lib.styles import ParagraphStyle, ListStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, KeepInFrame
from reportlab.platypus import ListFlowable, ListItem
from reportlab.platypus import Frame, PageBreak

# Добавляем импорт для работы с Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image


def crete_pdf_bill(
    specification,
    request,
    is_contract,
    order,
    # bill_name,
    type_delivery,
    post_update,
    type_save,
):

    try:
        print("crete_pdf_bill")
        print(type_save)
        directory = check_spesc_directory_exist(
            "bill",
        )
        specifications = Specification.objects.get(id=specification)
        if order.manager:
            name_admin = f"{order.manager.last_name} {order.manager.first_name}"
            if order.manager.middle_name:
                name_admin = f"{order.manager.last_name} {order.manager.first_name} {order.manager.middle_name}"
        else:

            if specifications.admin_creator:
                name_admin = f"{specifications.admin_creator.last_name} {specifications.admin_creator.first_name}"
                if specifications.admin_creator.middle_name:
                    name_admin = f"{specifications.admin_creator.last_name} {specifications.admin_creator.first_name} {specifications.admin_creator.middle_name}"
            else:
                name_admin = " "

        product_specification = ProductSpecification.objects.filter(
            specification=specification
        ).order_by("id")
        print(type_delivery)
        # type_delivery = TypeDelivery.objects.get(id=type_delivery)

        order = Order.objects.get(specification=specification)
        motrum_info = order.motrum_requisites.requisites
        motrum_info_req = order.motrum_requisites

        client_info = order.requisites
        client_info_req_kpp = order.account_requisites.requisitesKpp
        client_info_req = order.account_requisites

        date_now = transform_date(datetime.date.today().isoformat())
        date_name_dot = datetime.datetime.today().strftime("%d.%m.%Y")

        # if post_update:
        #     date_now = order.bill_date_start.isoformat()
        #     date_now = transform_date(date_now)
        #     date_name_dot = order.bill_date_start.strftime("%d.%m.%Y")
        # else:
        #     date_now = transform_date(datetime.date.today().isoformat())
        #     date_name_dot = datetime.datetime.today().strftime("%d.%m.%Y")

        if order.requisites.contract or specifications.total_amount > 99999.99:
            type_bill = "Счет"
            bill_name = motrum_info.counter_bill + 1
            motrum_info.counter_bill = bill_name
        else:
            type_bill = "Счет-оферта"
            bill_name = motrum_info.counter_bill + 1
            motrum_info.counter_bill = bill_name
            # bill_name = motrum_info.counter_bill_offer + 1
            # motrum_info.counter_bill_offer = bill_name

        print("type_bill", type_bill)
        print("*******************************")
        if type_save == "new":
            name_bill_text = f"{type_bill} № {bill_name}"
            motrum_info.save()
        elif type_save == "update":
            bill_name = order.bill_name
            name_bill_text = f"{type_bill} № {bill_name}"
        elif type_save == "hard_update":
            name_bill_text = f"{type_bill} № {bill_name}"
            motrum_info.save()
        else:
            bill_name = order.bill_name
            name_bill_text = f"{type_bill} № {bill_name}"
        print("name_bill_text", name_bill_text)
        print("*******************************")

        older_doc = OrderDocumentBill.objects.filter(
            order=order,
            bill_name=bill_name,
            bill_file_no_signature="",
            bill_date_start=datetime.datetime.today(),
        )
        if older_doc:
            print(1, older_doc)
            older_doc_ver = older_doc.last()
            print(2, older_doc_ver)
            version = older_doc_ver.version + 1
            print(2, older_doc_ver)
            text_version = f"_{version}"
        else:
            print(2, older_doc)
            version = 1
            text_version = ""
        print(older_doc)
        print(version)
        print(text_version)
        name_bill_to_fullname = f"{name_bill_text} от {date_now}{text_version}"
        name_bill_to_fullname_nosign = (
            f"{name_bill_text} от {date_now} без печати{text_version}"
        )
        name_bill = f"{name_bill_text} от {date_now}{text_version}.pdf"
        name_bill_no_signature = (
            f"{name_bill_text} от {date_now} без печати{text_version}.pdf"
        )
        document_info = BaseImage.objects.filter().first()

        fileName = os.path.join(directory, name_bill)
        fileName_no_sign = os.path.join(directory, name_bill_no_signature)

        story = []
        story_no_sign = []

        pdfmetrics.registerFont(TTFont("Roboto", "Roboto-Regular.ttf", "UTF-8"))
        pdfmetrics.registerFont(TTFont("Roboto-Bold", "Roboto-Bold.ttf", "UTF-8"))

        addMapping("Roboto", 0, 0, "Roboto")  # normal
        addMapping("Roboto-Bold", 1, 0, "Roboto-Bold")  # bold

        doc = SimpleDocTemplate(
            fileName,
            pagesize=A4,
            rightMargin=20,
            leftMargin=20,
            topMargin=40,
            bottomMargin=50,
            title="Счет",
        )
        doc_2 = SimpleDocTemplate(
            fileName_no_sign,
            pagesize=A4,
            rightMargin=20,
            leftMargin=20,
            topMargin=40,
            bottomMargin=50,
            title="Счет",
        )

        styles = getSampleStyleSheet()

        styles.add(ParagraphStyle(name="Roboto", fontName="Roboto", fontSize=7))
        styles.add(ParagraphStyle(name="Roboto-8", fontName="Roboto", fontSize=6))
        styles.add(
            ParagraphStyle(name="Roboto-Bold", fontName="Roboto-Bold", fontSize=7)
        )
        styles.add(
            ParagraphStyle(name="Roboto-Title", fontName="Roboto-Bold", fontSize=12)
        )
        styles.add(
            ParagraphStyle(
                name="Roboto-Bold-Center",
                fontName="Roboto-Bold",
                fontSize=7,
                alignment=TA_CENTER,
            )
        )
        styles.add(
            ParagraphStyle(
                name="Roboto-right",
                fontName="Roboto",
                fontSize=7,
                alignment=TA_RIGHT,
            )
        )
        styles.add(
            ParagraphStyle(
                name="Roboto-left",
                fontName="Roboto",
                fontSize=7,
                alignment=TA_LEFT,
            )
        )
        styles.add(
            ParagraphStyle(
                name="Roboto-centre",
                fontName="Roboto",
                fontSize=7,
                alignment=TA_CENTER,
            )
        )
        styles.add(
            ParagraphStyle(
                name="Roboto-Center-Gray-6",
                fontName="Roboto",
                fontSize=6,
                alignment=TA_CENTER,
            )
        )
        styles.add(
            ParagraphStyle(
                name="Roboto-Center-Gray-6-left",
                fontName="Roboto",
                fontSize=6,
                alignment=TA_LEFT,
            )
        )
        styles.add(
            ParagraphStyle(
                name="Roboto-Center-Gray-6-right",
                fontName="Roboto",
                fontSize=6,
                alignment=TA_RIGHT,
            )
        )
        styles.add(
            ParagraphStyle(
                name="Roboto-Bold-left",
                fontName="Roboto-Bold",
                fontSize=7,
                alignment=TA_RIGHT,
            )
        )

        bold_style = styles["Roboto-Bold"]
        normal_style = styles["Roboto"]
        normal_style_centre = styles["Roboto-centre"]

        normal_style_left = styles["Roboto-left"]
        normal_style_right = styles["Roboto-right"]
        normal_style_6 = styles["Roboto-Center-Gray-6"]
        normal_style_6_right = styles["Roboto-Center-Gray-6-right"]
        bold_style_center = styles["Roboto-Bold-Center"]
        normal_style_8 = styles["Roboto-8"]
        title_style_14 = styles["Roboto-Title"]
        bold_left_style = styles["Roboto-Bold-left"]

        doc_title = copy.copy(styles["Heading1"])
        doc_title.fontName = "Roboto-Bold"
        doc_title.fontSize = 7

        # name_image_logo = f"{MEDIA_ROOT}/documents/logo.png"
        name_image_logo = request.build_absolute_uri(document_info.logo.url)
        logo_motrum = Paragraph(
            f'<img width="155" height="35"  src="{name_image_logo}" />',
            normal_style,
        )

        story.append(logo_motrum)
        story_no_sign.append(logo_motrum)
        # тут вставки бик корпоратив
        data_bank = [
            (
                Paragraph(
                    f"{motrum_info_req.bank}",
                    normal_style,
                ),
                Paragraph("БИК", normal_style),
                Paragraph(f"{motrum_info_req.bic}", normal_style),
            ),
            (
                None,
                Paragraph("Сч. №", normal_style),
                Paragraph(f"{motrum_info_req.kpp}", normal_style),
            ),
            (
                Paragraph("Банк получателя", normal_style_8),
                None,
                None,
            ),
            (
                Paragraph(
                    f"ИНН {motrum_info.inn} &nbsp &nbsp &nbsp  KПП {motrum_info.kpp} ",
                    normal_style,
                ),
                Paragraph("Сч. №", normal_style),
                Paragraph(f"{motrum_info_req.account_requisites}", normal_style),
            ),
            (
                Paragraph(f"{motrum_info.full_name_legal_entity}", normal_style),
                None,
                None,
            ),
            (
                Paragraph("Получатель", normal_style_8),
                None,
                None,
            ),
        ]
        table_bank = Table(
            data_bank,
            colWidths=[
                9 * cm,
                2 * cm,
                7 * cm,
            ],
            rowHeights=0.5 * cm,
            hAlign="LEFT",
        )
        table_bank.setStyle(
            TableStyle(
                [
                    ("ALIGN", (0, 0), (0, -1), "RIGHT"),
                    ("BOX", (0, 0), (-1, -1), 0.25, colors.black),
                    ("BOX", (0, -3), (-1, -1), 0.25, colors.black),
                    ("BOX", (0, -3), (-3, -3), 0.25, colors.black),
                    ("BOX", (1, 0), (-1, -1), 0.25, colors.black),
                    ("BOX", (2, 0), (-1, -1), 0.25, colors.black),
                    # ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.black),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    # # ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
                    # ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.black),
                    # ("BOX", (0, 0), (-1, -1), 2, colors.black),
                    # ("BOX", (0, 0), (-1, 1), 2, colors.black),
                    # ("BOX", (-1, 1), (0, 0), 2, colors.red),
                    # # ('BOX',(-2,-1),(-1,1),2,colors.red),
                    # ("BOX", (-1, 0), (-1, -1), 2, colors.black),
                    # # ('LINEABOVE',(0,2),(-1,2),2,colors.black)
                    # # ('LINEABOVE',(0,-2),(-1,2),2,colors.black)
                ]
            )
        )
        story.append(table_bank)
        story_no_sign.append(table_bank)
        # name_image_logo_supplier = f"{MEDIA_ROOT}/documents/supplier.png"
        name_image_logo_supplier = request.build_absolute_uri(document_info.vendors.url)

        logo_supplier = Paragraph(
            f'<br></br><br></br><br></br><br></br><br></br><br></br><img width="555" height="55"  src="{name_image_logo_supplier}" /><br></br>',
            normal_style,
        )
        story.append(logo_supplier)
        story_no_sign.append(logo_supplier)
        if is_contract or specifications.total_amount > 99999.99:
            story.append(
                Paragraph(
                    f"Счет на оплату № {bill_name} от {date_now}<br></br><br></br>",
                    title_style_14,
                )
            )
            story_no_sign.append(
                Paragraph(
                    f"Счет на оплату № {bill_name} от {date_now}<br></br><br></br>",
                    title_style_14,
                )
            )
        else:
            story.append(
                Paragraph(
                    f"Счет-оферта № {bill_name} от {date_now}<br></br><br></br>",
                    title_style_14,
                )
            )
            story_no_sign.append(
                Paragraph(
                    f"Счет-оферта № {bill_name} от {date_now}<br></br><br></br>",
                    title_style_14,
                )
            )

        data_info = []
        data_info.append(
            (
                Paragraph(
                    f'<br></br>Поставщик<br></br><font  size="6">(исполнитель):</font>',
                    normal_style,
                ),
                Paragraph(
                    f"{motrum_info.full_name_legal_entity}, ИНН {motrum_info.inn}, КПП {motrum_info.kpp}, {motrum_info.legal_post_code}, {motrum_info.legal_city}, {motrum_info.legal_address}, тел.: {motrum_info.tel}<br></br>",
                    bold_style,
                ),
            )
        )

        if client_info.type_client == 1 or client_info.type_client == "1":
            # клиент юрлицо
            info_client = f"{client_info.legal_entity}, ИНН {client_info.inn}, КПП {client_info_req_kpp.kpp}, {client_info_req_kpp.legal_post_code}, {client_info_req_kpp.legal_city} {client_info_req_kpp.legal_address}"

        else:
            # клиент ип
            info_client = f"{client_info.legal_entity}, ИНН {client_info.inn}, ОГРНИП {client_info_req_kpp.ogrn}, {client_info_req_kpp.legal_post_code}, {client_info_req_kpp.legal_city} {client_info_req_kpp.legal_address}"

        if client_info_req_kpp.tel and client_info_req_kpp.tel != "":
            info_client = f"{info_client},  тел.: {client_info_req_kpp.tel}"

        data_info.append(
            (
                Paragraph(
                    f'Покупатель<br></br><font  size="6">(заказчик):</font>',
                    normal_style,
                ),
                Paragraph(
                    info_client,
                    bold_style,
                ),
            )
        )

        data_info.append(
            (
                Paragraph(f"Основание:", normal_style),
                Paragraph(f"Счет {bill_name} от {date_name_dot}", bold_style),
            )
        )

        table_info = Table(data_info, colWidths=[2.5 * cm, 14 * cm], hAlign="LEFT")
        table_info.setStyle(
            TableStyle(
                [
                    ("LINEABOVE", (0, 0), (-1, 0), 2, colors.black),
                    ("FONT", (0, 0), (-1, -1), "Roboto", 7),
                    ("ALIGN", (0, 0), (0, -1), "RIGHT"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.transparent),
                    # ("TOPPADDING", (0, 0), (-1, 0), 1, )
                ]
            )
        )
        story.append(table_info)
        story_no_sign.append(table_info)
        data = [
            (
                Paragraph("№ ", bold_style_center),
                Paragraph("Товар (Услуга)", bold_style_center),
                Paragraph("Код", bold_style_center),
                Paragraph("Кол-во", bold_style_center),
                Paragraph("Ед.", bold_style_center),
                Paragraph("Цена", bold_style_center),
                Paragraph("Сумма", bold_style_center),
                Paragraph("Срок поставки", bold_style_center),
            ),
            (
                Paragraph(f"1", normal_style_6),
                Paragraph(f"2", normal_style_6),
                Paragraph(f"3", normal_style_6),
                Paragraph(f"4", normal_style_6),
                Paragraph(f"5", normal_style_6),
                Paragraph(f"6", normal_style_6),
                Paragraph(f"7", normal_style_6),
                Paragraph(f"8", normal_style_6),
            ),
        ]

        i = 0
        date_ship = datetime.date.today()
        is_none_date_delivery = False
        total_product_quantity = 0
        for product in product_specification:
            i += 1
            print(i, product)
            try:
                product_stock_item = Stock.objects.get(prod=product.product)
                product_stock = product_stock_item.lot.name_shorts
            except Stock.DoesNotExist:
                product_stock = "шт"

            if product.product:
                if IS_TESTING:
                    link = product.product.get_url_document_test()
                else:
                    link = product.product.get_url_document()

                url_absolute = request.build_absolute_uri("/").strip("/")
                link = f"{url_absolute}/{link}"
                product_name_str = product.product.name
                
                if product.product.supplier.slug == "prompower" and product.product.description:
                    product_name_str = product.product.description
                    # if product.product.name not in product.product.description:
                    #     product_name_str = f"{product.product.name} {product_name_str}"
                
                if product.product.in_view_website:
                    product_name = (
                        f'<a href="{link}" color="blue">{str(product_name_str)}</a>'
                    )
                else:
                    product_name = f"{str(product_name_str)}"

                product_name = (Paragraph(product_name, normal_style_left),)
                product_code = product.product.article_supplier
                product_code = (Paragraph(product_code, normal_style_left),)

            else:
                product_name = product.product_new
                product_name = (Paragraph(product_name, normal_style_left),)
                product_code = product.product_new_article
                product_code = (Paragraph(product_code, normal_style_left),)

            product_price = product.price_one
            product_price = (
                "{0:,.2f}".format(product_price).replace(",", " ").replace(".", ",")
            )
            product_price_all = product.price_all
            product_price_all = (
                "{0:,.2f}".format(product_price_all).replace(",", " ").replace(".", ",")
            )
            product_quantity = product.quantity
            product_data = product.text_delivery

            if product.date_delivery_bill:
                product_data = str(product.date_delivery_bill.strftime("%d.%m.%Y"))
            else:
                product_data = product.text_delivery

            product_data = (Paragraph(f"{product_data}", normal_style_right),)
            # else:
            #     product_data = str("-")
            total_product_quantity += product_quantity
            data.append(
                (
                    Paragraph(f"{str(i)}", normal_style_centre),
                    product_name,
                    product_code,
                    Paragraph(f"{str(product_quantity)}", normal_style_right),
                    # product_quantity,
                    Paragraph(f"{str(product_stock)}.", normal_style_left),
                    # product_stock,
                    Paragraph(f"{product_price}", normal_style_right),
                    # product_price,
                    Paragraph(f"{product_price_all}", normal_style_right),
                    # product_price_all,
                    product_data,
                )
            )
        total_amount_str = (
            "{0:,.2f}".format(specifications.total_amount)
            .replace(",", " ")
            .replace(".", ",")
        )

        data.append(
            (
                None,
                None,
                None,
                Paragraph(f"{str(total_product_quantity)}", normal_style_right),
                # total_product_quantity,
                None,
                None,
                Paragraph(f"{str(total_amount_str)}", normal_style_right),
                # total_amount_str,
                None,
            )
        )

        table_product = Table(
            data,
            colWidths=[
                1 * cm,
                6 * cm,
                3 * cm,
                1.5 * cm,
                1 * cm,
                2.5 * cm,
                2.5 * cm,
                2 * cm,
            ],
            repeatRows=2,
        )
        table_product.setStyle(
            TableStyle(
                [
                    ("FONT", (0, 0), (-1, -1), "Roboto", 7),
                    ("ALIGN", (0, 0), (0, -1), "RIGHT"),
                    # ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
                    ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.black),
                    ("BOX", (0, 0), (-1, -1), 2, colors.black),
                    ("BOX", (0, 0), (-1, 1), 2, colors.black),
                    ("BOX", (-1, 1), (0, 0), 2, colors.red),
                    # ('BOX',(-2,-1),(-1,1),2,colors.red),
                    ("BOX", (-1, 0), (-1, -1), 2, colors.black),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    # ('LINEABOVE',(0,2),(-1,2),2,colors.black)
                    # ('LINEABOVE',(0,-2),(-1,2),2,colors.black)
                ]
            )
        )
        story.append(table_product)
        story_no_sign.append(table_product)
        if order.prepay_persent:
            if order.prepay_persent == 100:
                info_payment = f" Способ оплаты: 100% предоплата."

            else:
                info_payment = f"{order.prepay_persent}% предоплата, {order.postpay_persent}% в течение 5 дней с момента отгрузки со склада Поставщика."

        else:
            info_payment = ""

        total_amount_nds = float(specifications.total_amount) * 20 / (20 + 100)
        total_amount_nds = round(total_amount_nds, 2)

        total_amount = (
            "{0:,.2f}".format(specifications.total_amount)
            .replace(",", " ")
            .replace(".", ",")
        )
        total_amount_nds = (
            "{0:,.2f}".format(total_amount_nds).replace(",", " ").replace(".", ",")
        )
        final_table_all = []
        final_table_all.append(
            (
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
            )
        )
        final_table_all.append(
            (
                Paragraph(info_payment, normal_style),
                None,
                None,
                None,
                Paragraph("Итого:", bold_left_style),
                None,
                Paragraph(f"{total_amount}", bold_left_style),
                None,
            )
        )

        final_table_all.append(
            (
                None,
                None,
                None,
                None,
                Paragraph("В том числе  НДС:", bold_left_style),
                None,
                Paragraph(f"{total_amount_nds} ", bold_left_style),
                None,
            )
        )

        final_table_all.append(
            (
                None,
                None,
                None,
                None,
                Paragraph("Всего к оплате:", bold_left_style),
                None,
                Paragraph(f"{total_amount}<br></br>", bold_left_style),
                None,
            )
        )

        final_table_all_prod = Table(
            final_table_all,
            colWidths=[
                6 * cm,
                1 * cm,
                2.5 * cm,
                0 * cm,
                3 * cm,
                2.5 * cm,
                2.5 * cm,
                2 * cm,
            ],
            rowHeights=13,
        )
        story.append(final_table_all_prod)
        story_no_sign.append(final_table_all_prod)

        total_amount_word = num2words.num2words(
            int(specifications.total_amount), lang="ru"
        ).capitalize()
        total_amount_pens = str(specifications.total_amount).split(".")
        total_amount_pens = total_amount_pens[1]

        if len(total_amount_pens) < 2:
            total_amount_pens = f"{total_amount_pens}0"
        rub_word = rub_words(int(specifications.total_amount))
        data_text_info = []
        # data_text_info = [
        #     (
        #         Paragraph(
        #             f"Всего наименований {i}, на сумму {total_amount_str} руб.",
        #             normal_style,
        #         ),
        #     ),
        #     (
        #         Paragraph(
        #             f"{total_amount_word} {rub_word} {total_amount_pens} копеек",
        #             bold_style,
        #         ),
        #     ),
        # ]
        data_text_info.append(
            (
                Paragraph(
                    f"Всего наименований {i}, на сумму {total_amount_str} руб.",
                    normal_style,
                ),
            ),
        )
        data_text_info.append(
            (
                Paragraph(
                    f"{total_amount_word} {rub_word} {total_amount_pens} копеек",
                    bold_style,
                ),
            ),
        )

        if is_contract or specifications.total_amount > 99999.99:
            data_text_info.append(
                (
                    Paragraph(
                        f"<br></br><br></br>Оплата данного счета означает согласие с условиями поставки товара.",
                        normal_style,
                    ),
                )
            )
            data_text_info.append(
                (
                    Paragraph(
                        f"Уведомление об оплате обязательно, в противном случае не гарантируется наличие товара на складе.",
                        normal_style,
                    ),
                )
            )

            if type_delivery.company_delivery is None:

                data_text_info.append(
                    (
                        Paragraph(
                            f"Товар отпускается по факту прихода денег на р/с Поставщика, самовывозом, при наличии доверенности и паспорта.",
                            normal_style,
                        ),
                    )
                )

            else:

                data_text_info.append(
                    (
                        Paragraph(
                            f"Условия доставки: {type_delivery.text_long}",
                            normal_style,
                        ),
                    )
                )

        else:
            data_text_info.append(
                (
                    Paragraph(
                        f"<br></br><br></br>1. Оплата данного счет-оферты означает полное и безоговорочное согласие (акцепт) с условиями поставки товара по наименованию, ассортименту, количеству и цене. Срок действия счета 3 банковских дня.",
                        normal_style,
                    ),
                )
            )
            data_text_info.append(
                (
                    Paragraph(
                        f"2. Поставщик гарантирует отгрузку товара по ценам и в сроки, указанные в настоящем Счет-оферте, при условии зачисления денежных средств на расчетный счет Поставщика в течение 3 банковских дней с даты выставления счета.<br></br>При невыполнении Покупателем указанных условий оплаты, цена и сроки поставки товара могут измениться.",
                        normal_style,
                    ),
                )
            )
            data_text_info.append(
                (
                    Paragraph(
                        f"3. Товар отгружается после полной оплаты счета-оферты Покупателем.",
                        normal_style,
                    ),
                )
            )
            data_text_info.append(
                (
                    Paragraph(
                        f"5. Обязательства Поставщика по поставке Товара считаются выполненными с момента подписания УПД или товарной накладной представителями Поставщика и Покупателя или организации перевозчика.",
                        normal_style,
                    ),
                )
            )
            data_text_info.append(
                (
                    Paragraph(
                        f"6. Каждая партия поставляемой продукции сопровождается Универсальным передаточным документом (УПД): на бумажном носителе в двух экземплярах (1 экз. Покупателя, 1 экз. Поставщика). После каждой поставки продукции с документами, Покупатель обязан вернуть Поставщику один экземпляр, верно оформленного УПД, не позднее 1 месяца с даты подтверждения получения товара. В случае задержки Покупателем возврата, верно оформленного со стороны Покупателя, оригинала УПД на бумажном носителе на срок более 1 (одного) календарного месяца, Поставщик вправе предъявить Покупателю штрафные санкции в размере 5 000 (Пять тысяч) рублей (НДС не облагается) за каждый факт не предоставления подписанного оригинала УПД.",
                        normal_style,
                    ),
                )
            )

            if type_delivery.company_delivery is None:
                data_text_info.append(
                    (
                        Paragraph(
                            f"7. Доставка товара самовывозом со склада Поставщика",
                            normal_style,
                        ),
                    )
                )
            else:

                data_text_info.append(
                    (
                        Paragraph(
                            f"7.{type_delivery.text_long}",
                            normal_style,
                        ),
                    )
                )

            data_text_info.append(
                (
                    Paragraph(
                        f"8. В случае превышения более чем на 15 дней сроков поставки продукции, указанных в Счёте-оферте, Поставщик по требованию Покупателя обязан уплатить неустойку в размере 0,1 % от стоимости не поставленной в срок продукции за каждый день просрочки, но не более 5% от стоимости не поставленной в срок продукции.",
                        normal_style,
                    ),
                )
            )
            data_text_info.append(
                (
                    Paragraph(
                        f"9. Претензии по п. 8 должны быть заявлены Сторонами в письменной форме в течение 5 дней с момента наступления, указанных в них событий. В случае не выставления претензии в указанный срок, это трактуется как освобождение Сторон от уплаты неустойки.",
                        normal_style,
                    ),
                )
            )
            data_text_info.append(
                (
                    Paragraph(
                        f"10. Срок гарантии на поставляемую продукцию составляет не менее одного года с момента отгрузки.",
                        normal_style,
                    ),
                )
            )
            data_text_info.append(
                (
                    Paragraph(
                        f"11. Правила гарантийного обслуживания оговариваются в гарантийных талонах на поставляемую продукцию.",
                        normal_style,
                    ),
                )
            )
            data_text_info.append(
                (
                    Paragraph(
                        f"12.  Стороны принимают необходимые меры к тому, чтобы спорные вопросы и разногласия, возникающие при исполнении и расторжении настоящего договора, были урегулированы путем переговоров. В случае если стороны не достигнут соглашения по спорным вопросам путем переговоров, то спор передается заинтересованной стороной в арбитражный суд.",
                        normal_style,
                    ),
                )
            )

        table_data_text_info = Table(data_text_info, splitInRow=1, hAlign="LEFT")
        table_data_text_info2 = Table(data_text_info, splitInRow=1, hAlign="LEFT")

        table_data_text_info.setStyle(
            TableStyle(
                [
                    ("LINEABOVE", (0, 0), (-1, 0), 2, colors.transparent),
                    ("FONT", (0, 0), (-1, -1), "Roboto", 7),
                    # ("ALIGN", (0, 0), (0, -1), "LEFT"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.transparent),
                ]
            )
        )
        table_data_text_info2.setStyle(
            TableStyle(
                [
                    ("LINEABOVE", (0, 0), (-1, 0), 2, colors.transparent),
                    ("FONT", (0, 0), (-1, -1), "Roboto", 7),
                    # ("ALIGN", (0, 0), (0, -1), "LEFT"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.transparent),
                ]
            )
        )

        print(table_data_text_info)
        # story_no_sign = story.copy()
        story.append(table_data_text_info)
        story_no_sign.append(table_data_text_info2)

        name_image = request.build_absolute_uri(motrum_info.signature.url)
        signature_motrum = Paragraph(
            f'<br /><img width="100" height="30" src="{name_image}" valign="middle"/>',
            normal_style,
        )
        signature_motrum_name = Paragraph(f"Старостина В. П.", normal_style)

        name_image_press = request.build_absolute_uri(motrum_info.stamp.url)

        press_motrum = Paragraph(
            f'<br /><br /><br /><br /><br />М.П.<img width="100" height="100" src="{name_image_press}" valign="middle"/>',
            normal_style,
        )

        data_signature = [
            (
                Paragraph("Руководитель:", bold_style),
                signature_motrum,
                signature_motrum_name,
            ),
            (
                Paragraph("Бухгалтер:", bold_style),
                signature_motrum,
                signature_motrum_name,
            ),
            (
                None,
                press_motrum,
                None,
            ),
            (
                None,
                None,
                Paragraph(
                    f"<br></br><br></br><br></br>{name_admin}", normal_style_right
                ),
            ),
        ]

        table_signature = Table(
            data_signature, colWidths=[3 * cm, 5 * cm, 7 * cm, 7 * cm], hAlign="LEFT"
        )
        table_signature.setStyle(
            TableStyle(
                [
                    ("LINEABOVE", (0, 0), (-1, 0), 2, colors.black),
                    ("FONT", (0, 0), (-1, -1), "Roboto", 7),
                    ("ALIGN", (0, 0), (0, -1), "RIGHT"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.transparent),
                ]
            )
        )

        # Получаем высоту таблицы
        available_width = doc.width - doc.topMargin - doc.rightMargin
        available_height = doc.height - doc.bottomMargin - doc.topMargin
        table_height = table_signature.wrap(available_width, available_height)[1]

        # Считаем общую кумулятивную высоту всех элементов в story
        cumulative_height_of_story = 0
        for element in story:
            if hasattr(element, "wrap"):
                # Предполагается, что doc.height здесь - это максимальная высота,
                # которую элемент МОГ БЫ занять. wrap вернет реальную высоту элемента.
                _w, h = element.wrap(available_width, available_height)
                cumulative_height_of_story += h

        # Теперь вычисляем, какая часть последней страницы занята этим cumulative_height_of_story
        if cumulative_height_of_story == 0:
            height_used_on_final_page_of_story = 0
        else:
            # Если кумулятивная высота точно равна N страницам, то последняя страница полная.
            if (
                cumulative_height_of_story % available_height == 0
                and cumulative_height_of_story > 0
            ):
                height_used_on_final_page_of_story = available_height
            else:
                # Иначе это остаток на последней странице.
                height_used_on_final_page_of_story = (
                    cumulative_height_of_story % available_height
                )

        # total_height теперь представляет высоту, использованную на ПОСЛЕДНЕЙ странице,
        # где заканчивается текущий 'story'.
        total_height = height_used_on_final_page_of_story

        # Оставшееся место = высота страницы - использованная высота на этой последней странице
        remaining_height = available_height - total_height

        table_height_3 = table_height / 3
        table_height_23 = table_height_3 * 2
        print("remaining_height", remaining_height)
        print("table_height_23", table_height_23)
        # Если таблица занимает меньше половины оставшегося места
        if table_height_23 > remaining_height:
            print("66666666666666666666666666666666666666666666666666")
            signature_motrum = Paragraph(
                f'<br /><img width="95" height="25" src="{name_image}" valign="middle"/>',
                normal_style,
            )
            press_motrum = Paragraph(
                f'<br /><br /><br /><br />М.П.<img width="90" height="90" src="{name_image_press}" valign="middle"/>',
                normal_style,
            )

            data_signature = [
                (
                    Paragraph("Руководитель:", bold_style),
                    signature_motrum,
                    signature_motrum_name,
                ),
                (
                    Paragraph("Бухгалтер:", bold_style),
                    signature_motrum,
                    signature_motrum_name,
                ),
                (
                    None,
                    press_motrum,
                    Paragraph(f"{name_admin}", normal_style_right),
                ),
            ]

            table_signature = Table(
                data_signature,
                colWidths=[3 * cm, 5 * cm, 7 * cm, 7 * cm],
                hAlign="LEFT",
            )
            # Уменьшаем отступы в таблице
            table_signature.setStyle(
                TableStyle(
                    [
                        ("LINEABOVE", (0, 0), (-1, 0), 2, colors.black),
                        ("FONT", (0, 0), (-1, -1), "Roboto", 7),
                        ("ALIGN", (0, 0), (0, -1), "RIGHT"),
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.transparent),
                        ("TOPPADDING", (0, 0), (-1, -1), 0),  # Убираем верхний отступ
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),  # Убираем нижний отступ
                    ]
                )
            )

            # Уменьшаем отступы в параграфах
            for row in data_signature:
                for cell in row:
                    if isinstance(cell, Paragraph):
                        cell.style.spaceBefore = 0
                        cell.style.spaceAfter = 0

        story.append(table_signature)

        # Создаем PDF
        pdf = doc
        pdf = pdf.build(story, canvasmaker=MyCanvas)

        file_path = "{0}/{1}/{2}".format(
            "documents",
            "bill",
            name_bill,
        )
        print(file_path)

        print(333333333333333333)

        data_signature = [
            (
                Paragraph(f"<br /><br />Руководитель:", bold_style),
                Paragraph(f"________________________________", normal_style_centre),
                Paragraph(
                    f"__________________________________________________________",
                    normal_style_centre,
                ),
            ),
            (
                None,
                Paragraph(f"   подпись           ", normal_style_centre),
                Paragraph(f"       расшифровка подписи      ", normal_style_centre),
            ),
            (
                Paragraph("Бухгалтер:", bold_style),
                Paragraph(f"________________________________", normal_style_centre),
                Paragraph(
                    f"__________________________________________________________",
                    normal_style_centre,
                ),
            ),
            (
                None,
                Paragraph(f"   подпись           ", normal_style_centre),
                Paragraph(f"       расшифровка подписи      ", normal_style_centre),
            ),
            (
                Paragraph(f"<br /><br /><br /><br />МП.", normal_style_right),
                None,
                None,
            ),
            (
                None,
                None,
                Paragraph(
                    f"<br></br><br></br><br></br>{name_admin}", normal_style_right
                ),
            ),
        ]

        table_signature = Table(
            data_signature, colWidths=[3 * cm, 5 * cm, 7 * cm, 7 * cm], hAlign="LEFT"
        )
        table_signature.setStyle(
            TableStyle(
                [
                    ("LINEABOVE", (0, 0), (-1, 0), 2, colors.black),
                    ("FONT", (0, 0), (-1, -1), "Roboto", 7),
                    ("ALIGN", (0, 0), (0, -1), "RIGHT"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.transparent),
                ]
            )
        )

        
        # Получаем высоту таблицы
        available_width = doc_2.width - doc_2.topMargin - doc_2.rightMargin
        available_height = doc_2.height - doc_2.bottomMargin - doc_2.topMargin
        table_height = table_signature.wrap(available_width, available_height)[1]

        # Считаем общую кумулятивную высоту всех элементов в story
        cumulative_height_of_story = 0
        for element in story_no_sign:
            if hasattr(element, "wrap"):
                # Предполагается, что doc.height здесь - это максимальная высота,
                # которую элемент МОГ БЫ занять. wrap вернет реальную высоту элемента.
                _w, h = element.wrap(available_width, available_height)
                cumulative_height_of_story += h

        # Теперь вычисляем, какая часть последней страницы занята этим cumulative_height_of_story
        if cumulative_height_of_story == 0:
            height_used_on_final_page_of_story = 0
        else:
            # Если кумулятивная высота точно равна N страницам, то последняя страница полная.
            if (
                cumulative_height_of_story % available_height == 0
                and cumulative_height_of_story > 0
            ):
                height_used_on_final_page_of_story = available_height
            else:
                # Иначе это остаток на последней странице.
                height_used_on_final_page_of_story = (
                    cumulative_height_of_story % available_height
                )

        # total_height теперь представляет высоту, использованную на ПОСЛЕДНЕЙ странице,
        # где заканчивается текущий 'story'.
        total_height = height_used_on_final_page_of_story

        # Оставшееся место = высота страницы - использованная высота на этой последней странице
        remaining_height = available_height - total_height

        table_height_3 = table_height / 3
        table_height_23 = table_height_3 * 2
        print("remaining_height", remaining_height)
        print("table_height_23", table_height_23)
        # Если таблица занимает меньше половины оставшегося места
        if table_height_23 > remaining_height:
            print("66666666666666666666666666666666666666666666666666")
            signature_motrum = Paragraph(
                f'<br /><img width="95" height="25" src="{name_image}" valign="middle"/>',
                normal_style,
            )
            press_motrum = Paragraph(
                f'<br /><br /><br /><br />М.П.<img width="90" height="90" src="{name_image_press}" valign="middle"/>',
                normal_style,
            )

            data_signature = [
            (
                Paragraph(f"<br />Руководитель:", bold_style),
                Paragraph(f"________________________________", normal_style_centre),
                Paragraph(
                    f"__________________________________________________________",
                    normal_style_centre,
                ),
            ),
            (
                None,
                Paragraph(f"   подпись           ", normal_style_centre),
                Paragraph(f"       расшифровка подписи      ", normal_style_centre),
            ),
            (
                Paragraph("Бухгалтер:", bold_style),
                Paragraph(f"________________________________", normal_style_centre),
                Paragraph(
                    f"__________________________________________________________",
                    normal_style_centre,
                ),
            ),
            (
                None,
                Paragraph(f"   подпись           ", normal_style_centre),
                Paragraph(f"       расшифровка подписи      ", normal_style_centre),
            ),
            (
                Paragraph(f"<br /><br />МП.", normal_style_right),
                None,
                Paragraph(
                    f"{name_admin}", normal_style_right
                ),
            ),
            
        ]

            table_signature = Table(
                data_signature,
                colWidths=[3 * cm, 5 * cm, 7 * cm, 7 * cm],
                hAlign="LEFT",
            )
            # Уменьшаем отступы в таблице
            table_signature.setStyle(
                TableStyle(
                    [
                        ("LINEABOVE", (0, 0), (-1, 0), 2, colors.black),
                        ("FONT", (0, 0), (-1, -1), "Roboto", 7),
                        ("ALIGN", (0, 0), (0, -1), "RIGHT"),
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.transparent),
                        ("TOPPADDING", (0, 0), (-1, -1), 0),  # Убираем верхний отступ
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),  # Убираем нижний отступ
                    ]
                )
            )

            # Уменьшаем отступы в параграфах
            for row in data_signature:
                for cell in row:
                    if isinstance(cell, Paragraph):
                        cell.style.spaceBefore = 0
                        cell.style.spaceAfter = 0
        
        
        

        story_no_sign.append(table_signature)

        # Создаем PDF
        pdf_no_sign = doc_2
        pdf_no_sign = pdf_no_sign.build(story_no_sign, canvasmaker=MyCanvas)

        file_path_no_sign = "{0}/{1}/{2}".format(
            "documents",
            "bill",
            name_bill_no_signature,
        )

        return (
            file_path,
            bill_name,
            file_path_no_sign,
            version,
            name_bill_to_fullname,
            name_bill_to_fullname_nosign,
        )

    except Exception as e:
        tr = traceback.format_exc()
        error = "error"
        location = "Сохранение пдф счета "
        info = f"Сохранение пдф счета  ошибка {e}{tr}"
        e = error_alert(error, location, info)


def create_xlsx_bill(
    specification,
    request,
    is_contract,
    order,
    # bill_name,
    type_delivery,
    post_update,
    type_save,
):
    try:
        print("create_xlsx_bill")
        print(type_save)
        
        # Создаем директорию для сохранения файла
        directory = check_spesc_directory_exist("bill")
        
        base_image = BaseImage.objects.filter().first()
        name_image_logo = base_image.logo
        name_image_vendor = base_image.vendors
        
        
        
        
        specifications = Specification.objects.get(id=specification)
        if order.manager:
            name_admin = f"{order.manager.last_name} {order.manager.first_name}"
            if order.manager.middle_name:
                name_admin = f"{order.manager.last_name} {order.manager.first_name} {order.manager.middle_name}"
        else:
            if specifications.admin_creator:
                name_admin = f"{specifications.admin_creator.last_name} {specifications.admin_creator.first_name}"
                if specifications.admin_creator.middle_name:
                    name_admin = f"{specifications.admin_creator.last_name} {specifications.admin_creator.first_name} {specifications.admin_creator.middle_name}"
            else:
                name_admin = " "

        product_specification = ProductSpecification.objects.filter(
            specification=specification
        ).order_by("id")
        print(type_delivery)

        order = Order.objects.get(specification=specification)
        motrum_info = order.motrum_requisites.requisites
        motrum_info_req = order.motrum_requisites
        name_image_motrum_stamp = motrum_info.stamp
        name_image_motrum_signatures = motrum_info.signature
        
        client_info = order.requisites
        client_info_req_kpp = order.account_requisites.requisitesKpp
        client_info_req = order.account_requisites

        date_now = transform_date(datetime.date.today().isoformat())
        date_name_dot = datetime.datetime.today().strftime("%d.%m.%Y")

        if order.requisites.contract or specifications.total_amount > 99999.99:
            type_bill = "Счет"
            bill_name = motrum_info.counter_bill + 1
            motrum_info.counter_bill = bill_name
        else:
            type_bill = "Счет-оферта"
            bill_name = motrum_info.counter_bill + 1
            motrum_info.counter_bill = bill_name

        print("type_bill", type_bill)
        print("*******************************")
        if type_save == "new":
            name_bill_text = f"{type_bill} № {bill_name}"
            motrum_info.save()
        elif type_save == "update":
            bill_name = order.bill_name
            name_bill_text = f"{type_bill} № {bill_name}"
        elif type_save == "hard_update":
            name_bill_text = f"{type_bill} № {bill_name}"
            motrum_info.save()
        else:
            bill_name = order.bill_name
            name_bill_text = f"{type_bill} № {bill_name}"
        print("name_bill_text", name_bill_text)
        print("*******************************")

        older_doc = OrderDocumentBill.objects.filter(
            order=order,
            bill_name=bill_name,
            bill_file_no_signature="",
            bill_date_start=datetime.datetime.today(),
        )
        if older_doc:
            print(1, older_doc)
            older_doc_ver = older_doc.last()
            print(2, older_doc_ver)
            version = older_doc_ver.version + 1
            print(2, older_doc_ver)
            text_version = f"_{version}"
        else:
            print(2, older_doc)
            version = 1
            text_version = ""
        print(older_doc)
        print(version)
        print(text_version)
        name_bill_to_fullname = f"{name_bill_text} от {date_now}{text_version}"
        name_bill = f"{name_bill_text} от {date_now}{text_version}.xlsx"

        # Создаем новую книгу Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Лист_1"

        # Настройка ширины столбцов для соответствия шаблону
        column_widths = {
            'A': 2, 'B': 3, 'C': 2, 'D': 45, 'E': 2, 'F': 2, 'G': 2, 'H': 2, 'I': 2, 'J': 2, 'K': 2, 'L': 2, 'M': 2, 'N': 2, 'O': 2, 'P': 2, 'Q': 2, 'R': 2, 'S': 2, 'T': 2, 'U': 2, 'V': 2, 'W': 2, 'X': 2, 'Y': 2, 'Z': 2,
            'AA': 15, 'AB': 2, 'AC': 2, 'AD': 2, 'AE': 2, 'AF': 2, 'AG': 2, 'AH': 2, 'AI': 2, 'AJ': 2, 'AK': 2, 'AL': 2, 'AM': 2,
            'AN': 6, 'AO': 2, 'AP': 2, 'AQ': 2, 'AR': 2, 'AS': 2,
            'AT': 6, 'AU': 2, 'AV': 2, 'AW': 2, 'AX': 2,
            'AY': 12, 'AZ': 2, 'BA': 2, 'BB': 2, 'BC': 2, 'BD': 2, 'BE': 2, 'BF': 2,
            'BG': 15, 'BH': 2, 'BI': 2, 'BJ': 2,
            'BK': 15, 'BL': 2, 'BM': 2, 'BN': 2, 'BO': 2
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Настройка стилей
        header_font = Font(name='Arial', size=12, bold=True)
        normal_font = Font(name='Arial', size=10)
        small_font = Font(name='Arial', size=8)
        title_font = Font(name='Arial', size=14, bold=True)
        bold_font = Font(name='Arial', size=10, bold=True)
        bank_font = Font(name='Arial', size=9)  # Размер 9.0 как в шаблоне
        bank_small_font = Font(name='Arial', size=8)  # Размер 8.0 как в шаблоне
        
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        right_alignment = Alignment(horizontal='right', vertical='center')
        top_alignment = Alignment(horizontal='left', vertical='top')  # left, top как в шаблоне
        left_none_alignment = Alignment(horizontal='left', vertical=None)  # left, None как в шаблоне
        left_center_alignment = Alignment(horizontal='left', vertical='center')  # left, center как в шаблоне
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        thick_border = Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )

        # Логотип в начале документа (строки 1-6)
        if name_image_logo:
            try:
                img = Image(name_image_logo.path)
                # Устанавливаем размер изображения
                img.width = 155
                img.height = 35
                # Размещаем изображение в ячейке B1
                ws.add_image(img, 'B1')
            except Exception as e:
                print(f"Ошибка при добавлении логотипа: {e}")

        # Банковская информация (строки 8-14)
        # Строка 8: Банк получателя
        ws['B8'] = f"{motrum_info_req.bank}"
        ws['B8'].font = bank_font
        ws['B8'].alignment = top_alignment  # left, top как в шаблоне
        ws['B8'].border = thin_border
        ws.merge_cells('B8:AP9')
        
        ws['AQ8'] = 'БИК'
        ws['AQ8'].font = bank_small_font
        ws['AQ8'].alignment = left_alignment
        ws['AQ8'].border = thin_border
        ws.merge_cells('AQ8:AW8')
        
        ws['AX8'] = f"{motrum_info_req.bic}"
        ws['AX8'].font = bank_font
        ws['AX8'].alignment = left_alignment
        ws['AX8'].border = thin_border
        ws.merge_cells('AX8:BO8')

        # Строка 9: Сч. №
        ws['AQ9'] = 'Сч. №'
        ws['AQ9'].font = bank_small_font
        ws['AQ9'].alignment = left_alignment
        ws['AQ9'].border = thin_border
        ws.merge_cells('AQ9:AW10')
        
        ws['AX9'] = f"{motrum_info_req.account_requisites}"
        ws['AX9'].font = bank_font
        ws['AX9'].alignment = left_alignment
        ws['AX9'].border = thin_border
        ws.merge_cells('AX9:BO10')

        # Строка 10: Банк получателя
        ws['B10'] = 'Банк получателя'
        ws['B10'].font = bank_small_font
        ws['B10'].alignment = left_none_alignment  # left, None как в шаблоне
        ws['B10'].border = thin_border
        ws.merge_cells('B10:AP10')

        # Строка 11: ИНН и КПП
        ws['B11'] = 'ИНН'
        ws['B11'].font = bank_font
        ws['B11'].alignment = left_center_alignment  # left, center как в шаблоне
        ws['B11'].border = thin_border
        ws.merge_cells('B11:C11')
        
        ws['D11'] = f"{motrum_info.inn}"
        ws['D11'].font = bank_font
        ws['D11'].alignment = left_center_alignment  # left, center как в шаблоне
        ws['D11'].border = thin_border
        ws.merge_cells('D11:V11')
        
        ws['W11'] = 'КПП'
        ws['W11'].font = bank_font
        ws['W11'].alignment = left_center_alignment  # left, center как в шаблоне
        ws['W11'].border = thin_border
        ws.merge_cells('W11:AA11')
        
        ws['AB11'] = f"{motrum_info.kpp}"
        ws['AB11'].font = bank_font
        ws['AB11'].alignment = left_center_alignment  # left, center как в шаблоне
        ws['AB11'].border = thin_border
        ws.merge_cells('AB11:AP11')
        
        ws['AQ11'] = 'Сч. №'
        ws['AQ11'].font = bank_small_font
        ws['AQ11'].alignment = left_alignment
        ws['AQ11'].border = thin_border
        ws.merge_cells('AQ11:AW14')
        
        ws['AX11'] = f"{motrum_info_req.account_requisites}"
        ws['AX11'].font = bank_font
        ws['AX11'].alignment = left_alignment
        ws['AX11'].border = thin_border
        ws.merge_cells('AX11:BO14')

        # Строка 12: Получатель
        ws['B12'] = f"{motrum_info.full_name_legal_entity}"
        ws['B12'].font = bank_font
        ws['B12'].alignment = top_alignment  # left, top как в шаблоне
        ws['B12'].border = thin_border
        ws.merge_cells('B12:AP13')

        # Строка 14: Получатель
        ws['B14'] = 'Получатель'
        ws['B14'].font = bank_small_font
        ws['B14'].alignment = left_center_alignment  # left, center как в шаблоне
        ws['B14'].border = thin_border
        ws.merge_cells('B14:AP14')

        # Изображение поставщика под банковской информацией (строка 16)
        if name_image_vendor:
            try:
                img_vendor = Image(name_image_vendor.path)
                # Устанавливаем размер изображения
                img_vendor.width = 200
                img_vendor.height = 50
                # Размещаем изображение в ячейке B16
                ws.add_image(img_vendor, 'B16')
            except Exception as e:
                print(f"Ошибка при добавлении изображения поставщика: {e}")

        # Заголовок счета (строки 23-24)
        ws['B23'] = f"Счет на оплату № {bill_name} от {date_now} г."
        ws['B23'].font = title_font
        ws['B23'].alignment = center_alignment
        ws['B23'].border = thin_border
        ws.merge_cells('B23:BG24')

        # Информация о поставщике (строки 27-28)
        ws['B27'] = 'Поставщик'
        ws['B27'].font = normal_font
        ws['B27'].alignment = left_alignment
        ws.merge_cells('B27:E27')
        
        ws['F27'] = f" {motrum_info.full_name_legal_entity},  ИНН {motrum_info.inn},  КПП {motrum_info.kpp},  {motrum_info.legal_post_code}, {motrum_info.legal_city}, {motrum_info.legal_address},  тел.: {motrum_info.tel}"
        ws['F27'].font = normal_font
        ws['F27'].alignment = left_alignment
        ws.merge_cells('F27:BG28')

        ws['B28'] = '(исполнитель):'
        ws['B28'].font = normal_font
        ws['B28'].alignment = left_alignment
        ws.merge_cells('B28:E28')

        # Информация о покупателе (строки 30-31)
        ws['B30'] = 'Покупатель'
        ws['B30'].font = normal_font
        ws['B30'].alignment = left_alignment
        ws.merge_cells('B30:E30')
        
        if client_info.type_client == 1 or client_info.type_client == "1":
            # клиент юрлицо
            info_client = f" {client_info.legal_entity},  ИНН {client_info.inn},  КПП {client_info_req_kpp.kpp},  {client_info_req_kpp.legal_post_code}, {client_info_req_kpp.legal_city} {client_info_req_kpp.legal_address}"
        else:
            # клиент ип
            info_client = f" {client_info.legal_entity},  ИНН {client_info.inn},  ОГРНИП {client_info_req_kpp.ogrn}, {client_info_req_kpp.legal_post_code}, {client_info_req_kpp.legal_city} {client_info_req_kpp.legal_address}"

        if client_info_req_kpp.tel and client_info_req_kpp.tel != "":
            info_client = f"{info_client},  тел.: {client_info_req_kpp.tel}"

        ws['F30'] = info_client
        ws['F30'].font = normal_font
        ws['F30'].alignment = left_alignment
        ws.merge_cells('F30:BG31')

        ws['B31'] = '(заказчик):'
        ws['B31'].font = normal_font
        ws['B31'].alignment = left_alignment
        ws.merge_cells('B31:E31')

        # Основание (строка 33)
        ws['B33'] = 'Основание:'
        ws['B33'].font = normal_font
        ws['B33'].alignment = left_alignment
        ws.merge_cells('B33:E33')
        
        ws['F33'] = f"Счет {bill_name} от {date_name_dot}"
        ws['F33'].font = normal_font
        ws['F33'].alignment = left_alignment
        ws.merge_cells('F33:BG33')

        # Заголовки таблицы товаров (строка 36)
        ws['B36'] = '№'
        ws['B36'].font = header_font
        ws['B36'].alignment = center_alignment
        ws['B36'].border = thick_border
        ws.merge_cells('B36:C36')

        ws['D36'] = 'Товар (Услуга)'
        ws['D36'].font = header_font
        ws['D36'].alignment = center_alignment
        ws['D36'].border = thick_border
        ws.merge_cells('D36:Z36')

        ws['AA36'] = 'Код'
        ws['AA36'].font = header_font
        ws['AA36'].alignment = center_alignment
        ws['AA36'].border = thick_border
        ws.merge_cells('AA36:AM36')

        ws['AN36'] = 'Кол-во'
        ws['AN36'].font = header_font
        ws['AN36'].alignment = center_alignment
        ws['AN36'].border = thick_border
        ws.merge_cells('AN36:AS36')

        ws['AT36'] = 'Ед.'
        ws['AT36'].font = header_font
        ws['AT36'].alignment = center_alignment
        ws['AT36'].border = thick_border
        ws.merge_cells('AT36:AX36')

        ws['AY36'] = 'Цена'
        ws['AY36'].font = header_font
        ws['AY36'].alignment = center_alignment
        ws['AY36'].border = thick_border
        ws.merge_cells('AY36:BF36')

        ws['BG36'] = 'Сумма'
        ws['BG36'].font = header_font
        ws['BG36'].alignment = center_alignment
        ws['BG36'].border = thick_border
        ws.merge_cells('BG36:BJ36')

        ws['BK36'] = 'Срок поставки'
        ws['BK36'].font = header_font
        ws['BK36'].alignment = center_alignment
        ws['BK36'].border = thick_border
        ws.merge_cells('BK36:BN36')

        # Номера столбцов (строка 37)
        ws['B37'] = '1'
        ws['B37'].font = small_font
        ws['B37'].alignment = center_alignment
        ws['B37'].border = thin_border
        ws.merge_cells('B37:C37')

        ws['D37'] = '2'
        ws['D37'].font = small_font
        ws['D37'].alignment = center_alignment
        ws['D37'].border = thin_border
        ws.merge_cells('D37:Z37')

        ws['AA37'] = '3'
        ws['AA37'].font = small_font
        ws['AA37'].alignment = center_alignment
        ws['AA37'].border = thin_border
        ws.merge_cells('AA37:AM37')

        ws['AN37'] = '4'
        ws['AN37'].font = small_font
        ws['AN37'].alignment = center_alignment
        ws['AN37'].border = thin_border
        ws.merge_cells('AN37:AS37')

        ws['AT37'] = '5'
        ws['AT37'].font = small_font
        ws['AT37'].alignment = center_alignment
        ws['AT37'].border = thin_border
        ws.merge_cells('AT37:AX37')

        ws['AY37'] = '6'
        ws['AY37'].font = small_font
        ws['AY37'].alignment = center_alignment
        ws['AY37'].border = thin_border
        ws.merge_cells('AY37:BF37')

        ws['BG37'] = '7'
        ws['BG37'].font = small_font
        ws['BG37'].alignment = center_alignment
        ws['BG37'].border = thin_border
        ws.merge_cells('BG37:BJ37')

        ws['BK37'] = '8'
        ws['BK37'].font = small_font
        ws['BK37'].alignment = center_alignment
        ws['BK37'].border = thin_border
        ws.merge_cells('BK37:BN37')

        # Данные товаров (начиная со строки 38)
        row_num = 38
        total_product_quantity = 0
        
        for i, product in enumerate(product_specification, 1):
            try:
                product_stock_item = Stock.objects.get(prod=product.product)
                product_stock = product_stock_item.lot.name_shorts
            except Stock.DoesNotExist:
                product_stock = "шт"

            if product.product:
                product_name_str = product.product.name
                
                if product.product.supplier.slug == "prompower" and product.product.description:
                    product_name_str = product.product.description
                
                product_code = product.product.article_supplier
            else:
                product_name_str = product.product_new
                product_code = product.product_new_article

            product_price = product.price_one
            product_price_str = "{0:,.2f}".format(product_price).replace(",", " ").replace(".", ",")
            product_price_all = product.price_all
            product_price_all_str = "{0:,.2f}".format(product_price_all).replace(",", " ").replace(".", ",")
            product_quantity = product.quantity

            if product.date_delivery_bill:
                product_data = str(product.date_delivery_bill.strftime("%d.%m.%Y"))
            else:
                product_data = product.text_delivery

            total_product_quantity += product_quantity

            # Заполняем строку товара
            ws[f'B{row_num}'] = i
            ws[f'B{row_num}'].border = thin_border
            ws.merge_cells(f'B{row_num}:C{row_num}')
            
            ws[f'D{row_num}'] = product_name_str
            ws[f'D{row_num}'].border = thin_border
            ws.merge_cells(f'D{row_num}:Z{row_num}')
            
            ws[f'AA{row_num}'] = product_code
            ws[f'AA{row_num}'].border = thin_border
            ws.merge_cells(f'AA{row_num}:AM{row_num}')
            
            ws[f'AN{row_num}'] = product_quantity
            ws[f'AN{row_num}'].border = thin_border
            ws.merge_cells(f'AN{row_num}:AS{row_num}')
            
            ws[f'AT{row_num}'] = product_stock
            ws[f'AT{row_num}'].border = thin_border
            ws.merge_cells(f'AT{row_num}:AX{row_num}')
            
            ws[f'AY{row_num}'] = product_price_str
            ws[f'AY{row_num}'].border = thin_border
            ws.merge_cells(f'AY{row_num}:BF{row_num}')
            
            ws[f'BG{row_num}'] = product_price_all_str
            ws[f'BG{row_num}'].border = thin_border
            ws.merge_cells(f'BG{row_num}:BJ{row_num}')
            
            ws[f'BK{row_num}'] = product_data
            ws[f'BK{row_num}'].border = thin_border
            ws.merge_cells(f'BK{row_num}:BN{row_num}')

            row_num += 1

        # Итоговая строка
        total_amount_str = "{0:,.2f}".format(specifications.total_amount).replace(",", " ").replace(".", ",")
        
        ws[f'AN{row_num}'] = total_product_quantity
        ws[f'AN{row_num}'].font = normal_font
        ws[f'AN{row_num}'].border = thin_border
        ws.merge_cells(f'AN{row_num}:AS{row_num}')
        
        ws[f'AY{row_num}'] = total_amount_str
        ws[f'AY{row_num}'].font = normal_font
        ws[f'AY{row_num}'].border = thin_border
        ws.merge_cells(f'AY{row_num}:BJ{row_num}')

        # Итоги (строки row_num + 2, row_num + 3, row_num + 4)
        row_num += 2
        total_amount_nds = float(specifications.total_amount) * 20 / (20 + 100)
        total_amount_nds = round(total_amount_nds, 2)
        total_amount_nds_str = "{0:,.2f}".format(total_amount_nds).replace(",", " ").replace(".", ",")

        # Сначала устанавливаем значения, потом объединяем
        ws[f'AY{row_num}'] = total_amount_str
        ws[f'AY{row_num}'].font = bold_font
        ws[f'AY{row_num}'].alignment = right_alignment
        ws.merge_cells(f'AY{row_num}:BJ{row_num}')

        row_num += 1
        ws[f'AY{row_num}'] = total_amount_nds_str
        ws[f'AY{row_num}'].font = bold_font
        ws[f'AY{row_num}'].alignment = right_alignment
        ws.merge_cells(f'AY{row_num}:BJ{row_num}')

        row_num += 1
        ws[f'AY{row_num}'] = total_amount_str
        ws[f'AY{row_num}'].font = bold_font
        ws[f'AY{row_num}'].alignment = right_alignment
        ws.merge_cells(f'AY{row_num}:BJ{row_num}')

        # Сумма прописью (строки row_num + 2, row_num + 3)
        row_num += 2
        total_amount_word = num2words.num2words(int(specifications.total_amount), lang="ru").capitalize()
        total_amount_pens = str(specifications.total_amount).split(".")
        total_amount_pens = total_amount_pens[1]

        if len(total_amount_pens) < 2:
            total_amount_pens = f"{total_amount_pens}0"
        rub_word = rub_words(int(specifications.total_amount))

        ws[f'B{row_num}'] = f"Всего наименований {len(product_specification)}, на сумму {total_amount_str} руб."
        ws[f'B{row_num}'].font = normal_font
        ws[f'B{row_num}'].alignment = left_alignment
        ws.merge_cells(f'B{row_num}:BG{row_num}')

        row_num += 1
        ws[f'B{row_num}'] = f"{total_amount_word} {rub_word} {total_amount_pens} копеек"
        ws[f'B{row_num}'].font = bold_font
        ws[f'B{row_num}'].alignment = left_alignment
        ws.merge_cells(f'B{row_num}:BE{row_num}')

        # Условия (начиная со строки row_num + 2)
        row_num += 2
        if is_contract or specifications.total_amount > 99999.99:
            conditions = [
                "Оплата данного счета означает согласие с условиями поставки товара.",
                "Уведомление об оплате обязательно, в противном случае не гарантируется наличие товара на складе."
            ]
            
            if type_delivery.company_delivery is None:
                conditions.append("Товар отпускается по факту прихода денег на р/с Поставщика, самовывозом, при наличии доверенности и паспорта.")
            else:
                conditions.append(f"Условия доставки: {type_delivery.text_long}")
        else:
            conditions = [
                "1. Оплата данного счет-оферты означает полное и безоговорочное согласие (акцепт) с условиями поставки товара по наименованию, ассортименту, количеству и цене. Срок действия счета 3 банковских дня.",
                "2. Поставщик гарантирует отгрузку товара по ценам и в сроки, указанные в настоящем Счет-оферте, при условии зачисления денежных средств на расчетный счет Поставщика в течение 3 банковских дней с даты выставления счета. При невыполнении Покупателем указанных условий оплаты, цена и сроки поставки товара могут измениться.",
                "3. Товар отгружается после полной оплаты счета-оферты Покупателем.",
                "5. Обязательства Поставщика по поставке Товара считаются выполненными с момента подписания УПД или товарной накладной представителями Поставщика и Покупателя или организации перевозчика.",
                "6. Каждая партия поставляемой продукции сопровождается Универсальным передаточным документом (УПД): на бумажном носителе в двух экземплярах (1 экз. Покупателя, 1 экз. Поставщика). После каждой поставки продукции с документами, Покупатель обязан вернуть Поставщику один экземпляр, верно оформленного УПД, не позднее 1 месяца с даты подтверждения получения товара. В случае задержки Покупателем возврата, верно оформленного со стороны Покупателя, оригинала УПД на бумажном носителе на срок более 1 (одного) календарного месяца, Поставщик вправе предъявить Покупателю штрафные санкции в размере 5 000 (Пять тысяч) рублей (НДС не облагается) за каждый факт не предоставления подписанного оригинала УПД."
            ]

            if type_delivery.company_delivery is None:
                conditions.append("7. Доставка товара самовывозом со склада Поставщика")
            else:
                conditions.append(f"7.{type_delivery.text_long}")

            conditions.extend([
                "8. В случае превышения более чем на 15 дней сроков поставки продукции, указанных в Счёте-оферте, Поставщик по требованию Покупателя обязан уплатить неустойку в размере 0,1 % от стоимости не поставленной в срок продукции за каждый день просрочки, но не более 5% от стоимости не поставленной в срок продукции.",
                "9. Претензии по п. 8 должны быть заявлены Сторонами в письменной форме в течение 5 дней с момента наступления, указанных в них событий. В случае не выставления претензии в указанный срок, это трактуется как освобождение Сторон от уплаты неустойки.",
                "10. Срок гарантии на поставляемую продукцию составляет не менее одного года с момента отгрузки.",
                "11. Правила гарантийного обслуживания оговариваются в гарантийных талонах на поставляемую продукцию.",
                "12. Стороны принимают необходимые меры к тому, чтобы спорные вопросы и разногласия, возникающие при исполнении и расторжении настоящего договора, были урегулированы путем переговоров. В случае если стороны не достигнут соглашения по спорным вопросам путем переговоров, то спор передается заинтересованной стороной в арбитражный суд.",
            ])

        for condition in conditions:
            ws[f'B{row_num}'] = condition
            ws[f'B{row_num}'].font = normal_font
            ws[f'B{row_num}'].alignment = left_alignment
            ws.merge_cells(f'B{row_num}:BG{row_num}')
            row_num += 1

        # Подписи (начиная со строки row_num + 2)
        row_num += 2
        ws[f'B{row_num}'] = "Старостина В. П."
        ws[f'B{row_num}'].font = normal_font
        ws[f'B{row_num}'].alignment = left_alignment
        ws.merge_cells(f'B{row_num}:BD{row_num}')
        
        # Изображение подписи руководителя
        if name_image_motrum_signatures:
            try:
                img_signature = Image(name_image_motrum_signatures.path)
                # Устанавливаем размер изображения
                img_signature.width = 80
                img_signature.height = 30
                # Размещаем изображение подписи в ячейке C{row_num}
                ws.add_image(img_signature, f'C{row_num}')
            except Exception as e:
                print(f"Ошибка при добавлении подписи руководителя: {e}")

        row_num += 1
        ws[f'B{row_num}'] = "Старостина В. П."
        ws[f'B{row_num}'].font = normal_font
        ws[f'B{row_num}'].alignment = left_alignment
        ws.merge_cells(f'B{row_num}:BD{row_num}')
        
        # Изображение подписи бухгалтера (то же изображение)
        if name_image_motrum_signatures:
            try:
                img_signature2 = Image(name_image_motrum_signatures.path)
                # Устанавливаем размер изображения
                img_signature2.width = 80
                img_signature2.height = 30
                # Размещаем изображение подписи в ячейке C{row_num}
                ws.add_image(img_signature2, f'C{row_num}')
            except Exception as e:
                print(f"Ошибка при добавлении подписи бухгалтера: {e}")

        row_num += 2
        # Изображение печати рядом с М.П.
        if name_image_motrum_stamp:
            try:
                img_stamp = Image(name_image_motrum_stamp.path)
                # Устанавливаем размер изображения
                img_stamp.width = 60
                img_stamp.height = 60
                # Размещаем изображение печати в ячейке B{row_num}
                ws.add_image(img_stamp, f'B{row_num}')
            except Exception as e:
                print(f"Ошибка при добавлении печати: {e}")

        row_num += 3
        ws[f'B{row_num}'] = name_admin
        ws[f'B{row_num}'].font = normal_font
        ws[f'B{row_num}'].alignment = right_alignment

        # Настройка ширины столбцов (примерные значения из анализа)
        row_heights = {
            7: 11.0, 8: 13.0, 9: 11.0, 10: 11.0, 11: 13.0, 12: 11.0, 13: 11.0, 14: 11.0, 15: 11.0,
            22: 11.0, 23: 11.0, 24: 11.0, 25: 7.0, 26: 7.0, 27: 20.0, 28: 19.0, 29: 4.0
        }
        
        for row, height in row_heights.items():
            ws.row_dimensions[row].height = height

        # Сохранение файла
        fileName = os.path.join(directory, name_bill)
        wb.save(fileName)

        file_path = "{0}/{1}/{2}".format(
            "documents",
            "bill",
            name_bill,
        )
        convert_xlsx_to_pdf(fileName, pdf_file_path=None)
        print(file_path)

        return (
            file_path,
            bill_name,
            None,  # file_path_no_sign для xlsx не нужен
            version,
            name_bill_to_fullname,
            None,  # name_bill_to_fullname_nosign для xlsx не нужен
        )

    except Exception as e:
        tr = traceback.format_exc()
        error = "error"
        location = "Сохранение xlsx счета "
        info = f"Сохранение xlsx счета ошибка {e}{tr}"
        e = error_alert(error, location, info)


def convert_xlsx_to_pdf(xlsx_file_path, pdf_file_path=None):
    """
    Конвертирует XLSX файл в PDF используя pandas и pdfkit
    
    Args:
        xlsx_file_path (str): Путь к XLSX файлу
        pdf_file_path (str, optional): Путь для сохранения PDF. Если не указан, создается автоматически
    
    Returns:
        str: Путь к созданному PDF файлу
    """
    try:
        # Если путь к PDF не указан, создаем его автоматически
        if pdf_file_path is None:
            base_name = os.path.splitext(xlsx_file_path)[0]
            pdf_file_path = f"{base_name}.pdf"
        
        # Читаем XLSX файл с помощью pandas
        print(f"Читаем XLSX файл: {xlsx_file_path}")
        df = pd.read_excel(xlsx_file_path, engine='openpyxl')
        
        # Создаем временный HTML файл
        with tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False, encoding='utf-8') as temp_html:
            html_file_path = temp_html.name
            
            # Конвертируем DataFrame в HTML с улучшенным стилем
            html_content = df.to_html(
                index=False,
                classes=['table', 'table-striped', 'table-bordered'],
                table_id='data-table',
                escape=False
            )
            
            # Добавляем CSS стили для лучшего отображения
            full_html = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="utf-8">
                <title>Документ</title>
                <style>
                    body {{
                        font-family: Arial, sans-serif;
                        margin: 20px;
                        font-size: 12px;
                    }}
                    table {{
                        width: 100%;
                        border-collapse: collapse;
                        margin-bottom: 20px;
                    }}
                    th, td {{
                        border: 1px solid #ddd;
                        padding: 8px;
                        text-align: left;
                        vertical-align: top;
                    }}
                    th {{
                        background-color: #f2f2f2;
                        font-weight: bold;
                    }}
                    .table-striped tbody tr:nth-child(odd) {{
                        background-color: #f9f9f9;
                    }}
                    .table-bordered {{
                        border: 2px solid #ddd;
                    }}
                </style>
            </head>
            <body>
                {html_content}
            </body>
            </html>
            """
            
            temp_html.write(full_html)
        
        # Конвертируем HTML в PDF
        print(f"Конвертируем HTML в PDF: {pdf_file_path}")
        
        # Настройки для pdfkit
        options = {
            'page-size': 'A4',
            'margin-top': '0.75in',
            'margin-right': '0.75in',
            'margin-bottom': '0.75in',
            'margin-left': '0.75in',
            'encoding': "UTF-8",
            'no-outline': None,
            'enable-local-file-access': None
        }
        
        # Конвертируем HTML в PDF
        pdfkit.from_file(html_file_path, pdf_file_path, options=options)
        
        # Удаляем временный HTML файл
        os.unlink(html_file_path)
        
        print(f"PDF файл успешно создан: {pdf_file_path}")
        return pdf_file_path
        
    except Exception as e:
        tr = traceback.format_exc()
        error = "error"
        location = "Конвертация XLSX в PDF"
        info = f"Ошибка конвертации XLSX в PDF: {e}\n{tr}"
        error_alert(error, location, info)
        raise e


def create_xlsx_and_pdf_bill(
    specification,
    request,
    is_contract,
    order,
    type_delivery,
    post_update,
    type_save,
):
    """
    Создает XLSX счет и конвертирует его в PDF
    
    Returns:
        tuple: (xlsx_file_path, pdf_file_path, bill_name, version, name_bill_to_fullname)
    """
    try:
        # Сначала создаем XLSX файл
        xlsx_result = create_xlsx_bill(
            specification=specification,
            request=request,
            is_contract=is_contract,
            order=order,
            type_delivery=type_delivery,
            post_update=post_update,
            type_save=type_save,
        )
        
        xlsx_file_path, bill_name, _, version, name_bill_to_fullname, _ = xlsx_result
        
        # Получаем полный путь к XLSX файлу
        full_xlsx_path = os.path.join(MEDIA_ROOT, xlsx_file_path)
        
        # Конвертируем XLSX в PDF
        pdf_file_path = convert_xlsx_to_pdf(full_xlsx_path)
        
        # Создаем относительный путь для PDF
        relative_pdf_path = os.path.relpath(pdf_file_path, MEDIA_ROOT)
        relative_pdf_path = relative_pdf_path.replace('\\', '/')  # Для совместимости с Unix
        
        return (
            xlsx_file_path,
            relative_pdf_path,
            bill_name,
            version,
            name_bill_to_fullname,
        )
        
    except Exception as e:
        tr = traceback.format_exc()
        error = "error"
        location = "Создание XLSX и PDF счета"
        info = f"Ошибка создания XLSX и PDF счета: {e}\n{tr}"
        error_alert(error, location, info)
        raise e


def convert_xlsx_to_pdf_simple(xlsx_file_path, pdf_file_path=None):
    """
    Альтернативная функция конвертации XLSX в PDF без wkhtmltopdf
    Использует только pandas и reportlab
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        
        # Если путь к PDF не указан, создаем его автоматически
        if pdf_file_path is None:
            base_name = os.path.splitext(xlsx_file_path)[0]
            pdf_file_path = f"{base_name}.pdf"
        
        # Читаем XLSX файл с помощью pandas
        print(f"Читаем XLSX файл: {xlsx_file_path}")
        df = pd.read_excel(xlsx_file_path, engine='openpyxl')
        
        # Создаем PDF документ
        doc = SimpleDocTemplate(pdf_file_path, pagesize=A4)
        elements = []
        
        # Получаем данные из DataFrame
        data = [df.columns.tolist()]  # Заголовки
        data.extend(df.values.tolist())  # Данные
        
        # Создаем таблицу
        table = Table(data)
        
        # Стили для таблицы
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ])
        
        table.setStyle(style)
        elements.append(table)
        
        # Строим PDF
        doc.build(elements)
        
        print(f"PDF файл успешно создан: {pdf_file_path}")
        return pdf_file_path
        
    except Exception as e:
        tr = traceback.format_exc()
        error = "error"
        location = "Конвертация XLSX в PDF (простая версия)"
        info = f"Ошибка конвертации XLSX в PDF: {e}\n{tr}"
        error_alert(error, location, info)
        raise e


def convert_xlsx_to_pdf_advanced(xlsx_file_path, pdf_file_path=None):
    """
    Продвинутая функция конвертации XLSX в PDF с сохранением форматирования
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        
        # Если путь к PDF не указан, создаем его автоматически
        if pdf_file_path is None:
            base_name = os.path.splitext(xlsx_file_path)[0]
            pdf_file_path = f"{base_name}.pdf"
        
        # Читаем XLSX файл с помощью pandas
        print(f"Читаем XLSX файл: {xlsx_file_path}")
        df = pd.read_excel(xlsx_file_path, engine='openpyxl')
        
        # Создаем PDF документ
        doc = SimpleDocTemplate(
            pdf_file_path,
            pagesize=A4,
            rightMargin=20,
            leftMargin=20,
            topMargin=40,
            bottomMargin=50
        )
        
        # Регистрируем шрифты
        try:
            pdfmetrics.registerFont(TTFont("Roboto", "Roboto-Regular.ttf", "UTF-8"))
            pdfmetrics.registerFont(TTFont("Roboto-Bold", "Roboto-Bold.ttf", "UTF-8"))
        except:
            # Если шрифты не найдены, используем стандартные
            pass
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Добавляем стили
        styles.add(ParagraphStyle(
            name="CustomTitle",
            fontName="Helvetica-Bold",
            fontSize=16,
            alignment=1,  # Center
            spaceAfter=20
        ))
        
        styles.add(ParagraphStyle(
            name="CustomNormal",
            fontName="Helvetica",
            fontSize=10,
            spaceAfter=6
        ))
        
        # Заголовок
        title = Paragraph("Счет", styles["CustomTitle"])
        elements.append(title)
        elements.append(Spacer(1, 20))
        
        # Получаем данные из DataFrame
        data = [df.columns.tolist()]  # Заголовки
        data.extend(df.values.tolist())  # Данные
        
        # Создаем таблицу
        table = Table(data, repeatRows=1)
        
        # Стили для таблицы
        style = TableStyle([
            # Заголовки
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Данные
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
            
            # Границы
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            
            # Отступы
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ])
        
        table.setStyle(style)
        elements.append(table)
        
        # Строим PDF
        doc.build(elements)
        
        print(f"PDF файл успешно создан: {pdf_file_path}")
        return pdf_file_path
        
    except Exception as e:
        tr = traceback.format_exc()
        error = "error"
        location = "Конвертация XLSX в PDF (продвинутая версия)"
        info = f"Ошибка конвертации XLSX в PDF: {e}\n{tr}"
        error_alert(error, location, info)
        raise e