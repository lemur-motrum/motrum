from datetime import date, timedelta
from math import prod
import os
from struct import pack_into
import xml.etree.ElementTree as ET
from urllib.request import urlopen
from xml.etree import ElementTree, ElementInclude

import shutil
from zipfile import ZipFile

from apps.core.models import Currency
from apps.core.utils import get_price_supplier_rub
from apps.product.models import CurrencyRate, Price, Product
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl as op

from apps.specification.models import Specification
from project.settings import BASE_DIR, MEDIA_ROOT
from urllib.error import URLError, HTTPError


def pars():

    path = os.path.join(BASE_DIR, "docker/emas.xml")
    # # print(path)
    # o = open(path, 'r', encoding="utf-8")
    # print(o)
    # print(o)\
    xml_file = ET.parse(path)
    root = xml_file.getroot()
    ElementInclude.include(root)

    category = root.findall("./Классификатор/Группы/Группа/Наименование")
    for cat in category:
        cvb = root.findall(
            f"./Классификатор/Группы/Группа[Наименование='{cat.text}']/Группы/Группа//"
        )

        print("Категория", cat.text)

        if not cvb:
            pass
        else:
            cv = root.findall(
                f"./Классификатор/Группы/Группа[Наименование='{cat.text}']/Группы/Группа/Наименование"
            )

            for c in cv:
                print("Категория 22", c.text)

                cvlist = root.findall(
                    f"./Классификатор/Группы/Группа[Наименование='{cat.text}']/Группы/Группа/[Наименование='{c.text}']/Группы/Группа//"
                )

                if not cvlist:
                    pass
                else:
                    cvsd = root.findall(
                        f"./Классификатор/Группы/Группа[Наименование='{cat.text}']/Группы/Группа/[Наименование='{c.text}']/Группы/Группа/Наименование"
                    )
                    print(cvsd)
                    if not cvsd:
                        pass

                    else:

                        print(222222222222222222222222222222222222)
                        for rr in cvsd:

                            print("Категория 33", rr.text)
                    print(3333333333333333333333333333333)

    return xml_file


def pars_optimums():
    file_path = os.path.join(BASE_DIR, "docker/optimus.xlsx")
    file_path_name = os.path.join(BASE_DIR, "docker/optimus")
    file_path_name_zip = os.path.join(BASE_DIR, "docker/optimus.zip")
    print(file_path)
    # Создаем временную папку
    tmp_folder = os.path.join(BASE_DIR, "convert_wrong_excel")
    print(tmp_folder)
    os.makedirs(tmp_folder, exist_ok=True)

    # # Распаковываем excel как zip в нашу временную папку
    with ZipFile(file_path) as excel_container:
        excel_container.extractall(tmp_folder)

    # Переименовываем файл с неверным названием
    wrong_file_path = os.path.join(tmp_folder, "xl", "SharedStrings.xml")
    correct_file_path = os.path.join(tmp_folder, "xl", "sharedStrings.xml")
    os.rename(wrong_file_path, correct_file_path)

    # Запаковываем excel обратно в zip и переименовываем в исходный файл
    shutil.make_archive(file_path_name, "zip", tmp_folder)
    os.rename(file_path_name_zip, file_path)

    excel_doc = op.open(filename=file_path, data_only=True)

    sheetnames = excel_doc.sheetnames  # Получение списка листов книги
    sheet = excel_doc[sheetnames[0]]
    # a1 = sheet.cell(row = 11, column = 2).value
    # Считываем значения с ячейки A1
    wb = op.load_workbook(filename=file_path)
    sh = wb.worksheets[0]
    vendor = []
    all = []
    # sh.max_row
    for index in range(10, 100):

        row_level = sh.row_dimensions[index].outline_level + 1
        # ferst_level = sh.row_dimensions[index - 1].outline_level + 1
        # row_levels = sh.iter_rows(min_col=1, max_col=2, max_row=2):
        item_value = sh.cell(row=index, column=2).value
        if item_value == None:
            pass
        else:
            print(row_level, item_value)

        # if item_value == None and ferst_level != row_level:
        #     item_value_name = sh.cell(row=index-1, column=2).value
        #     print(item_value_name)

        # for col in sh[index]:
        #     vavl = col.value

        #     print("ТоЭвары", vavl)
        # i += 1

        # print(rgwert3erow_levels)


# def pars_optimums():
#     file_path = os.path.join(BASE_DIR, "docker/Odot.xlsx")


#     excel_doc = op.open(filename=file_path, data_only=True)


#     sheetnames = excel_doc.sheetnames #Получение списка листов книги
#     sheet = excel_doc[sheetnames[0]]
#     a1 = sheet.cell(row = 5, column = 2).value
#     i = 1
#     # while sheet[f'A{i}'].value is not None:

#     #     a = sheet[f'A{i}'].value #Первый способ чтения, обращаемся к имени ячейки
#     #     b = sheet[f'B{i}'].value
#     #     c = sheet[f'C{i}'].value
#     #     i += 1
#     #     print(f"{a} | {b} | {c}")

#     i = 1
#     # while sheet[f'A{i}'].value is not None:

#     #     a = sheet[f'A{i}'].value #Первый способ чтения, обращаемся к имени ячейки
#     #     b = sheet[f'B{i}'].value
#     #     c = sheet[f'C{i}'].value
#     #     i += 1
#     #     print(f"{a} | {b} | {c}")


#     #Считываем значения с ячейки A1

#     # wb = op.load_workbook(filename=file_path)
#     # sh = wb.worksheets[0]

#     # for index in range(sh.min_row, sh.max_row):
#     #     row_level = sh.row_dimensions[index].outline_level + 1
#     #     row_levels = sh.rows.value

#     #     print(row_levels)
