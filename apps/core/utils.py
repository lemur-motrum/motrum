# расчет цены

import csv
import datetime
import json
import random
import re
import shutil
import threading

import requests
import hashlib
import os
import traceback
from django.core.cache import cache
import traceback
from django.db import IntegrityError, transaction
from django.db.models import Q

from apps import supplier
from apps.core.models import Currency, CurrencyPercent, Vat

from apps.core.utils_web import send_email_message_html
from apps.logs.utils import error_alert
from requests.auth import HTTPBasicAuth


from apps.specification.utils import crete_pdf_specification


from project.settings import DOMIAN, FORM_LEMUR, MEDIA_ROOT, NDS
from simple_history.utils import update_change_reason
from django.utils.text import slugify
from pytils import translit
from openpyxl import load_workbook


def create_slug(name, arr_other_name):
    slug_text = name
    slugish = translit.translify(slug_text)
    slugish_name = slugish
    i = 0

    while slugish in arr_other_name:
        i += 1
        slugish = f"{slugish}_{i}"

    return slugish


# цена мотрум со скидкой
def get_price_motrum(
    item_category,
    item_group,
    vendors,
    rub_price_supplier,
    all_item_group,
    supplier,
    promo_groupe,
    is_need_vendor_serch=False
):
    from apps.supplier.models import (
        Discount,
    )
    print(item_category,
    item_group,
    vendors,
    rub_price_supplier,
    all_item_group,
    supplier,
    promo_groupe,
    is_need_vendor_serch
    )
    motrum_price = rub_price_supplier
    percent = 0
    sale = [None]

    # получение процента функция
    def get_percent(item):
        for i in item:
            return i.percent
    
    # промо группа
    if promo_groupe and percent == 0:
        discount_promo_groupe = Discount.objects.filter(
            promo_groupe=promo_groupe.id,
            is_tag_pre_sale=False,
        )

        if discount_promo_groupe:
            percent = get_percent(discount_promo_groupe)
            sale = discount_promo_groupe
    
    if all_item_group and percent == 0:
        discount_all_group = Discount.objects.filter(
            promo_groupe__isnull=True,
            category_supplier_all=all_item_group.id,
            is_tag_pre_sale=False,
        )
        if is_need_vendor_serch and discount_all_group:
            discount_all_group = discount_all_group.filter(
            vendor=vendors)

        if discount_all_group:
            percent = get_percent(discount_all_group)
            sale = discount_all_group

        # скидка по группе

    if item_group and percent == 0:
        print(item_group,"item_group")
        discount_group = Discount.objects.filter(
            promo_groupe__isnull=True,
            category_supplier_all__isnull=True,
            group_supplier=item_group.id,
            is_tag_pre_sale=False,
        )
        if is_need_vendor_serch and discount_group:
            print(discount_group,"discount_group")
            discount_group = discount_group.filter(
            vendor=vendors)
            

        if discount_group:
            percent = get_percent(discount_group)
            sale = discount_group

            # if percent != 0

    # скидка по категории
    if item_category and percent == 0:

        discount_categ = Discount.objects.filter(
            promo_groupe__isnull=True,
            category_supplier_all__isnull=True,
            group_supplier__isnull=True,
            category_supplier=item_category.id,
            is_tag_pre_sale=False,
        )
        if is_need_vendor_serch and discount_categ:
            discount_categ = discount_categ.filter(
            vendor=vendors)

        if discount_categ:
            percent = get_percent(discount_categ)
            sale = discount_categ

    if vendors and percent == 0:

        discount_all = Discount.objects.filter(
            vendor=vendors,
            group_supplier__isnull=True,
            category_supplier__isnull=True,
            category_supplier_all__isnull=True,
            promo_groupe__isnull=True,
            is_tag_pre_sale=False,
        )
        
        # скидка по всем вендору
        if discount_all:
            percent = get_percent(discount_all)
            sale = discount_all

    if percent == 0:

        discount_all = Discount.objects.filter(
            supplier=supplier,
            vendor__isnull=True,
            group_supplier__isnull=True,
            category_supplier__isnull=True,
            category_supplier_all__isnull=True,
            promo_groupe__isnull=True,
            is_tag_pre_sale=False,
        )
        # скидка по всем вендору
        if discount_all:
            percent = get_percent(discount_all)
            sale = discount_all
        # нет скидки
    if rub_price_supplier:
        print(sale)
        print(percent)
        motrum_price = rub_price_supplier - (rub_price_supplier / 100 * float(percent))
        # обрезать цены
        motrum_price = round(motrum_price, 2)
    else:
        motrum_price = None
    print("motrum_price",motrum_price)
    return motrum_price, sale[0]


# перевод валютной цены в рубли
def get_price_supplier_rub(currency, vat, vat_includ, price_supplier):
    from apps.product.models import CurrencyRate

    if vat_includ == True:
        vat = 0
    if price_supplier is not None:
        if currency == "RUB":
            price_supplier_vat = price_supplier + (price_supplier / 100 * vat)
            return price_supplier_vat
        else:
            currency_rate_query = CurrencyRate.objects.filter(
                currency__words_code=currency
            ).latest("date")
            currency_rate = currency_rate_query.vunit_rate
            current_percent = CurrencyPercent.objects.filter().latest("id")

            price_supplier_vat = price_supplier + (price_supplier / 100 * vat)

            price_supplier_rub = (
                price_supplier_vat * currency_rate * current_percent.percent
            )

            return round(price_supplier_rub, 2)
    else:
        return None


# получение комплектности и расчет штук
def get_lot(
    lot, stock_supplier, lot_complect, stock_supplier_unit, force_stock_supplier_unit
):
    from apps.product.models import Lot

    if lot_complect == None:
        lot_complect = 1
    else:
        lot_complect = lot_complect

    if lot == "base" or lot == "штука":
        print("lot == base or lot == штука:")
        lots = Lot.objects.get(name_shorts="шт")

        # lot_stock = stock_supplier

    else:
        print("lot != штука:")
        lots = Lot.objects.get(name=lot)
        # lot_stock = stock_supplier * lot_complect
    print("lot_complect", lot_complect)
    lot_stock = stock_supplier * lot_complect
    print("lot_stock", lot_stock)
    print("force_stock_supplier_unit", force_stock_supplier_unit)
    if force_stock_supplier_unit:
        lot_stock = stock_supplier_unit
    print("get_lot", lots, lot_stock, lot_complect)
    return (lots, lot_stock, lot_complect)


# артикул мотрум
def create_article_motrum(supplier):
    from apps.product.models import Product

    supplier_int = str(supplier).zfill(3)

    try:
        prev_product = Product.objects.filter(supplier=supplier).latest("id")
        last_item_id = str(prev_product.article)[3:]
        last_item_id_int = int(last_item_id) + 1
        name = f"{supplier_int}{last_item_id_int}"
    except Product.DoesNotExist:
        prev_product = None

        name = f"{supplier_int}1"

    return name


# категории поставщика для товара
def get_category(supplier, vendor, category_name):
    from apps.supplier.models import (
        SupplierCategoryProduct,
        SupplierCategoryProductAll,
        SupplierGroupProduct,
    )

    try:
        item_category_all = SupplierCategoryProductAll.objects.filter(
            supplier=supplier, name=category_name
        )
        item_category = item_category_all[0].category_supplier
        item_group = item_category_all[0].group_supplier
    except SupplierCategoryProductAll.DoesNotExist:
        item_category = None
        item_group = None
        item_category_all = None

    return (item_category, item_group, item_category_all[0])


# категории поставщика промповер для товара
def get_category_prompower(supplier, vendor, category_name):
    from apps.supplier.models import (
        SupplierCategoryProduct,
        SupplierCategoryProductAll,
        SupplierGroupProduct,
    )

    print(category_name)
    try:
        category_all = SupplierCategoryProductAll.objects.get(
            supplier=supplier, vendor=vendor, article_name=category_name
        )

        groupe = category_all.group_supplier
        categ = category_all.category_supplier
        print(" try:1", category_all, groupe, categ)
    except SupplierCategoryProductAll.DoesNotExist:
        try:
            groupe = SupplierGroupProduct.objects.get(
                supplier=supplier, vendor=vendor, article_name=category_name
            )
            category_all = None

            categ = groupe.category_supplier
        except SupplierGroupProduct.DoesNotExist:
            try:
                categ = SupplierCategoryProduct.objects.get(
                    supplier=supplier, vendor=vendor, article_name=category_name
                )
                category_all = None
                groupe = None

            except SupplierGroupProduct.DoesNotExist:
                category_all = None
                groupe = None
                categ = None

    ("return get_category_prompower", category_all, groupe, categ)

    return (category_all, groupe, categ)


# категории поставщика промповер для товара
def get_category_unimat(supplier, vendor, category_name):
    from apps.supplier.models import (
        SupplierCategoryProduct,
        SupplierCategoryProductAll,
        SupplierGroupProduct,
    )

    print(category_name)
    try:
        category_all = SupplierCategoryProductAll.objects.get(
            supplier=supplier,  article_name=category_name
        )

        groupe = category_all.group_supplier
        categ = category_all.category_supplier
        print(" try:1", category_all, groupe, categ)
    except SupplierCategoryProductAll.DoesNotExist:
        try:
            groupe = SupplierGroupProduct.objects.get(
                supplier=supplier,  article_name=category_name
            )
            category_all = None

            categ = groupe.category_supplier
        except SupplierGroupProduct.DoesNotExist:
            try:
                categ = SupplierCategoryProduct.objects.get(
                    supplier=supplier, article_name=category_name
                )
                category_all = None
                groupe = None

            except SupplierGroupProduct.DoesNotExist:
                category_all = None
                groupe = None
                categ = None

    ("return get_category_prompower", category_all, groupe, categ)

    return (category_all, groupe, categ)


# ктегории поставщика для еиас
def get_category_emas(supplier, category_name):
    from apps.supplier.models import (
        SupplierCategoryProduct,
        SupplierCategoryProductAll,
        SupplierGroupProduct,
    )

    try:
        category_all = SupplierCategoryProductAll.objects.get(
            supplier=supplier, article_name=category_name
        )
        groupe = category_all.group_supplier
        categ = category_all.category_supplier
    except SupplierCategoryProductAll.DoesNotExist:
        try:
            groupe = SupplierGroupProduct.objects.get(
                supplier=supplier, article_name=category_name
            )
            category_all = None

            categ = groupe.category_supplier
        except SupplierGroupProduct.DoesNotExist:
            try:
                categ = SupplierCategoryProduct.objects.get(
                    supplier=supplier, article_name=category_name
                )
                category_all = None
                groupe = None

            except SupplierGroupProduct.DoesNotExist:
                category_all = None
                groupe = None
                categ = None
    return (category_all, groupe, categ)


# проверка есть ли путь и папка
def check_media_directory_exist(
    base_dir, base_dir_supplier, base_dir_vendor, base_dir_type_file, article_suppliers
):
    new_dir = "{0}/{1}/{2}/{3}/{4}/{5}".format(
        MEDIA_ROOT,
        base_dir,
        base_dir_supplier,
        base_dir_vendor,
        base_dir_type_file,
        article_suppliers,
    )
    if not os.path.exists(new_dir):
        os.makedirs(new_dir)


# проверка директории для спецификаций
def check_spesc_directory_exist(
    base_dir,
):
    doc = "document"
    new_dir = "{0}/{1}/{2}".format(
        MEDIA_ROOT,
        "documents",
        base_dir,
    )

    if not os.path.exists(new_dir):
        os.makedirs(new_dir)
    return new_dir


# проверка директории для загрузки прайса из админки
def check_file_price_directory_exist(base_dir, base_dir_supplier):
    import shutil

    new_dir = "{0}/{1}/{2}".format(MEDIA_ROOT, base_dir, base_dir_supplier)

    if not os.path.exists(new_dir):
        os.makedirs(new_dir)
    else:
        shutil.rmtree(new_dir)

    return new_dir


# переименовывание изображений и документов по очереди
def create_name_file_downloading(article_suppliers, item_count):

    try:

        count = f"{item_count:05}"
        filename = "{0}_{1}".format(article_suppliers, count)
        return filename
    except item_count.DoesNotExist:
        return filename


def get_file_path(supplier, vendor, type_file, article_suppliers, item_count, place):
    base_dir = "products"
    base_dir_supplier = supplier
    base_dir_vendor = vendor
    base_dir_type_file = type_file
    filename = create_name_file_downloading(article_suppliers, item_count)

    check_media_directory_exist(
        base_dir,
        base_dir_supplier,
        base_dir_vendor,
        article_suppliers,
        base_dir_type_file,
    )
    if place == "utils":
        return "{0}/{1}/{2}/{3}/{4}".format(
            base_dir,
            base_dir_supplier,
            base_dir_vendor,
            article_suppliers,
            base_dir_type_file,
        )
    else:
        return "{0}/{1}/{2}/{3}/{4}/{5}".format(
            base_dir,
            base_dir_supplier,
            base_dir_vendor,
            article_suppliers,
            base_dir_type_file,
            filename,
        )


def save_file_product(link, image_path):
    r = requests.get(link, stream=True)
    with open(os.path.join(MEDIA_ROOT, image_path), "wb") as ofile:
        ofile.write(r.content)


def save_file_emas_product(link, image_path):

    out = f"{MEDIA_ROOT}/price/emas_site/{link}"
    in_out = f"{MEDIA_ROOT}/{image_path}"
    shutil.copy2(out, in_out)


# сохранение изображений и докуметов из админки и общее
def get_file_path_add(instance, filename):

    from apps.product.models import ProductDocument
    from apps.product.models import ProductImage
    from pytils import translit

    s = str(instance.product.article_supplier)
    item_instanse_name = re.sub("[^A-Za-z0-9]", "", s)

    base_dir = "products"
    base_dir_supplier = instance.product.supplier.slug

    if instance.product.vendor:
        base_dir_vendor = instance.product.vendor.slug
    else:
        base_dir_vendor = ""

    images_last_list = filename.split(".")
    type_file = "." + images_last_list[-1]

    if isinstance(instance, ProductDocument):

        path_name = "document"
        try:
            images_last = ProductDocument.objects.filter(
                product=instance.product
            ).latest("id")
            item_count = ProductDocument.objects.filter(
                product=instance.product
            ).count()
        except ProductDocument.DoesNotExist:
            item_count = 1

        filenames = create_name_file_downloading(item_instanse_name, item_count)
        filename = f"{filenames}_{instance.type_doc}{type_file}"

    elif isinstance(instance, ProductImage):
        path_name = "img"
        try:
            images_last = ProductImage.objects.filter(product=instance.product).latest(
                "id"
            )
            item_count = ProductImage.objects.filter(product=instance.product).count()
        except ProductImage.DoesNotExist:
            item_count = 1

        filenames = create_name_file_downloading(item_instanse_name, item_count)

        filename = filenames + type_file

    check_media_directory_exist(
        base_dir,
        base_dir_supplier,
        base_dir_vendor,
        item_instanse_name,
        path_name,
    )
    return "{0}/{1}/{2}/{3}/{4}/{5}".format(
        base_dir,
        base_dir_supplier,
        base_dir_vendor,
        item_instanse_name,
        path_name,
        filename,
    )


def get_file_path_add_more_vac(product, type_doc, filename):

    from apps.product.models import ProductDocument
    from apps.product.models import ProductImage

    from pytils import translit

    base_dir = "documents"
    path_name = "vacancy"
    type_doc = type_doc

    slug_text = str(filename)
    regex = r"[^A-Za-z0-9,А-ЯЁа-яё, ,-.]"
    slugish = re.sub(regex, "", slug_text)
    slugish = translit.translify(slugish)
    slugish = slugify(slugish)

    # link_file = f"{new_dir}/{slugish}"
    new_dir = "{0}/{1}/{2}".format(
        MEDIA_ROOT,
        base_dir,
        # base_dir_supplier,
        # base_dir_vendor,
        path_name,
    )
    if not os.path.exists(new_dir):
        os.makedirs(new_dir)

    link = "{0}/{1}".format(
        base_dir,
        # base_dir_supplier,
        # base_dir_vendor,
        path_name,
    )
    print(new_dir, link, slugish)

    return (new_dir, link, slugish)


def get_file_path_add_more_doc(product, type_doc, filename):

    from apps.product.models import ProductDocument
    from apps.product.models import ProductImage

    from pytils import translit

    base_dir = "products"
    path_name = "document_group"
    # base_dir_supplier = product.supplier.slug
    # if product.vendor:
    #     base_dir_vendor = product.vendor.slug
    # else:
    #     base_dir_vendor = "vendor-name"

    # if product.category:
    #     base_dir_vendor = product.category.slug
    # else:
    #     base_dir_vendor = "category-name"

    type_doc = type_doc

    slug_text = str(filename)
    regex = r"[^A-Za-z0-9,А-ЯЁа-яё, ,-.]"
    slugish = re.sub(regex, "", slug_text)
    slugish = translit.translify(slugish)
    slugish = slugify(slugish)

    # link_file = f"{new_dir}/{slugish}"
    new_dir = "{0}/{1}/{2}/{3}".format(
        MEDIA_ROOT,
        base_dir,
        # base_dir_supplier,
        # base_dir_vendor,
        path_name,
        type_doc,
    )
    if not os.path.exists(new_dir):
        os.makedirs(new_dir)

    link = "{0}/{1}/{2}".format(
        base_dir,
        # base_dir_supplier,
        # base_dir_vendor,
        path_name,
        type_doc,
    )
    print(new_dir, link, slugish)

    return (new_dir, link, slugish)


# сохранение изображений и докуметов из админки и общее
def doc_file_mass_upload(instance, filename):
    from apps.product.models import ProductDocument

    base_dir = "products"
    path_name = "document_group"
    base_dir_supplier = instance.product.supplier.slug
    base_dir_vendor = instance.product.vendor.slug
    group_name = instance.product.category_supplier.slug

    s = str(instance.product.article_supplier)
    item_instanse_name = re.sub("[^A-Za-z0-9]", "", s)
    new_dir = "{0}/{1}/{2}/{3}/{4}/{5}".format(
        MEDIA_ROOT,
        base_dir,
        base_dir_supplier,
        base_dir_vendor,
        path_name,
        group_name,
    )
    if not os.path.exists(new_dir):
        os.makedirs(new_dir)
    dir_no_path = "{0}/{1}/{2}/{3}/{4}".format(
        base_dir,
        base_dir_supplier,
        base_dir_vendor,
        path_name,
        group_name,
    )
    # base_dir = "products"
    # base_dir_supplier = instance.product.supplier.slug

    # if instance.product.vendor:
    #     base_dir_vendor = instance.product.vendor.slug
    # else:
    #     base_dir_vendor = ""

    images_last_list = filename.split(".")
    type_file = "." + images_last_list[-1]

    if isinstance(instance, ProductDocument):
        path_name = "document"
        try:
            images_last = ProductDocument.objects.filter(
                product=instance.product
            ).latest("id")
            item_count = ProductDocument.objects.filter(
                product=instance.product
            ).count()
        except ProductDocument.DoesNotExist:
            item_count = 1

        filenames = create_name_file_downloading(item_instanse_name, item_count)
        filename = f"{filenames}_{instance.type_doc}{type_file}"

    # check_media_directory_exist(
    #     base_dir,
    #     base_dir_supplier,
    #     base_dir_vendor,
    #     item_instanse_name,
    #     path_name,
    # )
    return "{0}/{1}/{2}/{3}/{4}/{5}".format(
        base_dir,
        base_dir_supplier,
        base_dir_vendor,
        item_instanse_name,
        path_name,
        filename,
    )


def get_file_path_add_motrum_base(instance, filename):
    from pytils import translit

    new_dir = "{0}/{1}".format(MEDIA_ROOT, "documents")
    if not os.path.exists(new_dir):
        os.makedirs(new_dir)

    slug_text = str(filename)
    regex = r"[^A-Za-z0-9,А-ЯЁа-яё, ,-.]"
    slugish = re.sub(regex, "", slug_text)
    slugish = translit.translify(slugish)

    link_file = f"{new_dir}/{slugish}"

    return "{0}/{1}".format(
        "documents",
        f"{slugish}",
    )


# проверка есть ли такой тип лота шт комп
def lot_chek(lot):
    from apps.product.models import Lot

    try:
        lot_item = Lot.objects.get(name_shorts=lot)
    except Lot.DoesNotExist:
        lot_item = Lot.objects.create(name_shorts=lot, name=lot)

    return lot_item


# ответ по соединению с апи
def response_request(response, location):
    if response >= 200 and response <= 399:
        return True
    else:
        error = "file api"
        if response == 502:
            pass
            # error_alert(error, location, response)
        # else:
        #     error_alert(error, location, response)
        return False


# создание времени окончания спецификации
def create_time_stop_specification():
    from apps.core.models import CalendarHoliday

    year_date = datetime.datetime.now().year
    year = str(year_date)
    data_bd = CalendarHoliday.objects.get(year=year)
    data_bd_holidays = data_bd.json_date
    now = datetime.datetime.now()

    day_need = 1
    i = 0
    while day_need < 3:

        i += 1
        date = now + datetime.timedelta(days=i)
        # расчет на стыке годов
        if date.year > year_date:
            data_bd = CalendarHoliday.objects.get(year=date.year)
            data_bd_holidays = data_bd.json_date
            holidays_day_need = data_bd_holidays["holidays"].count(str(date.date()))

            if "nowork" in data_bd_holidays and holidays_day_need == 0:
                holidays_day_need_nowork = data_bd_holidays["nowork"].count(
                    str(date.date())
                )
                holidays_day_need += holidays_day_need_nowork
        # расчет внутри текущего года
        else:

            holidays_day_need = data_bd_holidays["holidays"].count(str(date.date()))

            if "nowork" in data_bd_holidays and holidays_day_need == 0:
                holidays_day_need_nowork = data_bd_holidays["nowork"].count(
                    str(date.date())
                )
                holidays_day_need += holidays_day_need_nowork

        if holidays_day_need == 0:

            day_need += 1

    three_days = datetime.timedelta(i + 1)
    in_three_days = now + three_days
    data_stop = in_three_days.strftime("%Y-%m-%d")

    return data_stop
    # year_date = datetime.datetime.now().year
    # year = str(year_date)

    # data_bd = CalendarHoliday.objects.get(year=year)
    # data_bd_holidays = data_bd.json_date

    # now = datetime.datetime.now()
    # day_work = 3
    # list_day = []
    # for x in range(3):
    #     one_day = now + datetime.timedelta(days=x)
    #     list_day.append(one_day)

    # for list_day_item in list_day:
    #     list_day_item_date = list_day_item.date()
    #     holidays_day = data_bd_holidays["holidays"].count(str(list_day_item_date))
    #     day_work = day_work + holidays_day

    # three_days = datetime.timedelta(day_work)
    # in_three_days = now + three_days
    # data_stop = in_three_days.strftime("%Y-%m-%d")

    # return data_stop


# получение категорий мотрум из категорий поставщика
def get_motrum_category(self):
    category_catalog = None
    group_catalog = None
    if self.category_supplier_all != None:
        category_catalog = self.category_supplier_all.category_catalog
        group_catalog = self.category_supplier_all.group_catalog

    if self.group_supplier != None:
        print(self.group_supplier)
        if category_catalog == None and group_catalog == None:
            category_catalog = self.group_supplier.category_catalog
            print(category_catalog)
            group_catalog = self.group_supplier.group_catalog
            print(group_catalog)

    if self.category_supplier != None:
        if category_catalog == None and group_catalog == None:
            category_catalog = self.category_supplier.category_catalog
            group_catalog = self.category_supplier.group_catalog
    print(category_catalog, group_catalog)
    return (category_catalog, group_catalog)


# сохранение фаилов прайс из админки
def get_file_price_path_add(instance, filename):
    if instance.slug == "delta":
        base_dir = "price"
        base_dir_supplier = instance.slug

        current_date = datetime.date.today().isoformat()

        new_dir = check_file_price_directory_exist(
            base_dir,
            base_dir_supplier,
        )
        random_number = random.randint(1000, 9999)

        file = "{0}/{1}/{2}_{3}".format(
            base_dir,
            base_dir_supplier,
            random_number,
            instance.file,
        )

        return file

    elif instance.slug == "optimus-drive":
        base_dir = "price"
        base_dir_supplier = instance.slug

        current_date = datetime.date.today().isoformat()

        new_dir = check_file_price_directory_exist(
            base_dir,
            base_dir_supplier,
        )
        random_number = random.randint(1000, 9999)

        file = "{0}/{1}/{2}_{3}".format(
            base_dir,
            base_dir_supplier,
            random_number,
            instance.file,
        )

        return file

    elif instance.slug == "emas":
        base_dir = "price"
        base_dir_supplier = instance.slug

        current_date = datetime.date.today().isoformat()

        new_dir = check_file_price_directory_exist(
            base_dir,
            base_dir_supplier,
        )
        random_number = random.randint(1000, 9999)

        file = "{0}/{1}/{2}_{3}".format(
            base_dir,
            base_dir_supplier,
            random_number,
            instance.file,
        )

        return file

    elif instance.slug == "avangard":
        base_dir = "price"
        base_dir_supplier = instance.slug

        current_date = datetime.date.today().isoformat()

        new_dir = check_file_price_directory_exist(
            base_dir,
            base_dir_supplier,
        )
        random_number = random.randint(1000, 9999)

        file = "{0}/{1}/{2}_{3}".format(
            base_dir,
            base_dir_supplier,
            random_number,
            instance.file,
        )

        return file


# проверка заполненны ли поля продукта если нет добавить значение
def save_update_product_attr(
    product,
    supplier,
    vendor,
    additional_article_supplier,
    category_supplier_all,
    group_supplier,
    category_supplier,
    description,
    name,
):

    try:

        if product.supplier == None or product.supplier == "":
            product.supplier = supplier

        if product.vendor == None or product.vendor == "":
            product.vendor = vendor
        if product.vendor == None or product.vendor == "":
            product.vendor = vendor

        if (
            product.additional_article_supplier == None
            or product.additional_article_supplier == ""
        ):
            product.additional_article_supplier = additional_article_supplier
        if (
            product.additional_article_supplier == None
            or product.additional_article_supplier == ""
        ):
            product.additional_article_supplier = additional_article_supplier

        if product.category_supplier_all == None or product.category_supplier_all == "":
            product.category_supplier_all = category_supplier_all
        if product.category_supplier_all == None or product.category_supplier_all == "":
            product.category_supplier_all = category_supplier_all

        if product.group_supplier == None or product.group_supplier == "":
            product.group_supplier = group_supplier
        if product.group_supplier == None or product.group_supplier == "":
            product.group_supplier = group_supplier

        if product.category_supplier == None or product.category_supplier == "":
            product.category_supplier = category_supplier
        if product.category_supplier == None or product.category_supplier == "":
            product.category_supplier = category_supplier

        if product.description == None or product.description == "":
            product.description = description
        if product.description == None or product.description == "":
            product.description = description

        if product.name == None or product.name == "":
            product.name = name
        if product.name == None or product.name == "":
            product.name = name

        product._change_reason = "Автоматическое"
        product.save()
    except Exception as e:
        print(e)
        tr = traceback.format_exc()
        error = "file_api_error"
        location = "обновление товаров"
        info = f"ошибка при чтении товара артикул ИЗ ФУНКЦИИ save_update_product_attr: {name}. Тип ошибки:{e}{tr}"
        e = error_alert(error, location, info)
    # update_change_reason(product, "Автоматическое")


# проверка заполненны ли поля продукта если нет добавить значение
def save_update_product_attr_all(
    product,
    supplier,
    vendor,
    additional_article_supplier,
    category_supplier_all,
    group_supplier,
    category_supplier,
    description,
    name,
    promo_groupe,
):

    try:
        print( "save_update_product_attr_all")

        print(
            product,
            supplier,
            vendor,
            additional_article_supplier,
            category_supplier_all,
            group_supplier,
            category_supplier,
            description,
            name,
            promo_groupe,
        )
        print(promo_groupe, "promo_groupe")

        if product.supplier == None or product.supplier == "":
            product.supplier = supplier

        if product.vendor == None or product.vendor == "":
            product.vendor = vendor

        if (
            product.additional_article_supplier == None
            or product.additional_article_supplier == ""
        ):
            product.additional_article_supplier = additional_article_supplier
        if (
            product.additional_article_supplier == None
            or product.additional_article_supplier == ""
        ):
            product.additional_article_supplier = additional_article_supplier

        if category_supplier_all:
            product.category_supplier_all = category_supplier_all

        if group_supplier:
            product.group_supplier = group_supplier

        if category_supplier:
            product.category_supplier = category_supplier

        if product.description == None or product.description == "":
            product.description = description

        if product.name == None or product.name == "":
            product.name = name

        if promo_groupe:
            print(promo_groupe, "promo_groupe24")
            product.promo_groupe = promo_groupe
            print("product.promo_groupe", product.promo_groupe)

        product._change_reason = "Автоматическое"
        product.save()
    except Exception as e:
        print(e)
        tr = traceback.format_exc()
        error = "file_api_error"
        location = "обновление товаров"
        info = f"ошибка при чтении товара артикул ИЗ ФУНКЦИИ save_update_product_attr: {name}. Тип ошибки:{e}{tr}"
        e = error_alert(error, location, info)
    # update_change_reason(product, "Автоматическое")


def save_specification(
    received_data,
    pre_sale,
    request,
    motrum_requisites,
    account_requisites,
    requisites,
    id_bitrix,
    type_delivery,
    post_update,
    # specification_name,
    type_save,
):
    from apps.product.models import Price, Product
    from apps.specification.models import ProductSpecification, Specification
    from apps.specification.utils import crete_pdf_specification
    from apps.product.models import ProductCart
    from apps.core.utils import create_time_stop_specification

    # try:

    # сохранение спецификации
    id_bitrix = received_data["id_bitrix"]  # сюда распарсить значения с фронта
    admin_creator_id = received_data["admin_creator_id"]
    id_specification = received_data["id_specification"]
    specification_comment = received_data["comment"]
    date_delivery_all = received_data["date_delivery"]
    products = received_data["products"]
    id_cart = received_data["id_cart"]
    print("products", products)
    # первичное создание/взятие спецификации
    try:
        specification = Specification.objects.get(id=id_specification)
        if post_update:
            pass
        else:
            # if specification_name:
            #     specification.number = specification_name
            data_stop = create_time_stop_specification()
            specification.date_stop = data_stop
            specification.tag_stop = True

        # удалить продукты если удалили из спецификации
        product_old = ProductSpecification.objects.filter(specification=specification)
        for product_item_for_old in product_old:
            item_id = product_item_for_old.id
            having_items = False
            for products_new in products:
                if (
                    products_new["product_specif_id"] != "None"
                    and products_new["product_specif_id"] != None
                ):

                    if int(products_new["product_specif_id"]) == item_id:
                        having_items = True

            if having_items == False:

                specification._change_reason = "Ручное"
                product_item_for_old.delete()

            having_items = False

            # for i, dic in enumerate(products):
            #     if dic["product_specif_id"] == item_id:
            #         having_items = True

            # if having_items == False:

            #     product_item_for_old.delete()

    except Specification.DoesNotExist:
        specification = Specification(
            id_bitrix=id_bitrix, admin_creator_id=admin_creator_id, cart_id=id_cart
        )
        # if specification_name:
        #     specification.number = specification_name
        # specification.skip_history_when_saving = True
        data_stop = create_time_stop_specification()
        specification.date_stop = data_stop
        specification.tag_stop = True
        specification._change_reason = "Ручное"
        specification.save()

    # сохранение продуктов для спецификации
    # перебор продуктов и сохранение
    total_amount = 0.00
    total_amount_motrum = 0.00
    print(products)
    for product_item in products:
        print(product_item)
        # продукты которые есть в окт
        if (
            product_item["product_new_article"] == ""
            or product_item["product_new_article"] == None
            or product_item["product_new_article"] == "None"
        ):

            product = Product.objects.get(id=product_item["product_id"])
            price_data = float(product_item["price_one"])

            if product_item["sale_motrum"]:
                sale_motrum_data = product_item["sale_motrum"]
                sale_motrum_data = sale_motrum_data.replace(".", "")
                sale_motrum_data = sale_motrum_data.replace(",", ".")
                sale_motrum_data = float(sale_motrum_data)
            else:
                sale_motrum_data = 0.00

            # sale_motrum_data = float(product_item["sale_motrum"].replace(",", "."))

            price_okt = Price.objects.get(prod=product)

            # если цена по запросу взять ее если нет взять цену из бд
            if (
                product_item["price_exclusive"] != "0"
                and product_item["price_exclusive"] != ""
                and product_item["price_exclusive"] != 0
            ):
                price_one = price_data
                price_one_motrum = price_one - (price_one / 100 * sale_motrum_data)
                price_one_motrum = round(price_one_motrum, 2)
                product_price_catalog = None

            else:
                price_one = product_item["price_one"]
                price_one_motrum = price_one - (price_one / 100 * sale_motrum_data)
                price_one_motrum = round(price_one_motrum, 2)
                product_price_catalog = Price.objects.get(
                    prod=product
                ).rub_price_supplier

            # если есть доп скидка отнять от цены поставщика
            if (
                product_item["extra_discount"] != "0"
                and product_item["extra_discount"] != ""
                and product_item["extra_discount"] != 0
            ):

                persent_sale = float(product_item["extra_discount"])

                price_one_sale = price_one - (price_one / 100 * persent_sale)
                price_one = round(price_one_sale, 2)

            # если есть предоплата найти скидку по предоплате мотрум
            persent_pre_sale = 0

            if pre_sale and price_okt.in_auto_sale:
                price_pre_sale = get_presale_discount(product)
                if price_pre_sale:
                    persent_pre_sale = price_pre_sale.percent
                    price_one_motrum = price_one_motrum - (
                        price_one_motrum / 100 * float(persent_pre_sale)
                    )
                    price_one_motrum = round(price_one_motrum, 2)

            price_all = float(price_one) * int(product_item["quantity"])
            price_all = round(price_all, 2)

            price_all_motrum = float(price_one_motrum) * int(product_item["quantity"])
            price_all_motrum = round(price_all_motrum, 2)

            # выбор продукт из спецификации или заспись нового
            if (
                product_item["product_specif_id"] != "None"
                and product_item["product_specif_id"] != None
            ):
                product_spes = ProductSpecification.objects.get(
                    id=int(product_item["product_specif_id"]),
                )

            else:
                product_spes = ProductSpecification(
                    specification=specification,
                    product=product,
                )

            product_spes.price_exclusive = product_item["price_exclusive"]
            product_spes.product_currency = price_okt.currency
            product_spes.quantity = product_item["quantity"]
            product_spes.price_all = price_all
            product_spes.price_one = price_one

            if (
                product_item["extra_discount"] != "0"
                and product_item["extra_discount"] != ""
                and product_item["extra_discount"] != 0
            ):
                product_spes.extra_discount = product_item["extra_discount"]
            else:
                product_spes.extra_discount = None

            product_spes.price_one_motrum = price_one_motrum
            product_spes.price_all_motrum = price_all_motrum
            product_spes._change_reason = "Ручное"
            product_spes.comment = product_item["comment"]
            product_spes.id_cart_id = int(product_item["id_cart"])
            product_spes.product_price_catalog = product_price_catalog

            # запись дат
            date_delivery = product_item["date_delivery"]
            if date_delivery != "" and date_delivery != None:
                product_spes.date_delivery = datetime.datetime.strptime(
                    date_delivery, "%Y-%m-%d"
                )
                product_spes.date_delivery = date_delivery

            text_delivery = product_item["text_delivery"]
            if text_delivery != "" and text_delivery != None:
                product_spes.text_delivery = text_delivery

            item_comm = product_item["comment"]
            if item_comm != "" and item_comm != None:
                product_spes.text_delivery = item_comm

            product_spes.save()

            total_amount = total_amount + price_all
            total_amount_motrum = total_amount_motrum + price_all_motrum

        # продукты без записи в окт
        else:
            print("продукты без записи в окт", product_item)

            price_one = product_item["price_one"]
            price_one_original_new = price_one
            if product_item["sale_motrum"]:
                motrum_sale = product_item["sale_motrum"]
                motrum_sale = motrum_sale.replace(".", "")
                motrum_sale = motrum_sale.replace(",", ".")
                motrum_sale = float(motrum_sale)
            else:
                motrum_sale = 0.00

            price_one_motrum = price_one - (price_one / 100 * motrum_sale)
            price_one_motrum = round(price_one_motrum, 2)
            price_all_motrum = float(price_one_motrum) * int(product_item["quantity"])
            price_all_motrum = round(price_all_motrum, 2)

            if (
                product_item["extra_discount"] != "0"
                and product_item["extra_discount"] != ""
                and product_item["extra_discount"] != 0
            ):

                persent_sale = float(product_item["extra_discount"])

                price_one_sale = price_one - (price_one / 100 * persent_sale)
                price_one = round(price_one_sale, 2)

            price_all = float(price_one) * int(product_item["quantity"])
            price_all = round(price_all, 2)

            currency = Currency.objects.get(words_code="RUB")

            if (
                product_item["product_specif_id"] != "None"
                and product_item["product_specif_id"] != None
            ):
                product_spes = ProductSpecification.objects.get(
                    id=int(product_item["product_specif_id"]),
                )

            else:
                product_spes = ProductSpecification(
                    specification=specification,
                    product=None,
                )

            if (
                product_item["extra_discount"] != "0"
                and product_item["extra_discount"] != ""
                and product_item["extra_discount"] != 0
            ):
                product_spes.extra_discount = product_item["extra_discount"]
            else:
                product_spes.extra_discount = None

            product_spes.price_exclusive = product_item["price_exclusive"]
            product_spes.product_currency = currency
            product_spes.quantity = product_item["quantity"]
            product_spes.price_all = price_all
            product_spes.price_one = price_one
            product_spes.price_one_original_new = price_one_original_new
            product_spes.sale_motrum = motrum_sale
            product_spes.price_one_motrum = price_one_motrum
            product_spes.price_all_motrum = price_all_motrum
            product_spes.product_new = product_item["product_name_new"]
            product_spes.product_new_article = product_item["product_new_article"]
            product_spes._change_reason = "Ручное"
            product_spes.comment = product_item["comment"]
            product_spes.vendor_id = int(product_item["vendor"])
            product_spes.supplier_id = int(product_item["supplier"])
            product_spes.id_cart_id = int(product_item["id_cart"])

            date_delivery = product_item["date_delivery"]
            if date_delivery != "" and date_delivery != None:
                product_spes.date_delivery = datetime.datetime.strptime(
                    date_delivery, "%Y-%m-%d"
                )
                product_spes.date_delivery = date_delivery
            text_delivery = product_item["text_delivery"]
            if text_delivery != "" and text_delivery != None:
                product_spes.text_delivery = text_delivery

            item_comm = product_item["comment"]
            if item_comm != "" and item_comm != None:
                product_spes.text_delivery = item_comm

            product_spes.save()

            total_amount = total_amount + price_all
            total_amount_motrum = total_amount_motrum + price_all_motrum

    # обновить спецификацию пдф
    total_amount = round(total_amount, 2)
    total_amount_motrum = round(total_amount_motrum, 2)

    marginality = ((total_amount - total_amount_motrum) / total_amount) * 100
    marginality = round(marginality, 2)
    specification.total_amount = total_amount
    specification.comment = specification_comment
    specification.date_delivery = date_delivery_all
    specification.id_bitrix = id_bitrix
    specification.marginality = marginality
    specification._change_reason = "Ручное"

    specification.save()
    if requisites.contract:
        if type_save == "new":
            specification_name = requisites.number_spec + 1
            requisites.number_spec = specification_name

        elif type_save == "update":
            specification_name = specification.number
        elif type_save == "hard_update":
            specification_name = specification.number
        else:
            specification_name = specification.number

        if specification_name == None:
            specification_name = requisites.number_spec + 1
            requisites.number_spec = specification_name

        pdf = crete_pdf_specification(
            specification.id,
            requisites,
            account_requisites,
            request,
            motrum_requisites,
            date_delivery_all,
            type_delivery,
            post_update,
            specification_name,
        )

        if pdf:
            specification.file = pdf
            specification._change_reason = "Ручное"

            if post_update == False:
                specification.date_create_pdf = datetime.datetime.today()

            specification.number = specification_name
            specification.save()
            requisites.save()

    else:
        specification_name = None
        specification.file = None
        specification.number = specification_name
        specification.save()

    return specification


def get_presale_discount(product):
    from apps.supplier.models import Discount

    supplier = product.supplier
    try:
        discount = Discount.objects.get(supplier=supplier, is_tag_pre_sale=True)
        return discount
    except Discount.DoesNotExist:
        return False


def transform_date(date):

    months = [
        "января",
        "февраля",
        "марта",
        "апреля",
        "мая",
        "июня",
        "июля",
        "августа",
        "сентября",
        "октября",
        "ноября",
        "декабря",
    ]
    (
        year,
        month,
        day,
    ) = date.split("-")
    return f"{day} {months[int(month) - 1]} {year} г."


def rub_words(n):

    n = str(n)

    if n[-2:] in ("11", "12", "13", "14"):
        return f"рублей"

    elif n[-1] == "1":
        return f"рубль"
    elif n[-1] in ("2", "3", "4"):
        return f"рубля"

    else:
        return f"рублей"


def pens_words(n):

    if n[-2:] in ("11", "12", "13", "14"):
        return f"копеек"

    elif n[-1] == "1":
        return f"копейка"

    elif n[-1] in ("2", "3", "4"):
        return f"копейки"

    else:
        return f"копеек"


# общая функция кешировани
def loc_mem_cache(key, function, timeout=300):
    cache_data = cache.get(key)
    if not cache_data:
        cache_data = function()
        cache.set(key, cache_data, timeout)
    return cache_data


# сохранение товара без записи в бд в бд
def save_new_product_okt(product_new):
    from apps.product.models import Product, Price, Lot, Stock
    from apps.supplier.models import Supplier, Vendor

    if product_new.vendor:
        vendor = product_new.vendor

    else:
        vendor = Vendor.objects.get("drugoj")

    if product_new.supplier:

        supplier = product_new.supplier
    else:

        supplier = Supplier.objects.get("drugoj")

    if product_new.product:
        product_new_prod = product_new.product.id
        product = Product.objects.get(id=product_new_prod)

        product.name = product_new.product_new

        product.save()

    else:
        product_new_prod = None
        product = Product(
            supplier=supplier,
            vendor=vendor,
            article_supplier=product_new.product_new_article,
            name=product_new.product_new,
            in_view_website=False,
            autosave_tag=False,
        )
        product.save()
        update_change_reason(product, "Автоматическое")

        currency = Currency.objects.get(words_code="RUB")
        vat = Vat.objects.get(name=20)

        price = Price(
            prod=product,
            currency=currency,
            vat=vat,
            extra_price=True,
            in_auto_sale=False,
        )
        price.save()
        update_change_reason(price, "Автоматическое")

        lot_auto = Lot.objects.get(name_shorts="шт")
        product_stock = Stock(
            prod=product,
            lot=lot_auto,
        )
        product_stock.save()

        product_new.product = product
        product_new.save()

    return product
    # try:
    #     product = Product.objects.get(
    #         vendor=vendor, id=product_new_prod
    #     )
    #     product.name = product_new.product_new
    #     product.save()

    # except Product.DoesNotExist:
    #     product = Product(
    #         supplier=supplier,
    #         vendor=vendor,
    #         article_supplier=product_new.product_new_article,
    #         name=product_new.product_new,
    #         in_view_website=False,
    #         autosave_tag=False,
    #     )
    #     product.save()
    #     update_change_reason(product, "Автоматическое")

    #     currency = Currency.objects.get(words_code="RUB")
    #     vat = Vat.objects.get(name=20)

    #     price = Price(
    #         prod=product,
    #         currency=currency,
    #         vat=vat,
    #         extra_price=True,
    #         in_auto_sale=False,
    #     )
    #     price.save()
    #     update_change_reason(price, "Автоматическое")

    #     lot_auto = Lot.objects.get(name_shorts="шт")
    #     product_stock = Stock(
    #         prod=product,
    #         lot=lot_auto,
    #     )
    #     product_stock.save()

    #     return product


def number_specification(type_save):
    from apps.specification.models import Specification

    if type_save == "specification":
        last_spec_name = Specification.objects.filter(number__isnull=False).last()

        if last_spec_name:
            last_spec_name = last_spec_name.number
            specification_name = int(last_spec_name) + 1
        else:
            specification_name = 1

    elif type_save == "bill":
        specification_name = None

    return specification_name


def save_spesif_web(cart, products_cart, extra_discount, requisites):
    from apps.specification.api.serializers import (
        SpecificationSerializer,
        ProductSpecificationSaveSerializer,
    )
    from apps.specification.models import Specification
    from apps.product.models import Product
    from apps.core.utils_web import get_product_item_data

    print("save_spesif_web")
    try:
        with transaction.atomic():
            serializer_class_specification = SpecificationSerializer
            data_stop = create_time_stop_specification()
            # specification_name = number_specification("specification")

            # TODO:сохранять номер если все ок
            # specification_name = requisites.number_spec + 1
            # requisites.number_spec = specification_name

            data_specification = {
                "cart": cart.id,
                "admin_creator": None,
                "id_bitrix": None,
                "date_stop": data_stop,
                # "number": specification_name,
                "tag_stop": True,
            }

            old_spesif = Specification.objects.filter(cart=cart.id)
            if old_spesif:
                old_spesif = old_spesif[0]
                data_specification = {
                    # "cart": cart.id,
                    "admin_creator": None,
                    "id_bitrix": None,
                    "date_stop": data_stop,
                    # "number": specification_name,
                    "tag_stop": True,
                }
                serializer = serializer_class_specification(
                    old_spesif, data=data_specification, partial=True
                )
            else:
                old_spesif = None
                serializer = serializer_class_specification(
                    data=data_specification, partial=True
                )

            print(data_specification)
            date_delivery_all = None

            if serializer.is_valid():
                print("serializer.is_valid()")
                # serializer._change_reason = "Клиент с сайта"
                serializer.skip_history_when_saving = True
                specification = serializer.save()
                print(serializer.data)

                if specification:

                    total_amount = 0.00
                    for product_item in products_cart:
                        quantity = product_item.quantity
                        product = Product.objects.get(id=product_item.product.id)
                        item_data = get_product_item_data(
                            specification,
                            product,
                            extra_discount,
                            quantity,
                            product_item,
                        )

                        serializer_class_specification_product = (
                            ProductSpecificationSaveSerializer
                        )
                        serializer_prod = serializer_class_specification_product(
                            data=item_data
                        )
                        if serializer_prod.is_valid():
                            serializer_prod._change_reason = "Клиент с сайта"
                            # serializer_prod.skip_history_when_saving = True
                            specification_product = serializer_prod.save()

                        else:

                            return ("error", serializer_prod.errors, None)
                        total_amount += float(item_data["price_all"])

                    # обновить спецификацию пдф
                    total_amount = round(total_amount, 2)
                    specification.total_amount = total_amount
                    specification.date_delivery = date_delivery_all
                    specification._change_reason = "Клиент с сайта"
                    # specification.skip_history_when_saving = True
                    specification.save()
                    if requisites and requisites.contract:
                        # if requisites.contract:

                        specification_name = requisites.number_spec + 1
                        requisites.number_spec = specification_name

                        # креаотить пдф
                        pdf = None

                        if pdf:
                            specification.number = specification_name
                            specification.save()
                            requisites.save()

                    else:
                        specification_name = None
                        specification.file = None
                        specification.number = specification_name
                        specification.save()
                    return ("ok", specification, 11111)
                else:
                    error = "error"
                    location = "Сохранение спецификации в корзине сайта specification"
                    info = f" ошибка specification"
                    e = error_alert(error, location, info)
                    return ("error", None, None)

            else:
                print("ERROE SER")
                print(serializer.errors)
                tr = traceback.format_exc()
                error = "error"
                location = "Сохранение спецификации в корзине сайта"
                info = f" ошибка {e}{tr}{serializer.errors}"

                e = error_alert(error, location, info)
                return ("error", None, None)

    except Exception as e:
        print(e)
        tr = traceback.format_exc()
        error = "error"
        location = "Сохранение спецификации в корзине сайта"
        info = f" ошибка {e}{tr}"

        e = error_alert(error, location, info)
        return ("error", None, None)


def save_order_web(request, data_order, all_info_requisites, all_info_product):
    from apps.client.api.serializers import OrderSerializer
    from apps.notifications.models import Notification
    from apps.client.models import Order

    serializer_class = OrderSerializer
    serializer = serializer_class(data=data_order, many=False)

    if serializer.is_valid():
        order = serializer.save()
        if all_info_requisites and all_info_product:
            # Notification.add_notification(order.id, "STATUS_ORDERING")
            # Notification.add_notification(order.id, "DOCUMENT_SPECIFICATION")

            bill_name = (
                Order.objects.filter(bill_name__isnull=False)
                .order_by("bill_name")
                .last()
            )
            if bill_name:
                bill_name = int(bill_name.bill_name) + 1
            else:
                bill_name = 1

            if order.requisites.contract:
                is_req = True
            else:
                is_req = False

            order_pdf = order.create_bill(request, is_req, order, bill_name, False)
        else:
            pass

        return ("ok", serializer.data)

        # return Response(serializer.data, status=status.HTTP_200_OK)
    else:
        return ("error", serializer.errors)
        # return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


def client_info_bitrix(data, company_adress):
    from apps.client.models import RequisitesAddress
    from apps.client.models import (
        AccountRequisites,
        Requisites,
        RequisitesOtherKpp,
        ClientRequisites,
    )
    from dateutil.parser import parse

    error = "info_error_order"
    location = "отправка client_info_bitrix "
    info = f"отправка client_info_bitrix {data}{company_adress}"
    e = error_alert(error, location, info)

    if data["contract_date"]:
        data_contract = parse(data["contract_date"]).date()
        # data_contract = datetime.datetime.strptime(
        #     data["contract_date"], "%Y-%B-%dT%HH:%MM:%SS-%HH:%MM"
        # ).date()
    else:
        data_contract = None

    if len(data["contact_bd_arr"]) > 0:
        pass
    client_req, client_req_created = Requisites.objects.update_or_create(
        # id_bitrix=data["id_company"],
        inn=data["inn"],
        defaults={
            "id_bitrix": data["req_id_bitrix"],
            "legal_entity": data["legal_entity"],
            "type_client": data["type_client"],
            "manager_id": int(data["manager"]),
            "first_name": data["first_name"],
            "last_name": data["last_name"],
            "middle_name": data["middle_name"],
            # "contract": data["contract"],
            # "contract_date": data_contract,
            # "id_bitrix": data["id_bitrix"],
        },
        create_defaults={
            "id_bitrix": data["req_id_bitrix"],
            "legal_entity": data["legal_entity"],
            "inn": data["inn"],
            "contract": data["contract"],
            "contract_date": data_contract,
            "type_client": data["type_client"],
            "manager_id": int(data["manager"]),
            "first_name": data["first_name"],
            "last_name": data["last_name"],
            "middle_name": data["middle_name"],
            # "id_bitrix": data["id_bitrix"],
        },
    )
    print("client_req, client_req_created", client_req, client_req_created)

    if data["contract"] != "" and client_req.contract != data["contract"]:
        client_req.contract = data["contract"]
        client_req.contract_date = data_contract
        client_req.number_spec = 0
        client_req.save()

    client_req_kpp, client_req_kpp_created = (
        RequisitesOtherKpp.objects.update_or_create(
            requisites=client_req,
            id_bitrix=data["id_bitrix"],
            # kpp=data["kpp"],
            defaults={
                "kpp": data["kpp"],
                "ogrn": data["ogrn"],
                "legal_post_code": data["legal_post_code"],
                "legal_city": data["legal_city"],
                "legal_address": data["legal_address"],
                "postal_post_code": data["postal_post_code"],
                "postal_city": data["postal_city"],
                "postal_address": data["postal_address"],
                "tel": data["tel"],
            },
        )
    )
    print(
        "client_req_kpp, client_req_kpp_created", client_req_kpp, client_req_kpp_created
    )

    if len(data["contact_bd_arr"]) > 0:
        for contact_bd in data["contact_bd_arr"]:
            client_req_ty, client_req_created_ty = (
                ClientRequisites.objects.update_or_create(
                    client_id=int(contact_bd), requisitesotherkpp=client_req_kpp
                )
            )

    req_adress_web = RequisitesAddress.objects.filter(
        requisitesKpp=client_req_kpp,
        type_address_bx="web-lk-adress",
    )
    type_address_bx_9 = False
    for company_bx_adress in company_adress:

        if company_bx_adress["region"]:
            pass
        else:
            company_bx_adress["region"] == None

        if company_bx_adress["province"]:
            pass
        else:
            company_bx_adress["province"] == None

        client_req_kpp_address, client_req_kpp_created_address = (
            RequisitesAddress.objects.update_or_create(
                requisitesKpp=client_req_kpp,
                type_address_bx=company_bx_adress["type_address_bx"],
                defaults={
                    "country": company_bx_adress["country"],
                    "post_code": company_bx_adress["post_code"],
                    "region": company_bx_adress["province"],
                    "province": company_bx_adress["region"],
                    "city": company_bx_adress["city"],
                    "address1": company_bx_adress["address1"],
                    "address2": company_bx_adress["address2"],
                },
            )
        )

        if req_adress_web.count() == 0 and company_bx_adress["type_address_bx"] == "9":
            client_req_kpp_address, client_req_kpp_created_address = (
                RequisitesAddress.objects.update_or_create(
                    requisitesKpp=client_req_kpp,
                    type_address_bx="web-lk-adress",
                    defaults={
                        "country": company_bx_adress["country"],
                        "post_code": company_bx_adress["post_code"],
                        "region": company_bx_adress["province"],
                        "province": company_bx_adress["region"],
                        "city": company_bx_adress["city"],
                        "address1": company_bx_adress["address1"],
                        "address2": company_bx_adress["address2"],
                    },
                )
            )
            type_address_bx_9 == True

        if (
            req_adress_web.count() == 0
            and company_bx_adress["type_address_bx"] == "6"
            and type_address_bx_9 == False
        ):

            client_req_kpp_address, client_req_kpp_created_address = (
                RequisitesAddress.objects.update_or_create(
                    requisitesKpp=client_req_kpp,
                    type_address_bx="web-lk-adress",
                    defaults={
                        "country": company_bx_adress["country"],
                        "post_code": company_bx_adress["post_code"],
                        "region": company_bx_adress["province"],
                        "province": company_bx_adress["region"],
                        "city": company_bx_adress["city"],
                        "address1": company_bx_adress["address1"],
                        "address2": company_bx_adress["address2"],
                    },
                )
            )

        print(
            "client_req_kpp_address, client_req_kpp_created_address",
            client_req_kpp_address,
            client_req_kpp_created_address,
        )

    acc_req, acc_req_created = AccountRequisites.objects.update_or_create(
        requisitesKpp=client_req_kpp,
        account_requisites=data["account_requisites"],
        defaults={
            "bank": data["bank"],
            "kpp": data["ks"],
            "bic": data["bic"],
        },
    )
    return (client_req, acc_req)


def send_requests(url, headers, data, auth):
    import requests
    import certifi

    if auth == "1c":
        print("auth1c")
        payload = data
        #
        url = os.environ.get("1S_URL")
        headers = {}
        auth = HTTPBasicAuth(os.environ.get("1S_LOGIN"), os.environ.get("1S_PASSWORD"))
        import ssl

        paths = ssl.get_default_verify_paths()
        certifi1 = certifi.where()

        response = requests.request(
            "POST",
            url,
            auth=auth,
            headers=headers,
            data=payload,
            allow_redirects=False,
            verify=False,
        )
        
        error = "info_error_order"
        location = "отправка requests 1c"
        info = f"отправка requests 1c {response} / {response.text}"
        e = error_alert(error, location, info)

        print(response.text)

    else:
        response = requests.post(url, headers=headers, data=data, json=data)

    if response.status_code != 200:
        error = "error"
        location = "отправка requests"
        info = f"отправка requests {response}{response.text}"
        e = error_alert(error, location, info)

    return response.status_code


def json_serial(obj):
    from datetime import date, datetime

    """JSON serializer for objects not serializable by default json code"""

    if isinstance(obj, (datetime, date)):
        return obj.isoformat()
    raise TypeError("Type %s not serializable" % type(obj))


def after_save_order_products(products):
    from apps.specification.models import ProductSpecification

    order_products = []

    for obj in products:
        prod = ProductSpecification.objects.get(id=obj["id"])

        if prod.product_new_article != None:
            new_prod_db = save_new_product_okt(prod)

        if prod.product.vendor:
            vendor = prod.product.vendor.name
        else:
            vendor = None

        if prod.product.promo_groupe:
            promo = prod.product.promo_groupe.name
        else:
            promo = ""
        
        product_name_str =  prod.product.name
                
        if vendor and  prod.product.supplier.slug == "prompower" and prod.product.description:
            product_name_str = prod.product.description

        data_prod_to_1c = {
            "vendor": vendor,
            "article": prod.product.article_supplier,
            "article_motrum": prod.product.article,
            # "name": prod.product.name,
            "name": product_name_str,
            
            "price_one": prod.price_one,
            "quantity": prod.quantity,
            "price_all": prod.price_all,
            "text_delivery": prod.text_delivery,
            "data_delivery": prod.date_delivery.isoformat(),
            "promo_group": promo,
        }

        order_products.append(data_prod_to_1c)
    return order_products


def create_info_request_order_bitrix(
    order, pdf, pdf_signed, document_specification, order_products
):
    from apps.product.models import CurrencyRate

    data_for_bitrix = {
        "name_bill": order.bill_name,
        "pdf": pdf,
        "pdf_signed": pdf_signed,
        "bill_date_create": order.bill_date_start,
        "document_specification": document_specification,
        "order_products": order_products,
        # "currency": {},
    }

    currency = []
    date_now = datetime.datetime.today()
    currency_rate = CurrencyRate.objects.filter(date=date_now)
    for currency_rate_item in currency_rate:

        data_curr = {
            "currency_name": currency_rate_item.currency.words_code,
            "currency_rate": currency_rate_item.vunit_rate,
            "currency_date": currency_rate_item.date.strftime("%d.%m.%Y"),
        }

        currency.append(data_curr)

    data_for_bitrix["currency"] = currency

    return data_for_bitrix


def create_info_request_order_1c(order, order_products):
    from apps.specification.models import Specification

    specifications = Specification.objects.get(id=order.specification.id)

    print(specifications)
    if order.manager:
        name_admin = f"{order.manager.last_name} {order.manager.first_name}"
        if order.manager.middle_name:
            name_admin = f"{order.manager.last_name} {order.manager.first_name} {order.manager.middle_name}"
    else:

        if specifications.admin_creator:
            name_admin = f"{specifications.admin_creator.last_name} {specifications.admin_creator.first_name}"
            if specifications.admin_creator.middle_name:
                name_admin = f"{specifications.admin_creator.last_name} {specifications.admin_creator.first_name} {specifications.admin_creator.middle_name}"
        else:
            name_admin = " "
    contract_date = order.requisites.contract_date
    if contract_date:
        contract_date = order.requisites.contract_date.isoformat()

    if order.account_requisites.requisitesKpp.kpp:
        kpp = order.account_requisites.requisitesKpp.kpp
    else:
        kpp = None

    data_for_1c = {
        "motrum_requisites": {
            "legal_entity": order.motrum_requisites.requisites.full_name_legal_entity,
        },
        "client": {
            "type": order.requisites.get_type_client(),
            # "id_bitrix": order.id_bitrix,
            # "legal_entity_motrum": None,
            "contract": order.requisites.contract,
            "contract_date": contract_date,
            "legal_entity": order.requisites.legal_entity,
            "inn": order.requisites.inn,
            "kpp": kpp,
            "ogrn": order.account_requisites.requisitesKpp.ogrn,
            "legal_post_code": order.account_requisites.requisitesKpp.legal_post_code,
            "legal_city": order.account_requisites.requisitesKpp.legal_city,
            "legal_address": order.account_requisites.requisitesKpp.legal_address,
            # "postal_post_code": order.account_requisites.requisitesKpp.postal_post_code,
            # "postal_city": order.account_requisites.requisitesKpp.postal_city,
            # "postal_address": order.account_requisites.requisitesKpp.postal_address,
            "tel": order.account_requisites.requisitesKpp.tel,
            "account_requisites": order.account_requisites.account_requisites,
            "bank": order.account_requisites.bank,
            "ks": order.account_requisites.kpp,
            "bic": order.account_requisites.bic,
        },
        "invoice_options": {
            "id_bitrix": order.id_bitrix,
            "delivery": order.type_delivery.text_long,
            # "type_invoice": "счет" if order.requisites.contract else "счет-оферта",
            "type_invoice": "счет",
            "number_invoice": order.bill_name,
            "data_invoice": order.bill_date_start.isoformat(),
            "prepay_persent": order.requisites.prepay_persent,
            "postpay_persent": order.requisites.postpay_persent,
            "manager_invoice": name_admin,
        },
        "order_products": order_products,
    }
    # if order.manager:
    #     data_for_1c["invoice_options"]["manager_invoice"] = (
    #         f"{order.manager.last_name}{order.manager.first_name}{order.manager.middle_name}",
    #     )
    # else:
    #     data_for_1c["invoice_options"]["manager_invoice"] = ""

    return data_for_1c


def image_error_check():
    from apps.product.models import ProductImage

    product_image = ProductImage.objects.all()
    for img in product_image:
        if img.photo == "" or img.photo == None:
            img.delete()
        else:
            dir_img = "{0}/{1}".format(MEDIA_ROOT, img.photo)

            if not os.path.exists(dir_img):
                img.delete()
            else:

                if os.stat(dir_img).st_size < 1000:
                    img.delete()


def product_cart_in_file(file, cart):
    from apps.product.models import Product, ProductCart
    from apps.supplier.models import Vendor, Supplier

    try:

        def _serch_prod_in_file(article, vendor):
            product = Product.objects.filter(
                article_supplier=article, vendor__slug=vendor
            )

            if product:
                pass
            else:
                # product = Product.objects.filter(article_supplier=article)
                # product = Product.objects.filter(name__icontains=article)
                product = Product.objects.filter(
                    Q(article_supplier=article) | Q(name__icontains=article)
                )

            if product.count() == 1:
                return (1, product[0])
            elif product.count() == 0:
                return (0, None)
            else:
                # slugish = translit.translify(vendor)
                # vendor_name = slugify(slugish)
                # product = product.filter(vendor__slug=vendor_name)

                return (product.count(), product)

        def _save_prod_to_cart(product_okt_count, product_okt, cart, data):
            product_price_client = float(data["product_price_client_no_nds"])
            sale_client = 100 - (float(data["sale_client"]) * 100)
            sale_motrum = 100 - (float(data["sale_motrum"]) * 100)
            product_price = product_price_client * (100 / (100 - sale_client))
            product_price_nds = product_price + (product_price / 100 * NDS)
            product_price = round(product_price_nds)

            quantity = int(data["quantity"])
            print("product_okt_count", product_okt_count)

            if product_okt_count == 1:
                # 100% ПОПАДАНИЕ
                vendor = product_okt.vendor
                supplier = product_okt.supplier
                print(
                    "product_okt_count, product_okt, cart, data",
                    product_okt_count,
                    product_okt,
                    cart,
                    data,
                )

                ProductCart.objects.update_or_create(
                    cart_id=cart,
                    product=product_okt,
                    defaults={
                        "product_price": product_price,
                        "product_sale_motrum": sale_motrum,
                        "quantity": quantity,
                        "supplier": supplier,
                        "vendor": vendor,
                        "sale_client": sale_client,
                        "product_sale_motrum": sale_motrum,
                        "tag_auto_document": "ONE",
                    },
                )
            elif product_okt_count > 1:
                # НЕСКОЛЬКО ВАРИАНТОВ
                prod = None
                for product_ok in product_okt:
                    prod = product_ok

                    if product_ok.article == data["article_file"]:
                        prod = product_ok

                vendor = prod.vendor
                supplier = prod.supplier
                ProductCart.objects.update_or_create(
                    cart_id=cart,
                    product=prod,
                    defaults={
                        "product_price": product_price,
                        "product_sale_motrum": sale_motrum,
                        "quantity": quantity,
                        "supplier": supplier,
                        "vendor": vendor,
                        "sale_client": sale_client,
                        "product_sale_motrum": sale_motrum,
                        "tag_auto_document": "MULTI",
                    },
                )

            else:
                # 0 НАХОДОК
                vendor = Vendor.objects.filter(slug=data["vendor_slug"])
                if vendor.count() == 1:
                    vendor = vendor[0]
                else:
                    currency_catalog = Currency.objects.get(words_code="RUB")
                    vat_catalog = Vat.objects.get(name=NDS)

                    vendor = Vendor.objects.create(
                        name=data["vendor_name"],
                        currency_catalog=currency_catalog,
                        vat_catalog=vat_catalog,
                    )

                supplier = Supplier.objects.get(slug="drugoj")
                # ищем в товарах сохданных в корзине товар если нет создаеи новый
                product_in_cart = ProductCart.objects.filter(
                    cart_id=cart,
                    product=None,
                    product_new_article=data["article_file"],
                    vendor=vendor,
                )
                if product_in_cart.count() == 1:
                    product_in_cart.update(
                        cart_id=cart,
                        product_new=data["article_file"],
                        product_new_price=product_price,
                        product_new_sale=sale_client,
                        product_new_sale_motrum=sale_motrum,
                        quantity=quantity,
                        sale_client=sale_client,
                        tag_auto_document="NONE",
                    )
                else:

                    ProductCart.objects.create(
                        product_new_article=data["article_file"],
                        product_new=data["article_file"],
                        cart_id=cart,
                        product=None,
                        product_new_price=product_price,
                        product_new_sale=sale_client,
                        product_new_sale_motrum=sale_motrum,
                        supplier=supplier,
                        quantity=quantity,
                        vendor=vendor,
                        sale_client=sale_client,
                        tag_auto_document="NONE",
                    )

        workbook = load_workbook(file, data_only=True)
        data_sheet = workbook.active

        first_row_in_prod = None
        last_row_in_prod = False

        for index in range(1, data_sheet.max_row):

            column_b = data_sheet.cell(row=index, column=2).value

            if first_row_in_prod == None:
                if column_b == "Производитель":
                    first_row_in_prod = index
            else:
                if last_row_in_prod == False:
                    vendor_file = data_sheet.cell(row=index, column=2).value
                    if vendor_file == None:
                        last_row_in_prod = True
                    else:

                        article_file = data_sheet.cell(row=index, column=3).value
                        product_price_file = data_sheet.cell(row=index, column=5).value

                        if product_price_file:
                            product_okt_count, product_okt = _serch_prod_in_file(
                                article_file, vendor_file
                            )
                            slugish = translit.translify(vendor_file)
                            vendor_name = slugify(slugish)
                            data = {
                                "article_file": article_file,
                                "product_price_client_no_nds": data_sheet.cell(
                                    row=index, column=5
                                ).value,
                                "quantity": data_sheet.cell(row=index, column=6).value,
                                "vendor_name": vendor_file,
                                "vendor_slug": vendor_name,
                                "sale_client": data_sheet.cell(
                                    row=index, column=21
                                ).value,
                                "sale_motrum": data_sheet.cell(
                                    row=index, column=22
                                ).value,
                            }

                            _save_prod_to_cart(
                                product_okt_count, product_okt, cart, data
                            )
                else:
                    pass
    except Exception as e:

        tr = traceback.format_exc()
        print(e, tr)
        error = "error"
        location = "Добавление товаров из фаила"
        info = f"Добавление товаров из фаила{e}{tr}"
        e = error_alert(error, location, info)


def vendor_delta_optimus_after_load():
    from apps.product.models import Product

    # try:
    #     product = Product.objects.filter(supplier__slug__in=['delta', 'optimus-drive'],vendor__isnull=True)
    #     for product_one in product:
    #         print(product)
    #         if product_one.group_supplier is not None:
    #             if product_one.group_supplier.vendor is not None:
    #                 product_one.vendor = product_one.group_supplier.vendor
    #         product_one._change_reason = "Автоматическое"
    #         product_one.save()
    # except Exception as e:
    #     tr = traceback.format_exc()
    #     error = "file_error"
    #     location = "Загрузка 111 Delta"

    #     info = f"Загрузка 111 Delta{e}{tr}"
    #     e = error_alert(error, location, info)
    def background_task():
        product = Product.objects.filter(
            supplier__slug__in=["delta", "optimus-drive"], vendor__isnull=True
        )
        for product_one in product:
            print(product)
            if product_one.group_supplier is not None:
                if product_one.group_supplier.vendor is not None:
                    product_one.vendor = product_one.group_supplier.vendor
            product_one._change_reason = "Автоматическое"
            product_one.save()

    daemon_thread = threading.Thread(target=background_task)
    daemon_thread.setDaemon(True)
    daemon_thread.start()


def save_info_bitrix_after_web(data, req):
    from apps.client.models import RequisitesAddress
    from apps.client.models import AccountRequisites, Requisites, RequisitesOtherKpp
    from dateutil.parser import parse
    from apps.user.models import AdminUser

    if data["contract_date"]:
        data_contract = parse(data["contract_date"]).date()
        # data_contract = datetime.datetime.strptime(
        #     data["contract_date"], "%Y-%B-%dT%HH:%MM:%SS-%HH:%MM"
        # ).date()
    else:
        data_contract = None
    print(data)
    id_req_bx = data["req_bx_id"]
    manager = AdminUser.objects.filter(bitrix_id=int(data["manager"]))
    r = Requisites.objects.get(id=data["id_req"])
    if data["contract"]:
        r.contract = data["contract"]
        r.contract_date = data_contract
    if manager:
        r.manager = manager[0]

    r.id_bitrix = id_req_bx
    r.save()
    print(r)
    # client_req = Requisites.objects.filter(id=data["id_req"]).update(
    #     contract_date=data_contract, manager= manager ,contract= data["contract"],)
    # print(client_req)


def delete_everything_in_folder(folder_path):
    shutil.rmtree(folder_path)
    os.mkdir(folder_path)


# save_props_etim


def get_etim_prors_iek(prop_list, article):
    from apps.product.models import ProductProperty

    for item_prop in prop_list:
        if len(item_prop) > 0:
            pass_item = False
            name = item_prop["Attribute"]
            value = item_prop["value"]
            unit_measure = None
            names = name.split("_")

            if len(names) > 1:
                name = names[0]
                if names[1] == "Code":
                    pass_item = True

            if "unit" in item_prop:
                unit_measure = item_prop["unit"]
                if unit_measure != None:
                    name = f"{name} {unit_measure}"

            prop = ProductProperty(
                product=article,
                name=name,
                value=value,
                hide=False,
                unit_measure=unit_measure,
            )

            prop.save()
            update_change_reason(prop, "Автоматическое")


def send_lemur_form(data, request):
    #     data = {
    #         "site": DOMIAN,
    #         "form":name-form,
    #         "name":Имя с формы,
    #         "email":Email с формы,
    #         "phone":phone  с формы,
    #         "message":message   с формы,
    #         "link":link * Ссылка с какой страницы пришло сообщение,
    #     }
    headers = {}
    response = requests.request(
        "POST",
        FORM_LEMUR,
        headers=headers,
        data=data,
        allow_redirects=False,
        verify=False,
    )


def add_new_photo_adress_prompower():
    from apps.product.models import Product, ProductDocument
    from apps.supplier.models import (
        Supplier,
        Vendor,
    )

    prompower = Supplier.objects.get(slug="prompower")
    base_adress = "https://prompower.ru"
    vendori = Vendor.objects.get(slug="prompower")
    vendor_item = vendori

    url = "https://prompower.ru/api/prod/getProducts"
    payload = json.dumps(
        {
            "email": os.environ.get("PROMPOWER_API_EMAIL"),
            "key": os.environ.get("PROMPOWER_API_KEY"),
        }
    )
    headers = {
        "Content-type": "application/json",
        "Cookie": "nuxt-session-id=s%3Anp9ngMJIwPPIJnpKt1Xow9DA50eUD5OQ.IwH2nwSHFODHMKNUx%2FJRYeOVF9phtKXSV6dg6QQebAU",
    }
    response = requests.request("POST", url, headers=headers, data=payload)
    data = response.json()
    # old_doc =ProductDocument.objects.filter(link__contains="https://prompower.ru/")
    # for o_doc in old_doc:
    #     url = o_doc.link
    #     response = requests.get(url)
    #     if response.status_code == 404:
    #         o_doc.delete()

    for data_item in data:

        if data_item["article"] != None:

            def del_doc(product):
                directory = f"{MEDIA_ROOT}/products/prompower/prompower/{product.article_supplier}/document"
                print(directory)
                if not os.path.exists(directory):
                    print("directory_NOT")
                else:
                    print("YES", product.article_supplier)
                    delete_everything_in_folder(directory)

            def save_document(categ, product):
                base_dir = "products"
                path_name = "document_group"
                base_dir_supplier = product.supplier.slug
                base_dir_vendor = product.vendor.slug
                doc_list = data_item["cad"]
                if len(doc_list) > 0:
                    print("doc_list", doc_list)
                    for doc_item_individual in doc_list:
                        print("for")
                        if (
                            doc_item_individual["filename"] != ""
                            or doc_item_individual["filename"] != None
                        ):
                            print("doc_item_individual == None", product)
                            img = f"{base_adress}/catalog/CAD/{doc_item_individual["filename"]}"
                            image = ProductDocument.objects.create(product=article)
                            update_change_reason(image, "Автоматическое")
                            image_path = get_file_path_add(image, img)
                            print(image_path)
                            p = save_file_product(img, image_path)
                            image.photo = image_path
                            image.link = img
                            image.document = image_path
                            image.link = img

                            if (
                                doc_item_individual["title"] == ""
                                and doc_item_individual["title"] == None
                            ):
                                name = doc_item_individual["filename"]
                                images_last_list = name.split(".")
                                name = images_last_list[0]
                                print(name)
                                image.name = name
                            else:
                                name = doc_item_individual["title"]
                                print(name)
                                image.name = name

                            image.type_doc = "Other"
                            image.save()
                            print(image)
                            print(image.document)
                            update_change_reason(image, "Автоматическое")

            article_suppliers = data_item["article"]
            try:
                article = Product.objects.get(
                    supplier=prompower,
                    vendor=vendori,
                    article_supplier=article_suppliers,
                )
                del_doc(article)
                save_document(None, article)
            except Exception as e:
                print(e)
                # print("Нет такого твоара",article_suppliers)


def serch_products_web(search_text, queryset):
    print("queryset1", queryset)
    search_input = search_text
    search_input = search_input.replace(",", "")
    search_input = search_input.split()
    print("search_input", search_input)
    queryset = queryset.filter(
        Q(name__icontains=search_input[0])
        | Q(article_supplier__icontains=search_input[0])
    )
    if len(search_input) > 1:
        for search_item in search_input[1:]:
            queryset = queryset.filter(
                Q(name__icontains=search_item)
                | Q(article_supplier__icontains=search_item)
            )
    print("queryset", queryset)
    return queryset


def create_file_props_in_vendor_props():
    from apps.product.models import (
        Product,
        ProductProperty,
    )

    new_dir = "{0}/{1}".format(MEDIA_ROOT, "props_file")
    # path_delta = f"{new_dir}/delta.csv"
    # path_emas = f"{new_dir}/emas.csv"
    path_iek = f"{new_dir}/iek.csv"
    # path_optimus = f"{new_dir}/optimus.csv"
    # path_prompower = f"{new_dir}/prompower.csv"

    fieldnames_nomenclature_written = [
        "Частота упоминания",
        "Название характеристики",
        "Варианты значений",
        "Единица измерения",
    ]

    props = []

    all_product_supplier = Product.objects.filter(supplier__slug="iek")
    i = 0

    def key_val_upd(key, val, value, unit_measure):

        for val_i in val:
            print(val_i)
            val_v = val_i.get("value")
            val_u = val_i.get("unit_measure")
            val_c = val_i.get("count")
            if val_c:
                print(val_c)
                val_i["count"] = val_c + 1

            if val_v:
                if value not in val_v:
                    val_v.append(value)
            if val_u:
                if unit_measure != None:
                    if unit_measure not in val_u:
                        val_u.append(f"{unit_measure}")

    def if_name(name, value, unit_measure):
        print("startprops", props)
        need_add_name = True
        for prop_arr_name in props:
            print(prop_arr_name)

            for key, val in prop_arr_name.items():
                print("key,val", key, val)
                if name == key:
                    print("name == key")
                    need_add_name = False
                    key_val_upd(key, val, value, unit_measure)

            print("need_add_name", need_add_name)
        if need_add_name:
            props_new = {
                name: [
                    {"value": [value]},
                    {"unit_measure": [unit_measure]},
                    {"count": 1},
                ]
            }

            print("props_new", props_new)
            props.append(props_new)

    for prod in all_product_supplier:
        i += 1
        props_prod = ProductProperty.objects.filter(product=prod)
        if props_prod.count() > 0:
            for prop_prod in props_prod:
                print(prop_prod)
                name = if_name(prop_prod.name, prop_prod.value, prop_prod.unit_measure)

    print(props)

    props_count = len(props)
    print("props_count", props_count)
    with open(path_iek, "w", encoding="UTF-8") as writerFile:
        writer_nomenk = csv.DictWriter(
            writerFile, delimiter=";", fieldnames=fieldnames_nomenclature_written
        )
        for i in range(0, props_count):
            row = {}
            if i == 0:

                row["Частота упоминания"] = "Частота упоминания"
                row["Название характеристики"] = "Название характеристики"
                row["Варианты значений"] = "Варианты значений"
                row["Единица измерения"] = "Единица измерения"

                writer_nomenk.writerow(row)

            print(i)
            prop = props[i]
            print(prop)
            for key, value in prop.items():
                print(key, value)
                # Глубина (мм) [{'value': ['69.6']}, {'unit_measure': [None]}]

                row["Название характеристики"] = key
                for val_i in value:
                    val_v = val_i.get("value")
                    val_u = val_i.get("unit_measure")
                    val_c = val_i.get("count")
                    if val_c:
                        row["Частота упоминания"] = val_c
                    if val_v:
                        val_v_i = "|| ".join(val_v)
                        row["Варианты значений"] = val_v_i
                    if val_u:
                        if val_u != [None]:
                            val_u_i = "|| ".join(val_u)
                            row["Единица измерения"] = val_u_i
                writer_nomenk.writerow(row)


def email_manager_after_new_order_site(order):
    try:
        from django.template import loader

        id_bitrix = order.id_bitrix
        manager_client = order.manager.email
        phone_client = order.client.phone
        url_bitrix_deal = f"https://pmn.bitrix24.ru/crm/deal/details/{id_bitrix}/"
        html_message = loader.render_to_string(
            "core/emails/email_manager_neworder.html",
            {
                "id_bitrix": id_bitrix,
                "url_bitrix_deal": url_bitrix_deal,
                "phone": phone_client,
            },
        )
        subject = "Новый заказа с сайта"
        to_email = manager_client

        print(to_email)
        sending_result = send_email_message_html(
            subject, None, to_email, html_message=html_message
        )

    except Exception as e:
        tr = traceback.format_exc()
        error = "error"
        location = "Отправка оповещения менеджеру после создания заказа с сайта"
        info = f"order.id = {order.id} {e}{tr}"
        e = error_alert(error, location, info)
        return ("error", info)
