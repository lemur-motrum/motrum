import datetime
from locale import currency
import os
import re
from unicodedata import category
import openpyxl as openxl
from openpyxl import Workbook
from openpyxl import load_workbook

from apps import supplier
from apps.core.models import Currency, Vat
from simple_history.utils import update_change_reason

from apps.core.utils import create_article_motrum
from apps.logs.utils import error_alert
from apps.product.models import Lot, Price, Product, Stock
from apps.supplier.models import (
    Supplier,
    SupplierCategoryProduct,
    SupplierCategoryProductAll,
    Vendor,
)
from project.settings import BASE_DIR, MEDIA_ROOT, NDS


def get_category_avangard(name, supplier, vendor):

    try:
        category_item = SupplierCategoryProduct.objects.get(
            supplier=supplier, vendor=vendor, name=category
        )
        print(category_item)
        return category_item

    except SupplierCategoryProduct.DoesNotExist:
        category_item = SupplierCategoryProduct(
            name=category,
            supplier=supplier,
            vendor=vendor,
        )
        category_item.save()
        print(category_item)
        return category_item


def get_price_avangard(vendor, supplier, article, price_supplier, category_item):
    from apps.core.utils import get_price_supplier_rub, get_category

    currency = Currency.objects.get(words_code="CNY")
    vat_include = True
    vat = Vat.objects.get(name=NDS)

    try:
        price_product = Price.objects.get(prod=article)

    except Price.DoesNotExist:
        price_product = Price(prod=article)

    finally:
        if price_supplier == "снят с производства":
            price_product.currency = currency
            price_product.vat = vat
            price_product.vat_include = True
            price_product.extra_price = True
            price_product._change_reason = "Автоматическое"
            price_product.save()

            article.in_view_website = False
            article.save()
        else:
            price_product.currency = currency
            price_product.price_supplier = price_supplier
            price_product.vat = vat
            price_product.vat_include = True
            price_product.extra_price = False
            price_product._change_reason = "Автоматическое"
            price_product.save()
            # update_change_reason(price_product, "Автоматическое")


def get_avangard_file(new_file, obj):
    try:
        supplier = Supplier.objects.get(slug="avangard")
        vendor = Vendor.objects.get( slug="odot")
        vat = Vat.objects.get(name=NDS)
        currency = Currency.objects.get(words_code="CNY")

        file_path = "{0}/{1}".format(MEDIA_ROOT, new_file)
        excel_doc = openxl.open(filename=file_path, data_only=True)

        sheetnames = excel_doc.sheetnames
        sheet = excel_doc[sheetnames[0]]
        first = 0
        category = ""
        category_item = None
        for i in range(sheet.min_row, sheet.max_row + 1):

            if first == 0:
                if sheet[f"A{i}"].value == "P/N\nАртикул":
                    first = i

            elif first != 0:
                # запись категорий

                if (
                    sheet[f"C{i}"].value is None
                    and sheet[f"b{i}"].value is None
                    and sheet[f"a{i}"].value is not None
                ):
                    # print(sheet[f"b{i}"].value)
                    category = sheet[f"A{i}"].value
                    # print(category)
                    # category_item = get_category_avangard(category, supplier, vendor)
                    # print(category_item)
                    try:
                        category_item = SupplierCategoryProduct.objects.get(
                            supplier=supplier, vendor=vendor, name=category
                        )
                    except SupplierCategoryProduct.DoesNotExist:
                        category_item = SupplierCategoryProduct(
                            name=category,
                            supplier=supplier,
                            vendor=vendor,
                        )
                        category_item.save()
                   
                    print(category_item)
                elif (
                    sheet[f"C{i}"].value is None
                    and sheet[f"b{i}"].value is None
                    and sheet[f"a{i}"].value is None
                ):
                    category_item = None
                # запись товаров
                elif (
                    sheet[f"a{i}"].value is not None
                    and sheet[f"b{i}"].value is not None
                    and sheet[f"c{i}"].value is not None
                ):
                    article_supplier = sheet[f"A{i}"].value
                    price_supplier_noval = sheet[f"C{i}"].value
                    
                    
                    if price_supplier_noval in (None, ""):
                        price_supplier = 0
                    elif isinstance(price_supplier_noval, (int, float)):
                        price_supplier = float(price_supplier_noval)
                    elif str(price_supplier_noval).strip().lower() in (
                        "снят с производства!",
                        "снят с производства",
                    ):
                        price_supplier = "снят с производства"
                    else:
                        price_supplier_sub = re.sub(
                            "[^0-9.,]", "", str(price_supplier_noval)
                        ).replace(",", ".")
                        price_supplier = float(price_supplier_sub) if price_supplier_sub else 0
                        
                        
                    name = sheet[f"B{i}"].value

                    try:
                        article = Product.objects.get(
                            supplier=supplier,
                            vendor=vendor,
                            article_supplier=article_supplier,
                        )

                    except Product.DoesNotExist:
                        new_article = create_article_motrum(supplier.id)
                        article = Product(
                            article=new_article,
                            supplier=supplier,
                            vendor=vendor,
                            article_supplier=article_supplier,
                            name=name,
                            description=None,
                            category_supplier_all=None,
                            group_supplier=None,
                            category_supplier=category_item,
                        )

                        article.save()
                        update_change_reason(article, "Автоматическое")
                    
                        
                    price = get_price_avangard(
                        vendor, supplier, article, price_supplier, category_item
                    )
                

                    try:
                        stock_prod = Stock.objects.get(prod=article)
                    except Stock.DoesNotExist:
                        stock_supplier = None
                        lot = Lot.objects.get(name="штука")

                        stock_prod = Stock(
                            prod=article,
                            lot=lot,
                            stock_supplier=stock_supplier,
                            data_update=datetime.datetime.now(),
                        )
                        stock_prod._change_reason = "Автоматическое"
                        stock_prod.save()

                # обработка ошибок считыввания
                else:
                    error = "file structure"
                    location = f"Загрузка фаилов авангард строка:{i}"
                    info = "Фаил не соответствует обрабатываемой структуре фаила-считывание фаила невозможно"
                    e = error_alert(error, location, info)

        if first == 0:
            error = "file structure"
            location = "Загрузка фаилов авангард"
            info = "Фаил не соответствует обрабатываемой структуре фаила-считывание фаила невозможно"
            e = error_alert(error, location, info)

    except Exception as e:
        print(e)
        error = "file_error"
        location = "Загрузка фаилов Avangard"

        info = f"ошибка при чтении фаила"
        e = error_alert(error, location, info)
