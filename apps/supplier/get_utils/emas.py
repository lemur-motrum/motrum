import datetime
import re
import openpyxl as openxl
from simple_history.utils import update_change_reason

from apps.core.models import Currency, Vat
from apps.core.utils import (
    create_article_motrum,
    get_category_emas,
    get_file_path_add,
    lot_chek,
    save_file_emas_product,
)
from apps.logs.utils import error_alert
from apps.product.models import (
    Price,
    Product,
    ProductImage,
    ProductProperty,
    Stock,
)
from apps.supplier.models import (
    Supplier,
    SupplierCategoryProduct,
    SupplierCategoryProductAll,
    SupplierGroupProduct,
    Vendor,
)
from project.settings import  MEDIA_ROOT


def add_file_emas(new_file, obj):
    try:
    
        supplier = Supplier.objects.get(slug="emas")
        vendor = None
        vat = None
        currency = Currency.objects.get(words_code="RUB")

        # Разбираем фаил
        # file_path = os.path.join(BASE_DIR, "tmp/emas.xlsx")
        file_path = "{0}/{1}".format(MEDIA_ROOT, new_file)
        excel_doc = openxl.open(filename=file_path, data_only=True)
        sheetnames = excel_doc.sheetnames  # Получение списка листов книги
        sheet = excel_doc[sheetnames[0]]

        first = 0
        # разбираем строки
        for index in range(sheet.min_row, sheet.max_row + 1):
            # считываем артикулы после появления записи артикул в первом стобце
            if first == 0:
                if sheet[f"A{index}"].value == "Артикул":
                    first = index

            elif first != 0:
                if (
                    sheet[f"B{index}"].value is None
                    and sheet[f"D{index}"].value is None
                ):
                    try:
                        a = sheet[f"A{index}"].value
                        article_suppliers = sheet[f"A{index}"].value
                        name = sheet[f"C{index}"].value
                        count = sheet[f"F{index}"].value
                        lot_item = sheet[f"E{index}"].value
                        lot = lot_chek(lot_item)

                        try:
                            article = Product.objects.get(
                                supplier=supplier, article_supplier=article_suppliers
                            )
                        
                        except Product.DoesNotExist:
                            vendor_qs = Vendor.objects.get(slug="drugoe")
                            new_article = create_article_motrum(supplier.id)
                            article = Product(
                                article=new_article,
                                supplier=supplier,
                                vendor=vendor_qs,
                                article_supplier=article_suppliers,
                                name=name,
                                description=None,
                                category_supplier_all=None,
                                group_supplier=None,
                                category_supplier=None,
                            )
                            article.save()
                            update_change_reason(article, "Автоматическое")
                            
                        try:
                            stock_prod = Stock.objects.get(prod=article)

                        except Stock.DoesNotExist:
                            stock_prod = Stock(
                                prod=article,
                                lot=lot,
                                lot_complect=1,
                            )

                        finally:
                            stock_prod.data_update = datetime.datetime.now()
                            stock_prod.stock_supplier = count
                            stock_prod._change_reason = 'Автоматическое'
                            stock_prod.save()
                            # print(stock_prod)
                            # update_change_reason(stock_prod, "Автоматическое")
                            
                        # цены товара
                        vat_catalog = Vat.objects.get(name=20)
                        currency = Currency.objects.get(words_code="RUB")
                        price_supplier = None
                        try:
                            price_product = Price.objects.get(prod=article)

                        except Price.DoesNotExist:
                            price_product = Price(prod=article)

                        finally:
                            price_product.currency = currency
                            price_product.price_supplier = price_supplier
                            price_product.vat = vat_catalog
                            price_product.vat_include = True
                            price_product.extra_price = True
                            price_product._change_reason = "Автоматическое"
                            price_product.save()    
                            
                            
                    except Exception as e:
                        print(e)
                        error = "file_api_error"
                        location = "Загрузка фаилов EMAS"
                        info = f"ошибка при чтении товара артикул: {article_suppliers} Тип ошибки:{e}"
                        e = error_alert(error, location, info)
                    finally:
                        continue

                else:
                    error = "file structure"
                    location = "Загрузка фаилов Emas"
                    info = f"Ошибка структуры в строке {index}. Фаил не соответствует обрабатываемой структуре фаила-считывание фаила невозможно"
                    e = error_alert(error, location, info)
                    
        if first == 0:
            error = "file structure"
            location = "Загрузка фаилов Emas"
            info = "Фаил не соответствует обрабатываемой структуре фаила-считывание фаила невозможно"
            e = error_alert(error, location, info)

    except Exception as e:
        print(e)
        error = "file_error"
        location = "Загрузка фаилов emas"

        info = f"ошибка при чтении фаила"
        e = error_alert(error, location, info)


def add_group_emas():
    from bs4 import BeautifulSoup

    try:
        supplier = Supplier.objects.get(slug="emas")
        vendor = Vendor.objects.get(slug="emas")

        path = f"{MEDIA_ROOT}/price/emas_site/emas.xml"

        with open(f"{MEDIA_ROOT}/price/emas_site/emas.xml", "r") as f:
            file = f.read()

        soup = BeautifulSoup(file, "xml")
        # classes = soup.find("Группы",)
        # for class_item in classes:
        class_groupe = soup.Классификатор.Группы.find_all("Группа", recursive=False)
        for class_item in class_groupe:
            # Первый уровень категории
            categ_name = class_item.find("Наименование", recursive=False)
            categ_id = class_item.find("Ид", recursive=False)
            categ_name_text = categ_name.get_text()
            categ_id_text = categ_id.get_text()

            try:
                categ = SupplierCategoryProduct.objects.get(
                    supplier=supplier,
                    name=categ_name_text,
                    article_name=categ_id_text,
                )
               

            except SupplierCategoryProduct.DoesNotExist:
                categ = SupplierCategoryProduct(
                    name=categ_name_text,
                    article_name=categ_id_text,
                    supplier=supplier,
                    autosave_tag=True,
                    vendor=vendor,
                )
                categ.save()

            # второй уровень -группы
            groupe = class_item.Группы.find_all("Группа", recursive=False)

            for group_item in groupe:
                group_name = group_item.find("Наименование", recursive=False)
                group_id = group_item.find("Ид", recursive=False)
                group_name_text = group_name.get_text()
                group_id_text = group_id.get_text()
                all_groupe = group_item.Группы.find_all("Группа", recursive=False)

                try:
                    groupe = SupplierGroupProduct.objects.get(
                        supplier=supplier,
                        name=group_name_text,
                        article_name=group_id_text,
                    )

                except SupplierGroupProduct.DoesNotExist:
                    groupe = SupplierGroupProduct(
                        name=group_name_text,
                        article_name=group_id_text,
                        supplier=supplier,
                        autosave_tag=True,
                        vendor=vendor,
                        category_supplier=categ,
                    )
                    groupe.save()

                for all_groupe_item in all_groupe:
                    all_groupe_name = all_groupe_item.find("Наименование")
                    all_groupe_id = all_groupe_item.find("Ид", recursive=False)
                    all_groupe_name_text = all_groupe_name.get_text()
                    all_groupe_id_text = all_groupe_id.get_text()

                    try:
                        all_groupe = SupplierCategoryProductAll.objects.get(
                            supplier=supplier,
                            name=all_groupe_name_text,
                            article_name=all_groupe_id_text,   
                        )
                        
                    except SupplierCategoryProductAll.DoesNotExist:
                        all_groupe = SupplierCategoryProductAll(
                            name=all_groupe_name_text,
                            article_name=all_groupe_id_text,
                            supplier=supplier,
                            autosave_tag=True,
                            vendor=vendor,
                            category_supplier=categ,
                            group_supplier=groupe,
                        )
                        
                        all_groupe.save()
        
        add_props_emas_product()
    except Exception as e:

        print(e)
        error = "file_error"
        location = "Загрузка фаилов emas"
        info = f"ошибка при чтении фаила Загрузка Групп"
        e = error_alert(error, location, info)


def add_props_emas_product():
    from bs4 import BeautifulSoup
    print("9090")
    try:
        print("0909")
        supplier = Supplier.objects.get(slug="emas")
        vendor = Vendor.objects.get(slug="emas")
        product = Product.objects.filter(supplier=supplier)

        path = f"{MEDIA_ROOT}/price/emas_site/emas.xml"
     
        with open(f"{MEDIA_ROOT}/price/emas_site/emas.xml", "r") as f:
            file = f.read()

        soup = BeautifulSoup(file, "xml")
   
        # перебор товаров
        non_props = 0
        yes_props = 0
        for product_item in product:
            print("121")

            name_art = product_item.article_supplier
            product_soup = soup.Товары.find_all("Значение", string=name_art)
            if product_soup == []:

                non_props += 1
            else:
                yes_props += 1
        
                for product_soup_items in product_soup:
                    parent_product_soup = product_soup_items.parent.parent.parent
                    if parent_product_soup.Группы is not None:
                        parent_product_soup_group = parent_product_soup.Группы.Ид
                        if parent_product_soup_group is not None:
                            text = parent_product_soup.ЗначенияСвойств.find(
                                "Ид", string="CML2_PREVIEW_TEXT"
                            )
                      
                            description_bs_parent = text.parent.Значение.get_text()
                      
                            if description_bs_parent != "":
                                item_product_bs = parent_product_soup
                            else:
                                item_product_bs = parent_product_soup
                            
                groupe_bs = item_product_bs.Группы.Ид.get_text()
                groupe_items = get_category_emas(supplier, groupe_bs)

                image_old = ProductImage.objects.filter(product=product_item).exists()
                if image_old == False:
                    # сохранение картинок
                    image_bs = item_product_bs.Картинка.get_text()
             
                    if image_bs != "":
                        image = ProductImage.objects.create(product=product_item)
                        update_change_reason(image, "Автоматическое")
                        image_path = get_file_path_add(image, image_bs)

                        p = save_file_emas_product(image_bs, image_path)
                        image.photo = image_path
                        image.link = image_bs
                        image.save()
                        update_change_reason(image, "Автоматическое")

                # сохранение пропсов и описаний из текста
                description_bs = item_product_bs.ЗначенияСвойств.find(
                    "Ид", string="CML2_PREVIEW_TEXT"
                )
                description_bs_parent = description_bs.parent.Значение.get_text()
                
                if description_bs_parent == "":
                    description_bs = item_product_bs.ЗначенияСвойств.find(
                    "Ид", string="CML2_DETAIL_TEXT"
                )
                    
                description_bs_parent = description_bs.parent.Значение.get_text()
                tds = []
                soup_desc = BeautifulSoup(description_bs_parent, "html.parser")

                quotes = soup_desc.find_all(
                    "tbody",
                )
                if len(quotes) == 1:
                

                    for div in quotes:
                        rows = div.findAll("tr")
                        for row in rows:
                            r = row.findAll("td")
                            if len(r) == 2:
                                tds.append(r)

                text_desc = ""
                # текст из р
                quotes_text = soup_desc.find_all("p")
                for p_text in quotes_text:
                    text = str(p_text).replace("<br>", ". ")
                    text = text.replace("</li>", ". ")
                    text = text.replace("<li>", "")
                    
                    text = re.sub(r"<[^>]+>", "", text, flags=re.S)
                    text = " ".join(text.split())
                    

                    if text != "":
                        if text_desc == "":
                            text_desc = text
                        else:
                            text_desc = f"{text_desc}{text}"

                quotes_li = soup_desc.find_all("li")
                for li_text in quotes_li:
                    
                   
                    text_li = str(li_text).replace("<br>", ". ")
                    text_li = text_li.replace("</li>", ". ")
                    text_li = text_li.replace("<li>", "")
                    
                    text_li = re.sub(r"<[^>]+>", "", text_li, flags=re.S)

                    if text_li != "":
                        if text_desc == "":
                            text_desc = text_li
                        else:
                            text_desc = f"{text_desc}{text_li}"
        
                print(soup_desc)
                if text_desc == "" and soup_desc != "":
                    text_desc = str(soup_desc)

                if text_desc == "":
                    text_desc = None

                if product_item.description == None or product_item.description == '' :
                    product_item.description = text_desc
                if product_item.category_supplier == None:
                    product_item.category_supplier_all = groupe_items[0]
                    product_item.group_supplier = groupe_items[1]
                    product_item.category_supplier = groupe_items[2]
                if product_item.vendor == None:
                    product_item.vendor = vendor

               
                product_item._change_reason = 'Автоматическое'
                product_item.save()
            
                # update_change_reason(product_item, "Автоматическое")
            

                props_old = ProductProperty.objects.filter(
                    product=product_item
                ).exists()
                if props_old == False:
                    # if props_old == False:

                    if tds != []:
                        for td in tds:
                            if td != []:
                                name_props = re.sub(
                                    r"<[^>]+>", "", str(td[0]), flags=re.S
                                )
                                name_props = " ".join(name_props.split())
                                value_props = re.sub(
                                    r"<[^>]+>", "", str(td[1]), flags=re.S
                                )
                                value_props = " ".join(value_props.split())
                                if name_props == "Характеристики":
                                    pass
                                else:
                                    props_product = ProductProperty(
                                        product=product_item
                                    )
                                    props_product.name = name_props

                                    props_product.value = value_props
                                    props_product.save()
                                    update_change_reason(
                                        props_product, "Автоматическое"
                                    )


    except Exception as e:
        print(e)
        error = "file_error"
        location = "Загрузка фаилов Delta"

        info = f"ошибка при чтении строки артикул"
        e = error_alert(error, location, info)
