import datetime
from itertools import count
import os
import traceback
from bs4 import BeautifulSoup
import openpyxl as openxl
from regex import E
import requests
from simple_history.utils import update_change_reason

from apps.core.models import Currency, Vat
from apps.core.utils import create_article_motrum, get_file_path_add, save_file_product
from apps.logs.utils import error_alert
from apps.product.models import (
    Lot,
    Price,
    Product,
    ProductDocument,
    ProductImage,
    ProductProperty,
    Stock,
)
from apps.supplier.models import (
    Supplier,
    SupplierCategoryProduct,
    SupplierGroupProduct,
    Vendor,
)
from project.settings import BASE_DIR, MEDIA_ROOT, NDS
from urllib.request import urlopen

def get_instart_file(new_file):
    print(999999999999999999)
    print(new_file)
    try:
  
        supplier = Supplier.objects.get(slug="instart")
        vendor = Vendor.objects.get(slug="instart")
        file_path = "{0}/{1}".format(MEDIA_ROOT, new_file)
        pars_instart_xlsx(file_path)
    except Exception as e:
        tr = traceback.format_exc()
        error = "file_error"
        location = "Загрузка фаилов Instart"

        info = f"ошибка при чтении фаила get_instart_file{tr}{e}"
        e = error_alert(error, location, info)
    
    
    
def pars_instart_xlsx(file_path):
    try:
        # file_path = "{0}/{1}/{2}/{3}".format(MEDIA_ROOT, "price", "instart",'instar.xlsx')
        # file_path = os.path.join(BASE_DIR, "media/price/instart/instar.xlsx")
        print(file_path)
        # media/price/instart/instar.xlsx
        excel_doc = openxl.open(filename=file_path, data_only=True)

        sheetnames = excel_doc.sheetnames
        sheet = excel_doc[sheetnames[0]]

        # базовые параметры
        supplier = Supplier.objects.get(slug="instart")
        vendor = Vendor.objects.get(slug="instart")
        currency = Currency.objects.get(words_code="RUB")
        vat = Vat.objects.get(name=NDS)
        lot_item = Lot.objects.get(name="штука")
        # sheet.max_row + 1)
        for index in range(6, sheet.max_row + 1):
            article_supplier = sheet[f"A{index}"].value
            print(article_supplier)
            try:
                if article_supplier:
                    # категория поставщика поиск создание
                    category_okt = None
                    groupe_okt = None
                    category = sheet[f"B{index}"].value

                    if category:
                        category_okt = SupplierCategoryProduct.objects.get_or_create(
                            supplier=supplier,
                            vendor=vendor,
                            name=category,
                        )
                        category_okt = category_okt[0]

                        # группа поставщика поиск созданеи
                        group = sheet[f"C{index}"].value
                        if group:
                            groupe_okt = SupplierGroupProduct.objects.get_or_create(
                                category_supplier=category_okt,
                                vendor=vendor,
                                supplier=supplier,
                                name=group,
                            )
                            groupe_okt = groupe_okt[0]
                    print(category_okt, groupe_okt)
                    model = sheet[f"D{index}"].value
                    name = sheet[f"S{index}"].value
                    description = sheet[f"T{index}"].value
                    try:
                        # поиск по товарам до интеграции
                        product = Product.objects.get(
                            vendor=vendor,
                            article_supplier=model,
                            additional_article_supplier=None,
                        )
                        print("1product", product)
                        _instart_product_upd(
                            product,
                            supplier=supplier,
                            additional_article_supplier=article_supplier,
                            category_supplier=category_okt,
                            group_supplier=groupe_okt,
                            description=description,
                            name=name,
                        )
                    except Product.DoesNotExist:
                        try:
                            # поиск по товарам после интеграции
                            product = Product.objects.get(
                                supplier=supplier,
                                vendor=vendor,
                                additional_article_supplier=article_supplier,
                            )
                            _instart_product_upd(
                                product,
                                supplier=supplier,
                                additional_article_supplier=article_supplier,
                                category_supplier=category_okt,
                                group_supplier=groupe_okt,
                                description=description,
                                name=name,
                            )
                        except Product.DoesNotExist:

                            new_article = create_article_motrum(supplier.id)
                            product = Product(
                                article=new_article,
                                supplier=supplier,
                                vendor=vendor,
                                article_supplier=model,
                                additional_article_supplier=article_supplier,
                                name=name,
                                description=description,
                                category_supplier=category_okt,
                                group_supplier=groupe_okt,
                            )
                            product.save()
                            update_change_reason(product, "Автоматическое")

                    # прайс для товара
                    # price_supplier = sheet[f"R{index}"].value
                    # is_float_price_supplier = _is_number_instart(price_supplier)
                    # if is_float_price_supplier:
                    #     price_supplier = float(price_supplier)
                    #     extra = False
                    # else:
                    #     price_supplier = None
                    #     extra = True
                    price_supplier = None
                    extra = True
                    try:
                        price_product = Price.objects.get(prod=product)

                    except Price.DoesNotExist:
                        price_product = Price(prod=product)
                        price_product.currency = currency
                        price_product.price_supplier = price_supplier
                        price_product.vat = vat
                        price_product.vat_include = True
                        price_product.extra_price = extra
                        price_product._change_reason = "Автоматическое"
                        price_product.save()

                    print("after price")
                

                    # остаток товара
                    try:
                        stock_prod = Stock.objects.get(prod=product)

                    except Stock.DoesNotExist:
                        stock_prod = Stock(
                            prod=product,
                            lot=lot_item,
                            lot_complect=1,
                        )
                        stock_prod._change_reason = "Автоматическое"
                        stock_prod.save()
                        
                    print("after stock_prod")
                    # картинкки товара
                    img_1 = sheet[f"AW{index}"].value
                    img_2 = sheet[f"AX{index}"].value
                    img_3 = sheet[f"AY{index}"].value
                    img_4 = sheet[f"AZ{index}"].value
                    img_5 = sheet[f"BA{index}"].value

                    image = ProductImage.objects.filter(product=product).exists()
                    if image == False:
                        _save_image_instart(product, img_1, img_2, img_3, img_4, img_5)
                    print("after image")
                    # документы товара
                    doc_1 = sheet[f"BC{index}"].value
                    doc_2 = sheet[f"BD{index}"].value
                    doc_3 = sheet[f"BE{index}"].value
                    doc_4 = sheet[f"BF{index}"].value
                    doc_5 = sheet[f"BG{index}"].value

                    doc = ProductDocument.objects.filter(product=product).exists()
                    if doc == False:
                        _save_document_instart(product, doc_1, "Other", "Каталог")
                        _save_document_instart(
                            product, doc_2, "InstallationProduct", "Руководство по эксплуатации"
                        )
                        _save_document_instart(product, doc_3, "Passport", "Паспорт")
                        _save_document_instart(product, doc_4, "Certificates", "Cертификат")
                        _save_document_instart(product, doc_5, "Models3d", "3d модель")
                    print("after doc")
                    # пропсы товара
                    ip = sheet[f"X{index}"].value
                    ip = ip.split("-")
                    ip = ip[1].strip()

                    props = {
                        "Мощность, кВт(общепромыш-ленный режим (G))": sheet[f"E{index}"].value,
                        "Мощность, кВт(насосный режим (P))": sheet[f"F{index}"].value,
                        "Входной ток, А(общепромыш-ленный режим (G))": sheet[f"G{index}"].value,
                        "Входной ток, А(насосный режим (P))": sheet[f"H{index}"].value,
                        "Номинальный выходной ток": sheet[f"I{index}"].value,
                        "Напряжение, В": sheet[f"J{index}"].value,
                        "Макс.вых-ое напр-ие": sheet[f"K{index}"].value,
                        "С встроенными тормозными сопротивлениями": sheet[f"L{index}"].value,
                        "фильтр ЭМС": sheet[f"M{index}"].value,
                        "Для кабеля. м.": sheet[f"N{index}"].value,
                        "Сопротивление, Ом": sheet[f"O{index}"].value,
                        "Номинальный ток, А": sheet[f"P{index}"].value,
                        "Пиковый ток, А": sheet[f"Q{index}"].value,
                        "Комплексная защита двигателя от перегрузки": sheet[f"V{index}"].value,
                        "Внутренний байпас": sheet[f"W{index}"].value,
                        "Степень защиты (IP)": ip,
                        "Поддержка протокола PROFINET IO": sheet[f"Y{index}"].value,
                        "Макс. частота на выходе": sheet[f"Z{index}"].value,
                        "Поддержка протокола TCP": sheet[f"AA{index}"].value,
                        "Поддержка протокола MODBUS": sheet[f"AB{index}"].value,
                        "Поддержка протокола PROFIBUS": sheet[f"AC{index}"].value,
                        "Количество цифров. входов": sheet[f"AD{index}"].value,
                        "Количество аналог. выходов": sheet[f"AE{index}"].value,
                        "Количество аналог. входов": sheet[f"AF{index}"].value,
                        "Количество цифров. выходов": sheet[f"AG{index}"].value,
                        "Количество вход. фаз": sheet[f"AH{index}"].value,
                        "Количество выход. фаз": sheet[f"AI{index}"].value,
                        "Количество HW-интерфейсов RS-485": sheet[f"AJ{index}"].value,
                        "Импульсные входы": sheet[f"AK{index}"].value,
                        "Импульсные выходы": sheet[f"AL{index}"].value,
                        "Частота сети": sheet[f"AM{index}"].value,
                        "Рабочая температура": sheet[f"AN{index}"].value,
                        "Cтрана производства": sheet[f"AO{index}"].value,
                        "Вес, кг": sheet[f"AP{index}"].value,
                        "Длина, м": sheet[f"AQ{index}"].value,
                        "Ширина, м": sheet[f"AR{index}"].value,
                        "Высота, м": sheet[f"AS{index}"].value,
                        "Объем, м3": sheet[f"AT{index}"].value,
                    }
                    
                    _save_prop_instart(product,props)
                    print("after prop")
            except Exception as e:
                print("ERROR ",article_supplier)
                print(e)
                tr = traceback.format_exc()
                error = "file_error"
                location = "Разбор фаилов Instart конкретный товар "

                info = f"{article_supplier}pars_instart_xlsx{tr}{e}"
                e = error_alert(error, location, info)
          
    except Exception as e:
        print(e)
        tr = traceback.format_exc()
        error = "file_error"
        location = "Разбор фаилов Instart"

        info = f"ошибка при чтении фаила pars_instart_xlsx{tr}{e}"
        e = error_alert(error, location, info)


def _instart_product_upd(product, **kwargs):
    for k, v in kwargs.items():
        setattr(product, k, v)
    product.save()


def _is_number_instart(s):
    try:
        float(s)
        return True
    except ValueError:
        return False


def _save_image_instart(product, *img_link):
    for img in img_link:
        if img != None and img != "#N/A":
            image = ProductImage.objects.create(product=product)
            update_change_reason(image, "Автоматическое")
            try:
                image_path = get_file_path_add(image, img)
                p = save_file_product(img, image_path)
                image.photo = image_path
                image.link = img
                image.save()
                update_change_reason(image, "Автоматическое")
            except Exception as e:
                image.delete()


def _save_document_instart(product, link, type, name):
    print("_save_document_instart",link)
    if link != None:
     
        old_doc = ProductDocument.objects.filter(link=link)
        if old_doc.count() > 0:
            old_doc = old_doc[0]
            old_doc.pk = None
            old_doc.product = product
            old_doc.save()
        else:
            document = ProductDocument.objects.create(product=product)
   
            document_path = get_file_path_add(document, link)
            p = save_file_product(link, document_path)
            document.document = document_path
            document.link = link
            document.name = name
            document.save()
            update_change_reason(document, "Автоматическое")
          


def _save_prop_instart(product,prop_arr):
    for k, v in prop_arr.items():
        if v != None:
            property_product = ProductProperty(
                product=product,
                name=k,
                value=v,
            )
            property_product.save()
            update_change_reason(property_product, "Автоматическое")


def get_instart_price_stock():
    try:
        supplier = Supplier.objects.get(slug="instart")
        vendor = Vendor.objects.get(slug="instart")
        
        url = os.environ.get("INSTART_FEED")
        response  = requests.get(url)
        soup = BeautifulSoup(response.content, "xml")
    
        items = soup.find_all("offer")
        for item in items:
            article_sup = item.get('id')
            print(article_sup)
            price_item = item.find('price').text
            print(price_item)
            outlets = item.find('outlets')
            stock_value_spb = 0
            stock_value_msk = 0
            if outlets:
                main_warehouse = outlets.find('outlet', {'warehouse_name': 'Основной/производство - СПБ (FBS)'})
                if main_warehouse:
                    stock_value_spb = main_warehouse.get('instock')
                main_warehouse_msk = outlets.find('outlet', {'warehouse_name': 'склад Москва-Вагоноремонтная (FBS)'})
                if main_warehouse_msk:
                    stock_value_msk = main_warehouse_msk.get('instock')
                    
            all_stock = stock_value_spb + stock_value_msk
            
            try:
                product = Product.objects.get(
                                supplier=supplier,
                                vendor=vendor,
                                additional_article_supplier=article_sup,
                            )
                price_product = Price.objects.get(prod=product)
                if price_item == 0:
                    price_product.extra_price = True
                    price_product.price_supplier = None
                else:
                    price_product.price_supplier = float(price_item)
                    
                price_product._change_reason = "Автоматическое"    
                price_product.save()
                    
                stock_prod = Stock.objects.get(prod=product)
                stock_prod.stock_supplier = all_stock
                stock_prod._change_reason = "Автоматическое"
                stock_prod.save()
        
                
            except Product.DoesNotExist:
                if all_stock > 0:
                    tr = traceback.format_exc()
                    error = "info_error"
                    location = "Фид ежедневный Instart отдельный товар"

                    info = f"{article_sup} - такого твоара нет в каталоге- обновите каталог Product.DoesNotExist{tr}"
                    e = error_alert(error, location, info)
            except Product.MultipleObjectsReturned:
                tr = traceback.format_exc()
                error = "info_error"
                location = "Фид ежедневный Instart отдельный товар "

                info = f"{article_sup}Product.MultipleObjectsReturned НЕСКОЛЬКО ТОВАРОВ С ТАКИМ АРТИКУЛОМ(ДОП АРТИКУЛОМ поставщика){tr}"
                e = error_alert(error, location, info)
            
    except Exception as e:
        print(e)
        tr = traceback.format_exc()
        error = "file_error"
        location = "Фид ежедневный Instart"
        info = f"get_instart_price_stock{tr}{e}"
        e = error_alert(error, location, info)