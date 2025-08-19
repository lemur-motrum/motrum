import traceback
import openpyxl
import re
import pandas as pd

from apps.core.utils import (
    serch_prod_to_motrum_props_article,
    serch_prod_to_motrum_props_categ,
    serch_props_prod_and_add_motrum_props,
    serch_props_prod_and_add_motrum_props_diapason,
)
from apps.logs.utils import error_alert
from apps.product.models import (
    Product,
    ProductPropertyMotrum,
    ProductPropertyValueMotrum,
    VendorPropertyAndMotrum,
)
from apps.supplier.models import Supplier
from project.settings import MEDIA_ROOT


def xlsx_props_motrum():

    try:
        print(2)

        def smart_cast(x):
            try:
                return int(x)
            except ValueError:
                try:
                    return float(x.replace(",", "."))
                except ValueError:
                    return x

        # Открываем исходный файл
        input_path = f"{MEDIA_ROOT}/documents/filter/filter_comma2.xlsx"
        wb = openpyxl.load_workbook(input_path)
        sheet_names = wb.sheetnames

        for sheet_name in sheet_names:
            print(sheet_name)
            if (
                sheet_name == "РromPower (привод.техника) ПЧ "
                or "РromPower (привод.техника) УПП "
                or "PromPower  (Резисторы и "
                or "PP (привод.техника)Синус-фильтр"
            ):
                supplier_sheets = Supplier.objects.get(slug="prompower")
            elif sheet_name == "ONI ":
                supplier_sheets = Supplier.objects.get(slug="iek")

            if sheet_name == "Характеристики Мотрум ":
                continue  # пропускаем этот лист

            if sheet_name != "ONI ":
                continue  # пропускаем этот лист

            ws = wb[sheet_name]

            print(f"Обрабатывается лист: {sheet_name}")

            prev_name_motrum = None
            prev_supplier = None
            prev_name_supplier = None
            not_product = []
            for idx, row in enumerate(
                ws.iter_rows(min_row=2, min_col=1, max_col=8, values_only=True), start=2
            ):

                print(3)
                try:
                    if idx == 14:

                        (
                            name_motrum,
                            value_motrum,
                            supplier,
                            name_supplier,
                            value_supplier,
                            unit,
                            values,
                            articles,
                        ) = [str(x).strip() if x is not None else None for x in row]

                        # Если есть значение в value_motrum
                        if value_motrum:
                            # Если name_motrum пустой, берем из предыдущей строки
                            if not name_motrum:
                                name_motrum = prev_name_motrum
                            else:
                                prev_name_motrum = name_motrum
                            # Если supplier пустой, берем из предыдущей строки
                            if not supplier:
                                supplier = prev_supplier
                            else:
                                prev_supplier = supplier
                            # Если name_supplier пустой, берем из предыдущей строки
                            if not name_supplier:
                                name_supplier = prev_name_supplier
                            else:
                                prev_name_supplier = name_supplier
                        else:
                            # Если value_motrum пустой, не обновляем prev_* значения
                            pass

                        print("Наименование характеристик Мотрум:", name_motrum)
                        print("Вариант значений мотрум:", value_motrum)
                        print("Поставщик:", supplier)
                        print("Название характеристики поставщика:", name_supplier)
                        print("Вариант значений поставщика:", value_supplier)
                        print("Единица измерения (если есть):", unit)
                        print("Значения:", values)
                        print("артикул:", articles)

                        print("-----")
                except Exception as e:
                    print(e)
                    tr = traceback.format_exc()
                    error = "file_error"
                    location = "РАзбор х-к мотрум из фаила"

                    info = f"Конкретный товар строка {sheet_name}{idx}{row}{e}{tr}"
                    e = error_alert(error, location, info)
            print(not_product)
    except Exception as e:
        print(e)
        tr = traceback.format_exc()
        error = "file_error"
        location = "РАзбор х-к мотрум из фаила"

        info = f"общая{e}{tr}"
        e = error_alert(error, location, info)


def create_motrum_props(name_motrum, value_motrum, is_diapason,name_to_slug):
    prod_prop_motrum, created = ProductPropertyMotrum.objects.get_or_create(
        name=name_motrum, is_diapason=is_diapason, name_to_slug=name_to_slug
    )

    if prod_prop_motrum.is_diapason:
        prod_prop_value_motrum, created = (
            ProductPropertyValueMotrum.objects.get_or_create(
                property_motrum=prod_prop_motrum, is_diapason=is_diapason
            )
        )
    else:
        prod_prop_value_motrum, created = (
            ProductPropertyValueMotrum.objects.get_or_create(
                property_motrum=prod_prop_motrum, value=value_motrum
            )
        )

    return (prod_prop_motrum, prod_prop_value_motrum)


def create_motrum_props_and_vendor(
    supplier,
    vendor,
    property_motrum,
    property_value_motrum,
    name_supplier,
    value_supplier,
    is_diapason,
    is_category,
):
    obj, created = VendorPropertyAndMotrum.objects.get_or_create(
        supplier=supplier,
        property_motrum=property_motrum,
        property_value_motrum=property_value_motrum,
        property_vendor_name=name_supplier,
        property_vendor_value=value_supplier,
        is_diapason=is_diapason,
        is_category=is_category,
    )
    print("create_motrum_props_and_vendor", obj, created)
    return (obj, created)


def xlsx_props_motrum_pandas():
    try:
        input_path = f"{MEDIA_ROOT}/documents/filter/filter.xlsx"
        # Используем converters, чтобы все значения читались как строки
        df_dict = pd.read_excel(
            input_path,
            sheet_name=None,
            dtype=str,
            # converters={i: str for i in range(8)}  # 9 столбцов, все как строки
        )
        print("PANDAS")
        for sheet_name, df in df_dict.items():
            print(sheet_name)
            if (
                sheet_name == "РromPower (привод.техника) ПЧ"
                or sheet_name == "РromPower (привод.техника) УПП"
                or sheet_name == "PromPower Резисторы"
                or sheet_name == "PP (привод.техника)Синус-фильтр"
                or sheet_name == "РP (привод.техника)Электродвиг"
                or sheet_name == "Фильтр Датчики (Индуктивные)"
                or sheet_name == "РP (привод.техника)Серво"
                or sheet_name == "РP Систем. автомат ПЛК 12.08"
            ):
                supplier_sheets = Supplier.objects.get(slug="prompower")
            elif sheet_name == "ONI ПЧ":
                supplier_sheets = Supplier.objects.get(slug="iek")
            if sheet_name == "Характеристики Мотрум ":
                continue  # пропускаем этот лист

            # if sheet_name != "РP (привод.техника)Серво":
            #     continue  # пропускаем этот лист
            print(f"Обрабатывается лист: {sheet_name}")
            prev_name_motrum = None
            prev_supplier = None
            prev_name_supplier = None
            prev_to_slug_name = None
            not_product = []
            for idx, row in df.iterrows():
                try:
                    # pandas индексация с 0, Excel с 2, поэтому idx+2
                    if idx + 2 > 0:
                        # row[1] и row[4] — потенциально числа с точкой
                        # row[3] — проверка на 'ТТТТ'
                        val_1 = (
                            str(row[1]).replace(".", ",")
                            if pd.notna(row[1]) and str(row[3]).strip() != "ТТТТ"
                            else str(row[1]) if pd.notna(row[1]) else None
                        )
                        val_4 = (
                            str(row[4]).replace(".", ",")
                            if pd.notna(row[4]) and str(row[3]).strip() != "ТТТТ"
                            else str(row[4]) if pd.notna(row[4]) else None
                        )
                        name_motrum = str(row[0]).strip() if pd.notna(row[0]) else None
                        value_motrum = val_1.strip() if val_1 is not None else None
                        supplier = str(row[2]).strip() if pd.notna(row[2]) else None
                        name_supplier = (
                            str(row[3]).strip() if pd.notna(row[3]) else None
                        )
                        value_supplier = val_4.strip() if val_4 is not None else None
                        unit = str(row[5]).strip() if pd.notna(row[5]) else None
                        values = str(row[6]).strip() if pd.notna(row[6]) else None
                        articles = str(row[7]).strip() if pd.notna(row[7]) else None
                        category = str(row[8]).strip() if pd.notna(row[8]) else None
                        to_slug_name = str(row[9]).strip() if pd.notna(row[9]) else None
                        
                        
                        if value_motrum:
                            if not name_motrum:
                                name_motrum = prev_name_motrum
                            else:
                                prev_name_motrum = name_motrum
                            if not supplier:
                                supplier = prev_supplier
                            else:
                                prev_supplier = supplier
                            if not name_supplier:
                                name_supplier = prev_name_supplier
                            else:
                                prev_name_supplier = name_supplier
                            if not to_slug_name:
                                to_slug_name = prev_to_slug_name
                            else:
                                prev_to_slug_name = to_slug_name
                                
                        print("Наименование характеристик Мотрум:", name_motrum)
                        print("Вариант значений мотрум:", value_motrum)
                        print("Поставщик:", supplier)
                        print("Название характеристики поставщика:", name_supplier)
                        print("Вариант значений поставщика:", value_supplier)
                        print("Единица измерения (если есть):", unit)
                        print("Значения:", values)
                        print("Список товаров:", articles)
                        print("Категория группы :", category)
                        print("Слаг нейм:", to_slug_name)
                        type_save = "normal"
                        if value_motrum == "диапазон":
                            is_diapason = True
                            type_save = "diapason"
                        else:
                            is_diapason = False
                        if values == "Одинаковые значения":
                            is_more_val = True
                            type_save = "multi"
                        else:
                            is_more_val = False
                        if values == "артикулы":
                            is_article = True
                            type_save = "article"
                        if values == "Группа":
                            is_article = True
                            type_save = "categ"
                        else:
                            is_article = False
                        print(type_save)
                        if (
                            type_save == "normal"
                            and name_supplier
                            and value_supplier != "-"
                        ):
                            prod_prop_motrum, prod_prop_value_motrum = (
                                create_motrum_props(
                                    name_motrum, value_motrum, is_diapason,to_slug_name
                                )
                            )
                            vendor_property_and_motrum, created = (
                                create_motrum_props_and_vendor(
                                    supplier_sheets,
                                    None,
                                    prod_prop_motrum,
                                    prod_prop_value_motrum,
                                    name_supplier,
                                    value_supplier,
                                    is_diapason,
                                    False,
                                )
                            )
                            serch_props_prod_and_add_motrum_props(
                                vendor_property_and_motrum, supplier_sheets
                            )
                        if type_save == "multi":
                            val = value_motrum.split("||")
                            for val_item in val:
                                val_item = val_item.strip()
                                prod_prop_motrum, prod_prop_value_motrum = (
                                    create_motrum_props(
                                        name_motrum, val_item, is_diapason,to_slug_name
                                    )
                                )
                                vendor_property_and_motrum, created = (
                                    create_motrum_props_and_vendor(
                                        supplier_sheets,
                                        None,
                                        prod_prop_motrum,
                                        prod_prop_value_motrum,
                                        name_supplier,
                                        val_item,
                                        is_diapason,
                                        False,
                                    )
                                )
                                serch_props_prod_and_add_motrum_props(
                                    vendor_property_and_motrum, supplier_sheets
                                )

                        if type_save == "article":
                            prod_prop_motrum, prod_prop_value_motrum = (
                                create_motrum_props(
                                    name_motrum, value_motrum, is_diapason,to_slug_name
                                )
                            )
                            if articles:
                                articles_list = articles.split(",")
                                for article in articles_list:
                                    article = article.strip()
                                    if supplier_sheets.slug == "prompower":
                                        article_rpl = article.replace("-", "")
                                    elif supplier_sheets.slug == "iek":
                                        article_rpl = article
                                    else:
                                        article_rpl = article
                                    serch_prod_to_motrum_props_article(
                                        prod_prop_motrum,
                                        prod_prop_value_motrum,
                                        article_rpl,
                                        supplier_sheets,
                                        is_diapason,
                                    )
                                    products = Product.objects.filter(
                                        supplier=supplier_sheets,
                                        article_supplier=article_rpl,
                                    )
                                    if not products:
                                        if article not in not_product:
                                            not_product.append(article)
                        if type_save == "diapason":
                            prod_prop_motrum, prod_prop_value_motrum = (
                                create_motrum_props(
                                    name_motrum, value_motrum, is_diapason,to_slug_name
                                )
                            )
                            vendor_property_and_motrum, created = (
                                create_motrum_props_and_vendor(
                                    supplier_sheets,
                                    None,
                                    prod_prop_motrum,
                                    None,
                                    name_supplier,
                                    None,
                                    is_diapason,
                                    False,
                                )
                            )
                            serch_props_prod_and_add_motrum_props_diapason(
                                vendor_property_and_motrum, supplier_sheets
                            )

                        if type_save == "categ":
                            prod_prop_motrum, prod_prop_value_motrum = (
                                create_motrum_props(
                                    name_motrum, value_motrum, is_diapason,to_slug_name
                                )
                            )
                            print("prod_prop_motrum, prod_prop_value_motrum",prod_prop_motrum, prod_prop_value_motrum)
                            vendor_property_and_motrum, created = (
                                create_motrum_props_and_vendor(
                                    supplier_sheets,
                                    None,
                                    prod_prop_motrum,
                                    prod_prop_value_motrum,
                                    name_supplier,
                                    value_supplier,
                                    is_diapason,
                                    True,
                                )
                            )
                            print("vendor_property_and_motrum, created",vendor_property_and_motrum, created)
                            if category:
                                category_name = category.strip()
                                print(category_name)
                                serch_prod_to_motrum_props_categ(
                                    prod_prop_motrum,
                                    prod_prop_value_motrum,
                                    None,
                                    supplier_sheets,
                                    is_diapason,
                                    None,
                                    None,
                                    category_name,
                                )

                        print("-----")
                except Exception as e:
                    print(e)
                    tr = traceback.format_exc()
                    error = "file_error"
                    location = "РАзбор х-к мотрум из фаила (pandas)"
                    info = f"Конкретный товар строка {sheet_name}{idx}{row}{e}{tr}"
                    e = error_alert(error, location, info)
            print(not_product)
    except Exception as e:
        print(e)
        tr = traceback.format_exc()
        error = "file_error"
        location = "РАзбор х-к мотрум из фаила (pandas)"
        info = f"общая{e}{tr}"
        e = error_alert(error, location, info)


def xlsx_to_csv_one_sheet(xlsx_path, output_csv_path):
    import pandas as pd

    # Сначала узнаём, сколько столбцов максимум на всех листах
    xls_file = pd.ExcelFile(xlsx_path)
    max_cols = 0
    for sheet in xls_file.sheet_names:
        df = pd.read_excel(xlsx_path, sheet_name=sheet, header=0)
        max_cols = max(max_cols, len(df.columns))
    # Готовим converters для всех возможных столбцов
    converters = {i: str for i in range(max_cols)}
    # Теперь читаем все листы с converters
    xls = pd.read_excel(
        xlsx_path,
        sheet_name=None,
        converters=converters,
        dtype=object,
        keep_default_na=False,
        na_values=[],
    )
    df_all = []
    for sheet_name, df in xls.items():
        df = df.copy()
        df["sheet_name"] = sheet_name
        df_all.append(df)
    df_merged = pd.concat(df_all, ignore_index=True)
    df_merged.to_csv(output_csv_path, sep=";", index=False, encoding="utf-8")
    print(f"Сохранён общий CSV: {output_csv_path}")


def convert_all_to_text(xlsx_path, output_path):
    wb = openpyxl.load_workbook(xlsx_path)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.value = str(cell.value)
                    cell.number_format = "@"
    wb.save(output_path)
    print(f"Сохранено: {output_path}")
