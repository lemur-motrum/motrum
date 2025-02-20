import datetime
import os
import re
import traceback
import zipfile
import csv
from openpyxl import load_workbook
from simple_history.utils import update_change_reason
from pytils import translit
from django.utils.text import slugify

from apps import supplier
from apps.core.models import Currency, Vat
from apps.logs.utils import error_alert

from apps.product.models import Lot, Price, Product, Stock
from apps.supplier.models import Supplier, Vendor
from project.settings import MEDIA_ROOT


def get_motrum_nomenclature():
    try:
        new_dir = "{0}/{1}".format(MEDIA_ROOT, "ones")
        path_nomenclature = f"{new_dir}/Номенктатура_first.csv"
        path_nomenclature_write_file = f"{new_dir}/Номенктатура_last.csv"
        arr_nomenclature = []

        fieldnames_nomenclature = [
            "Номенклатура",
            "Артикул",
            "Единица измерения",
            "Категория",
            "Склад",
            "Тип",
            "В группе",
            "Код",
        ]
        
        fieldnames_nomenclature_written = [
            "Номенклатура",
            "Артикул",
            "Единица измерения",
            "Категория",
            "Склад",
            "Тип",
            "В группе",
            "Код",
            "Артикул мотрум",
        ]
        path_nomenclature_xlsx = f"{new_dir}/Номенктатура-25.xlsx"
        workbook = load_workbook(path_nomenclature_xlsx)
        sheet = workbook.active
        
        with open(path_nomenclature, "r", newline="", encoding="UTF-8") as csvfile, open(
            path_nomenclature_write_file, "w", encoding="UTF-8"
        ) as writerFile:
            reader_nomenk = csv.DictReader(
                csvfile, delimiter=",", fieldnames=fieldnames_nomenclature
            )
            writer_nomenk = csv.DictWriter(
                writerFile, delimiter=",",fieldnames=fieldnames_nomenclature_written
            )
            # writer_nomenk.writeheader()
            
          
            i = 0
            vendor = ""
            vendor_arr = []
            for row_nomenk in reader_nomenk:
                i += 1
                try:
                        

                    if (
                        i > 2 
                        and row_nomenk["Артикул"] != ""
                        and row_nomenk["Артикул"] != None
                    ):
                        print(i,"NEW STR")
                        cell_value_pass = sheet.cell(row=i, column=8).fill.fgColor.value
                        cell_value_vendor = sheet.cell(row=i, column=7).fill.fgColor.value
                        print("cell_value_pass",cell_value_pass,type(cell_value_pass))
                        print("cell_value_vendor",cell_value_vendor)
                        if cell_value_pass == "00000000":
                            if cell_value_vendor == 8:
                                tag_view = True
                            else:
                                tag_view = False
                            print("tag_view",tag_view)    
                            vendor_row = str(row_nomenk["В группе"]).strip()
                            supplier_qs, vendor_qs = get_or_add_vendor(vendor_row)
                            article_supplier = str(row_nomenk["Артикул"]).strip()
                            article_supplier = " ".join(article_supplier.split())
                            print(article_supplier)
                            lot_str = (
                                str(row_nomenk["Единица измерения"]).replace(".", "").strip()
                            )
                            print("lot",lot_str)
                            lot = Lot.objects.get_or_create(
                                name_shorts=lot_str, defaults={"name": lot_str}
                            )[0]
                            print(lot)
                            name = str(row_nomenk["Номенклатура"]).strip()
                            description = None
                            # поиск товара в окт:
                            if vendor_qs.slug == "emas" or vendor_qs.slug == "tbloc":

                                product = Product.objects.filter(
                                    supplier=supplier_qs, article_supplier=article_supplier
                                )

                                if product:
                                    product = product[0]
                                    product.vendor = vendor_qs
                                    if (
                                        product.name == article_supplier
                                        and article_supplier != name
                                    ):
                                        product.name = name
                                    if product.description != None and description:
                                        product.description = description
                                else:

                                    product = add_new_product(
                                        supplier_qs,
                                        article_supplier,
                                        vendor_qs,
                                    )
                                    product.name = name
                                    product.description = description

                            else:
                                product = Product.objects.filter(
                                    vendor=vendor_qs, article_supplier=article_supplier
                                )

                                if product:
                                    product = product[0]
                                    if (
                                        product.name == article_supplier
                                        and article_supplier != name
                                    ):
                                        product.name = name
                                    if product.description != None and description:
                                        product.description = description
                                else:
                                    product = add_new_product(
                                        supplier_qs,
                                        article_supplier,
                                        vendor_qs,
                                    )
                                    product.name = name
                                    # product.description = description
                            product.autosave_tag = False
                            product.in_view_website = tag_view
                            product._change_reason = "Автоматическое"
                            product.save()
                    
                            # update_change_reason(product, "Автоматическое")

                            add_stok_motrum_article(
                                product,
                                lot,
                            )
                        
                            row_nomenk["Артикул мотрум"] = product.article
                            writer_nomenk.writerow(row_nomenk)
                            
                    elif i == 2:
        
                        row_nomenk["Артикул мотрум"] = "Артикул мотрум"
                    
                        writer_nomenk.writerow(row_nomenk)
                    elif i < 2:
                    
                        writer_nomenk.writerow(row_nomenk)
                    else:
                        pass
                except Exception as e:
                    tr =  traceback.format_exc()
                    print(e)
                    error = "file_error"
                    location = "Загрузка фаилов номенклатура "

                    info = f"ошибка при чтении фаила{i}-{e}{tr}"
                    e = error_alert(error, location, info)

        csvfile.close()
        writerFile.close()
    except Exception as e:
        print(e)
        tr =  traceback.format_exc()
        error = "file_error"
        location = "Обработка фаилов номенклатура мотрум"

        info = f"ошибка при чтении фаила{e} { tr}"
        e = error_alert(error, location, info)    
    
def add_new_product(
    supplier_qs,
    article_supplier,
    vendor_qs,
):
    prod_new = Product(
        supplier=supplier_qs,
        name=article_supplier,
        vendor=vendor_qs,
        article_supplier=article_supplier,
        category_supplier_all=None,
        group_supplier=None,
        category_supplier=None,
        add_in_nomenclature=True,
    )
    prod_new.save()
    update_change_reason(prod_new, "Автоматическое")
    currency = Currency.objects.get(words_code="RUB")
    vat = Vat.objects.get(name=20)
    print("save price start")
    price = Price(
        prod=prod_new, currency=currency, vat=vat, extra_price=True, in_auto_sale=False
    )
    price.save()
    print("save price OK", price)
    update_change_reason(price, "Автоматическое")
    return prod_new


def get_or_add_vendor(vendor):
    vendor_name = vendor.strip()
    if vendor_name == "AuCom Electronics":
        vendor_name = "AuCom"
        supplier_qs = Supplier.objects.get(slug="delta")
        
    elif vendor_name == "OptimusDrive":
        vendor_name = "Optimus Drive"
        supplier_qs = Supplier.objects.get(slug="optimus-drive")
    elif vendor_name == "DELTA":
        vendor_name = "Delta Electronics"
        supplier_qs = Supplier.objects.get(slug="delta")
    elif vendor_name == "HIKROBOT":
        supplier_qs = Supplier.objects.get(slug="optimus-drive")
    elif vendor_name == "IEK":
        supplier_qs = Supplier.objects.get(slug="iek")
    elif vendor_name == "ITK":
        supplier_qs = Supplier.objects.get(slug="iek")
    elif vendor_name == "ODOT":
        supplier_qs = Supplier.objects.get(slug="avangard")
    elif vendor_name == "ONI":
        supplier_qs = Supplier.objects.get(slug="iek")
    elif vendor_name == "RAAD" :
        supplier_qs = Supplier.objects.get(slug="emas")
    elif vendor_name == "Roundss":
        supplier_qs = Supplier.objects.get(slug="optimus-drive")
    elif vendor_name == "SINE":
        supplier_qs = Supplier.objects.get(slug="optimus-drive")
    elif vendor_name == "TBLOC" :
        supplier_qs = Supplier.objects.get(slug="emas")
    elif vendor_name == "VEDA" :
        supplier_qs = Supplier.objects.get(slug="veda")
    elif vendor_name == "Veichi" :
        supplier_qs = Supplier.objects.get(slug="optimus-drive")
    elif vendor_name == "Emas" :
        supplier_qs = Supplier.objects.get(slug="emas")
    else:
        supplier_qs = Supplier.objects.get(slug="drugoj")
        
        
        
    slugish = translit.translify(vendor_name)
    slugish = slugify(slugish)

    try:
        vendor_qs = Vendor.objects.get(slug=slugish)
        # supplier_qs = vendor_qs.supplier

    except Vendor.DoesNotExist:

        supplier_qs = Supplier.objects.get(slug="drugoj")
        vat_catalog = Vat.objects.get(name=20)
        currency_catalog = Currency.objects.get(words_code="RUB")

        vendor_qs = Vendor.objects.create(
            name=vendor_name,
            vat_catalog=vat_catalog,
            currency_catalog=currency_catalog,
        )

    return (supplier_qs, vendor_qs)


def add_stok_motrum_article(
    product,
    lot,
):
    product_stock = Stock.objects.get_or_create(
        prod=product,
        defaults={
            "lot": lot,
            "stock_supplier": 1,
        },
    )

    stock = product_stock[0]

    stock.save()

