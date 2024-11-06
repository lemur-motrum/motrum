import datetime
from math import prod
import re
import openpyxl as openxl
from openpyxl import load_workbook
from pytils import translit
from django.utils.text import slugify
from simple_history.utils import update_change_reason

from apps import supplier
from apps.core.models import Currency, Vat
from apps.product.models import Lot, Price, Product, Stock
from apps.supplier.models import Supplier, Vendor
from project.settings import MEDIA_ROOT


def get_motrum_storage():
    new_dir = "{0}/{1}".format(MEDIA_ROOT, "ones")
    path_storage_motrum = f"{new_dir}/test_ooo.xlsx"
    path_storage_pnm = f"{new_dir}/test_pmn.xlsx"

    def get_group_rows(sheet):
        # a =  [row[0].row for row in sheet.iter_rows(2) if row[0].alignment.indent == 0.0]
        row_index_groupe = [
            row_index
            for row_index, row_dimension in sheet.row_dimensions.items()
            if row_index > 2 and row_dimension.outline_level == 0
        ]

        return row_index_groupe

    def reed_workbook(input_file):
        workbook = load_workbook(input_file)
        data_sheet = workbook.active
        group_rows = get_group_rows(data_sheet)

        # vendor_name = None
        vendor_qs = None
        supplier_qs = None
        lot_auto = Lot.objects.get(name_shorts="шт")

        for index in range(3, data_sheet.max_row):
            row_level = data_sheet.row_dimensions[index].outline_level + 1
            item_value = data_sheet.cell(row=index, column=2).value

            # работа с стоками поставщиков
            if row_level == 1:
                vendor_str = data_sheet.cell(row=index, column=1).value
                vendor_str = vendor_str.replace(",", "").strip()
                if (
                    vendor_str == "Товары"
                    or vendor_str == "Разное"
                    or vendor_str == ""
                    or vendor_str == "Оборудование (объекты основных средств)"
                ):
                    vendor_str = "Hеизвестный"
                elif vendor_str == "OptimusDrive":
                    vendor_str = "Optimus Drive"
                elif vendor_str == "Delta":
                    vendor_str = "Delta Electronics"

                slugish = translit.translify(vendor_str)
                slugish = slugify(slugish)

                try:
                    vendor_qs = Vendor.objects.get(slug=slugish)
                    supplier_qs = vendor_qs.supplier

                except Vendor.DoesNotExist:
                    supplier_qs = Supplier.objects.get(slug="neizvestnyij")
                    vat_catalog = Vat.objects.get(name=20)
                    currency_catalog = Currency.objects.get(words_code="RUB")

                    Vendor.objects.create(
                        name=vendor_str,
                        supplier=supplier_qs,
                        vat_catalog=vat_catalog,
                        currency_catalog=currency_catalog,
                    )

            # работа с товарами поставщиков
            else:

                all_fredom_remaining = data_sheet.cell(row=index, column=16).value
                all_reserve_remaining = data_sheet.cell(row=index, column=15).value
                int_stock_motrum = int(all_fredom_remaining)
                int_stock_reserve_motrum = int(all_reserve_remaining)
                article_supplier = data_sheet.cell(row=index, column=1).value


                if article_supplier != "" and article_supplier != None:
                    article_supplier = article_supplier.strip()
                    article_supplier = " ".join(article_supplier.split())
                    # товары находяться в окт с артикулом и производителем
                    try:
                        product = Product.objects.get(
                            vendor=vendor_qs, article_supplier=article_supplier
                        )
                        add_stok_motrum_old_article(product,lot_auto,int_stock_motrum)
                    
                    # товары НЕ находяться в окт с артикулом и производителем
                    except Product.DoesNotExist:

                        if vendor_qs.slug == "emas" or vendor_qs.slug == "tbloc":
                            try:
                                product = Product.objects.get(
                                    supplier=supplier_qs, article_supplier=article_supplier
                                )
                                product.vendor = vendor_qs
                                
                                product.save()
                                update_change_reason(product, "Автоматическое")
                                add_stok_motrum_old_article(product,lot_auto,int_stock_motrum,int_stock_reserve_motrum)
                            except Product.DoesNotExist:   
                                add_new_product_and_stock(supplier_qs,article_supplier,vendor_qs,lot_auto,int_stock_motrum,int_stock_reserve_motrum)
                            
                                
                        # порлностью новый товар
                        else:
                            add_new_product_and_stock(supplier_qs,article_supplier,vendor_qs,lot_auto,int_stock_motrum,int_stock_reserve_motrum)

    reed_workbook(path_storage_pnm)
    # reed_workbook(path_storage_motrum)

# вспомогательные функции
def add_new_product_and_stock(supplier_qs,article_supplier,vendor_qs,lot_auto,int_stock_motrum,int_stock_reserve_motrum):
    prod_new = Product(
        supplier=supplier_qs,
        name=article_supplier,
        vendor=vendor_qs,
        article_supplier=article_supplier,
        category_supplier_all=None,
        group_supplier=None,
        category_supplier=None,
    )
    prod_new.save()
    update_change_reason(prod_new, "Автоматическое")
    product_stock = Stock(
        prod=prod_new,
        lot=lot_auto,
        stock_motrum=int_stock_motrum,
        stock_motrum_reserve = int_stock_reserve_motrum,
    )
    product_stock.save()
    
    update_change_reason(product_stock, "Автоматическое")
    currency = Currency.objects.get(words_code="RUB")
    vat = Vat.objects.get(name=20)
    price =  Price(prod=prod_new,currency=currency,vat=vat,extra_price=True)
    price.save()
    update_change_reason(price, "Автоматическое")
    
def add_stok_motrum_old_article(product,lot_auto,int_stock_motrum,int_stock_reserve_motrum):
    product_stock = Stock.objects.get_or_create(
        prod=product,
        defaults={
            "lot": lot_auto,
        },
    )

    stock = product_stock[0]

    stock.stock_motrum = int_stock_motrum
    stock.stock_motrum_reserve = int_stock_reserve_motrum
    stock._change_reason = "Автоматическое"
    stock.save()
    # update_change_reason(stock, "Автоматическое")
    