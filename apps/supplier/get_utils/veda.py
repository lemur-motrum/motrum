import datetime
import os
import gc
import traceback
from apps.core.models import Currency, Vat
from apps.core.utils import (
    create_article_motrum,
    get_file_path_add,
    save_file_product,
    save_update_product_attr,
)
from apps.logs.utils import error_alert
from apps.product.models import Lot, Price, Product, Stock
from apps.supplier.models import (
    Supplier,
    SupplierCategoryProduct,
    SupplierCategoryProductAll,
    SupplierGroupProduct,
    Vendor,
)
import requests
from project.settings import MEDIA_ROOT
from simple_history.utils import update_change_reason
import re
from django.utils.text import slugify
from pytils import translit
from urllib.parse import urljoin


def veda_api():
    supplier = Supplier.objects.get(slug="veda-mc")
    vendors = Vendor.objects.filter(slug="veda")

    for vendors_item in vendors:
        if vendors_item.slug == "veda":
            vendor_id = vendors_item.id
            vendor_names = vendors_item.slug
            vendor = vendors_item

    # currency = Currency.objects.get(words_code="USD")
    currency = Currency.objects.get(words_code="CNY")
    # добавление товаров

    url = "https://driveshub.ru/vedaorder/api/stock_list/6312174204"

    headers = {
        "Authorization": f"{os.environ.get("VEDA_API_TOKEN")}",
        # 'Cookie': 'csrftoken=SGCeKpFZXtWm7YnZm9Bmf5ThpnuJpHFxCGXIbqv5SWD0w1rT7NCI7WoGsDD7rLMA'
    }

    response = requests.request(
        "GET",
        url,
        headers=headers,
    )
    data = response.json()

    for data_item in data["prices"]:
        try:
            article_suppliers = data_item["materialCode"]
            name = data_item["description"]
            # цены
            vat_include = True
            vat_catalog = Vat.objects.get(name="20")

            price_supplier = float(data_item["salesPrice"])
            if price_supplier == 0:
                price_supplier = None
            else:
                price_supplier = price_supplier + (price_supplier / 100 * 20)

            sale_persent = float(data_item["discountPercent"])
            # остатки
            lot = Lot.objects.get(name="штука")
            stock_supplier = data_item["avaliability"]
            if stock_supplier == "":
                stock_supplier = 0
            lot_complect = 1

            # транзитная инфа
            data_transit = None
            transit_count = None
            transit = data_item["transit"]
            if transit != []:
                for transit_item in transit:

                    fromisoformat_data_transit_new = datetime.datetime.fromisoformat(
                        transit_item["deliveryDate "]
                    )

                    if data_transit != None:
                        fromisoformat_data_transit = datetime.datetime.fromisoformat(
                            data_transit
                        )
                        if fromisoformat_data_transit_new < fromisoformat_data_transit:
                            data_transit = transit_item["deliveryDate "]
                            transit_count = transit_item["count"]
                    else:
                        data_transit = transit_item["deliveryDate "]
                        transit_count = transit_item["count"]

                if data_transit != None:
                    data_transit = datetime.datetime.fromisoformat(data_transit)
                    data_transit = data_transit.date()

            try:
                # если товар есть в бдd

                article = Product.objects.get(
                    supplier=supplier,
                    vendor=vendor,
                    article_supplier=article_suppliers,
                )

                save_update_product_attr(
                    article, supplier, vendor, None, None, None, None, None, name
                )

            except Product.DoesNotExist:
                # если товара нет в бд
                new_article = create_article_motrum(supplier.id)
                article = Product(
                    article=new_article,
                    supplier=supplier,
                    vendor=vendor,
                    article_supplier=article_suppliers,
                    name=name,
                    category_supplier_all=None,
                    group_supplier=None,
                    category_supplier=None,
                )
                article.save()
                update_change_reason(article, "Автоматическое")
            # цены товара
            try:
                price_product = Price.objects.get(prod=article)

            except Price.DoesNotExist:
                price_product = Price(prod=article)

            finally:
                price_product.extra_price = False
                price_product.currency = currency
                price_product.price_supplier = price_supplier
                price_product.vat = vat_catalog
                price_product.vat_include = vat_include
                price_product._change_reason = "Автоматическое"
                price_product.save()
                # update_change_reason(price_product, "Автоматическое")

            # остатки
            try:
                stock_prod = Stock.objects.get(prod=article)
            except Stock.DoesNotExist:
                stock_prod = Stock(
                    prod=article,
                    lot=lot,
                )
            finally:

                stock_prod.transit_count = transit_count
                stock_prod.data_transit = data_transit
                stock_prod.stock_supplier = stock_supplier
                stock_prod.stock_supplier_unit = stock_supplier
                stock_prod.data_update = datetime.datetime.now()
                stock_prod._change_reason = "Автоматическое"
                stock_prod.save()
                # update_change_reason(stock_prod, "Автоматическое")

        except Exception as e:
            print(e)
            tr = traceback.format_exc()
            error = "file_api_error"
            location = "Загрузка api Veda MC"
            info = f"ошибка при чтении товара артикул: {data_item["materialCode"]}. Тип ошибки:{e}{tr}"
            e = error_alert(error, location, info)
        finally:
            continue


def parse_drives_ru_products():
    """
    Парсит товары с drives.ru по артикулу (article_supplier) для товаров с vendor.slug == 'veda'.
    Получает фото (главные и доп), описание и характеристики, сохраняет их в переменные.
    Проверяет совпадение артикула на странице товара.
    """
    from apps.product.models import (
        Product,
        ProductImage,
        ProductDocument,
        ProductProperty,
    )
    from apps.supplier.models import Vendor
    import requests
    from bs4 import BeautifulSoup

    supplier = Supplier.objects.get(slug="veda-mc")
    vendor = Vendor.objects.get(slug="veda")
    # products = Product.objects.filter(article_supplier__in=["PBC01014"])
    products = Product.objects.filter(supplier=supplier, vendor=vendor).iterator(chunk_size=50)
    results = []

    for product in products:
        try:
            result_one = []
            print(product)
            type_code = product.article_supplier  # используем артикул напрямую
            if not type_code:
                continue

            search_url = f"https://drives.ru/search/?query={type_code}"

            response = requests.get(search_url)
            if response.status_code != 200:
                continue
            soup = BeautifulSoup(response.text, "html.parser")

            found = False
            for item in soup.find_all("div", class_="s-products__item"):
                link_tag = item.find("a", href=True)
                if not link_tag:
                    continue
                product_link = link_tag["href"]
                if not product_link.startswith("http"):
                    product_link = f"https://drives.ru{product_link}"

                prod_resp = requests.get(product_link)
                if prod_resp.status_code != 200:
                    continue
                prod_soup = BeautifulSoup(prod_resp.text, "html.parser")

                code_div = prod_soup.find("div", class_="product__code")
                page_article = None
                if code_div:
                    span = code_div.find("span")
                    if span:
                        page_article = span.get_text(strip=True)
                if not page_article or page_article != product.article_supplier:
                    continue  # если артикул не совпал, пропускаем

                # --- Парсинг изображений ---
                main_images = []
                for a in prod_soup.find_all("a", class_="p-images__slider-item"):
                    href = a.get("href")
                    if href:
                        if not href.startswith("http"):
                            href = f"https://drives.ru{href}"
                        main_images.append(href)

                dop_images = []
                for img in prod_soup.find_all("img", class_="p-images__dop-img"):
                    src = img.get("data-src") or img.get("src")
                    if src:
                        if not src.startswith("http"):
                            src = f"https://drives.ru{src}"
                        dop_images.append(src)
                # --- Конец парсинга изображений ---

                # --- Преобразование маленьких доп. изображений в большие ---
                big_dop_images = []
                for img_url in dop_images:
                    big_url = re.sub(r"\.[0-9]+x[0-9]+\.png$", ".970.png", img_url)
                    big_dop_images.append(big_url)
                # --- Конец преобразования ---

                # --- Новый парсинг названия, описания и характеристик ---
                # Название
                name_tag = prod_soup.find("h1", class_="product__h1")
                name = name_tag.get_text(strip=True) if name_tag else None

                # Описание
                desc = None
                desc_block = prod_soup.find("div", id="description")
                if desc_block:
                    desc_inner = desc_block.find("div", class_="desc desc_max ")
                    if desc_inner:
                        desc = desc_inner.get_text(strip=True)
                if desc is None:
                    # fallback: ищем по заголовку "Описание"
                    title_name = prod_soup.find(
                        "div", class_="in-blocks__title-name", string="Описание"
                    )
                    if title_name:
                        parent_item = title_name.find_parent(
                            "div", class_="in-blocks__item"
                        )
                        if parent_item:
                            desc_inner = parent_item.find("div", class_="desc desc_max")
                            if desc_inner:
                                desc = desc_inner.get_text(strip=True)
                if desc is None:
                    # ищем все desc desc_max h-hidden-show, над которыми есть in-blocks__title с Описание
                    for desc_inner in prod_soup.find_all("div", class_="desc desc_max"):
                        found = False
                        for sib in desc_inner.previous_siblings:
                            # Пропускаем не-теги
                            if not getattr(sib, "name", None):
                                continue
                            sib_classes = set(sib.get("class", []))

                            if "in-blocks__title" in sib_classes:
                                title_name = sib.find(
                                    "div", class_="in-blocks__title-name"
                                )
                                if (
                                    title_name
                                    and title_name.get_text(strip=True) == "Описание"
                                ):
                                    desc = desc_inner.get_text(strip=True)
                                    found = True
                                    break
                        if found:
                            break

                # Характеристики
                features = {}
                features_block = prod_soup.find(
                    "div", class_=lambda x: x and x.startswith("features")
                )
                if features_block:
                    for block in features_block.find_all(
                        "div", class_="features-two-val__block"
                    ):
                        name_div = block.find("div", class_="features-two-val__name")
                        value_div = block.find("div", class_="features-two-val__value")
                        if name_div and value_div:
                            fname = name_div.get_text(strip=True)
                            fval = value_div.get_text(strip=True)
                            if fname in ("Описание", "Типовой код", "Бренд"):
                                continue
                            if fname == "ВхШхГ, мм":
                                nums = re.findall(r"[\d,\.]+", fval)
                                if len(nums) == 3:
                                    features["Высота (мм)"] = nums[0]
                                    features["Ширина (мм)"] = nums[1]
                                    features["Глубина (мм)"] = nums[2]
                                else:
                                    features[fname] = fval
                            else:
                                features[fname] = fval
                # --- Конец нового парсинга ---

                # --- Документы с категории ---
                documents = []
                # Найти ссылку на категорию
                cat_block = prod_soup.find("div", class_="product__category")
                cat_a = (
                    cat_block.find("a", class_="product__category-item")
                    if cat_block
                    else None
                )
                if cat_a and cat_a.has_attr("href"):
                    cat_url = cat_a["href"]

                    if not cat_url.startswith("http"):
                        cat_url = f"https://drives.ru{cat_url}"
                    try:

                        cat_resp = requests.get(cat_url)
                        if cat_resp.status_code == 200:
                            cat_soup = BeautifulSoup(cat_resp.text, "html.parser")
                            res_block = cat_soup.find("div", class_="series__resources")
                            if res_block:
                                for item in res_block.find_all(
                                    "div", class_="series__resources-item"
                                ):
                                    h3 = item.find("h3")
                                    doc_type = None
                                    if h3:
                                        h3_text = h3.get_text(strip=True)
                                        if h3_text == "Инструкции и руководства":
                                            doc_type = {
                                                "type": "Doc",
                                                "type_name": "Руководства и Спецификации",
                                            }
                                        elif h3_text == "Каталоги и брошюры":
                                            doc_type = {
                                                "type": "Other",
                                                "type_name": "Другое",
                                            }
                                        elif h3_text == "Чертежи":
                                            doc_type = {
                                                "type": "DimensionDrawing",
                                                "type_name": "Габаритные чертежи",
                                            }
                                    ul = item.find("ul")
                                    if doc_type and ul:
                                        for li in ul.find_all("li"):
                                            a = li.find("a")

                                            if a and a.has_attr("href"):
                                                doc_url = a["href"]
                                                doc_url = urljoin(
                                                    "https://drives.ru", doc_url
                                                )
                                                doc_name = a.get_text(strip=True)

                                                documents.append(
                                                    {
                                                        "name": doc_name,
                                                        "url": doc_url,
                                                        "type": doc_type["type"],
                                                        "type_name": doc_type[
                                                            "type_name"
                                                        ],
                                                    }
                                                )
                    except Exception as e:
                        print(f"Ошибка при парсинге документов: {e}")

                # --- Парсинг хлебных крошек (категория и группа) ---
                category_name = None
                group_name = None
                group_all_name = None
                nav = prod_soup.find("nav", class_="bread__wrap")
                if nav:
                    crumb_items = nav.find_all(attrs={"itemprop": "itemListElement"})
                    if len(crumb_items) >= 2:
                        meta_cat = crumb_items[1].find(
                            "meta", attrs={"itemprop": "name"}
                        )
                        if meta_cat:
                            category_name = meta_cat.get("content")
                    if len(crumb_items) >= 3:
                        meta_group = crumb_items[2].find(
                            "meta", attrs={"itemprop": "name"}
                        )
                        if meta_group:
                            group_name = meta_group.get("content")
                    if len(crumb_items) >= 4:
                        meta_group_all = crumb_items[3].find(
                            "meta", attrs={"itemprop": "name"}
                        )
                        if meta_group_all:
                            group_all_name = meta_group_all.get("content")

                
                result_one = [
                    {
                        "product_id": product.id,
                        "type_code": type_code,
                        "product_link": product_link,
                        "main_images": main_images,
                        # 'dop_images': dop_images,
                        # 'big_dop_images': big_dop_images,
                        "description": desc,
                        "features": features,
                        "page_article": page_article,
                        "name": name,
                        "documents": documents,
                        "category_name": category_name,
                        "group_name": group_name,
                         "group_all_name": group_all_name,
                    }
                ]
                

                for result in result_one:

                    def save_document(doc, product):

                        base_dir = "products"
                        path_name = "documents"
                        base_dir_supplier = supplier
                        base_dir_vendor = vendor

                        new_dir = "{0}/{1}/{2}/{3}/{4}".format(
                            MEDIA_ROOT,
                            base_dir,
                            base_dir_supplier,
                            base_dir_vendor,
                            path_name,
                        )
                        dir_no_path = "{0}/{1}/{2}/{3}".format(
                            base_dir,
                            base_dir_supplier,
                            base_dir_vendor,
                            path_name,
                        )
                        if not os.path.exists(new_dir):
                            os.makedirs(new_dir)

                        if doc:
                            for doc_item in doc:
                                url = doc_item["url"]
                                try:
                                    d = ProductDocument.objects.get(
                                        product=product, link=url
                                    )

                                except ProductDocument.DoesNotExist:
                                    # Проверка наличия расширения файла
                                    doc_filename = url.split("/")[-1]
                                    _, ext = os.path.splitext(doc_filename)
                                    if not ext:
                                        continue  # если нет расширения, пропускаем

                                    other_doc = ProductDocument.objects.filter(
                                        link=url,
                                    )
                                    print(other_doc)
                                    if other_doc:
                                        other_doc_first = other_doc.first()
                                        doc = ProductDocument.objects.create(
                                            product=product,
                                            document=other_doc_first.document,
                                            link=other_doc_first.link,
                                            name=doc_item["name"],
                                            type_doc=doc_item["type"].capitalize(),
                                        )
                                    else:
                                        slugish = translit.translify(doc_item["name"])
                                        base_slug = slugify(slugish)
                                        doc_name = base_slug
                                        doc = ProductDocument.objects.create(
                                            product=product
                                        )
                                        update_change_reason(doc, "Автоматическое")
                                        doc_list_name = url.split("/")
                                        doc_name = doc_list_name[-1]
                                        # Удаляем неразрывные пробелы и обычные пробелы, ограничиваем длину
                                        doc_name = doc_name.replace(
                                            "\u00a0", ""
                                        ).replace(" ", "")[:100]
                                        images_last_list = url.split(".")
                                        type_file = "." + images_last_list[-1]
                                        link_file = f"{new_dir}/{doc_name}"
                                        r = requests.get(url, stream=True)
                                        r.raise_for_status()
                                        with open(link_file, 'wb') as f:
                                            for chunk in r.iter_content(chunk_size=8192):
                                                f.write(chunk)
                                                
                                        # r = requests.get(url, stream=True)
                                        # with open(
                                        #     os.path.join(link_file), "wb"
                                        # ) as ofile:
                                        #     ofile.write(r.content)
                                            
                                        doc.document = f"{dir_no_path}/{doc_name}"
                                        doc.link = url
                                        doc.name = doc_item["name"]
                                        doc.type_doc = doc_item["type"].capitalize()
                                        doc.save()
                                        update_change_reason(doc, "Автоматическое")

                    def save_image(
                        product,
                    ):
                        if len(result["main_images"]) > 0:
                            for img_item in result["main_images"]:
                                img = img_item
                                image = ProductImage.objects.create(product=product)
                                update_change_reason(image, "Автоматическое")
                                image_path = get_file_path_add(image, img)
                                p = save_file_product(img, image_path)
                                image.photo = image_path
                                image.link = img
                                image.save()
                                update_change_reason(image, "Автоматическое")

                    category_all, groupe, categ = get_category_delta(
                        supplier, vendor, result["category_name"], result["group_name"],result["group_all_name"]
                    )
                    if product.category_supplier_all == None or product.category_supplier_all == "":
                        product.category_supplier_all = category_all
                        
                    if product.group_supplier == None or product.group_supplier == "":
                        product.group_supplier = groupe

                    if (
                        product.category_supplier == None
                        or product.category_supplier == ""
                    ):
                        product.category_supplier = categ
                    

                    # if product.description == None or product.description == "":
                    #     product.description = result["description"]
                    product.description = result["description"]
                    product.name = result["name"]
                    # if product.name == None or product.name == "":
                    #     product.name = result["name"]
                    product.autosave_tag = True
                    product._change_reason = "Автоматическое"
                    product.save()

                    image = ProductImage.objects.filter(product=product).exists()
                    if image == False:
                        save_image(product)

             
                    save_document(result["documents"], product)

                    props = ProductProperty.objects.filter(product=product).exists()

                    if props == False:
                        for name, value in features.items():
                            property_product = ProductProperty(
                                product=product,
                                name=name,
                                value=value,
                            )
                            property_product.save()
                            update_change_reason(property_product, "Автоматическое")

                found = True
                break  # нашли нужный товар, дальше не ищем
            if not found:
                print(f"Товар с артикулом {type_code} не найден на drives.ru")
        except Exception as e:
            print(e)
            tr = traceback.format_exc()
            error = "file_api_error"
            location = "Загрузка крыса Veda MC"
            info = f"ошибка при чтении товара артикул: {product}. Тип ошибки:{e}{tr}"
            e = error_alert(error, location, info)
        finally:
            # --- ОЧИСТКА ПАМЯТИ ---
            
            # Удаляем тяжёлые локальные переменные
            if 'prod_soup' in locals():
                del prod_soup
            if 'soup' in locals():
                del soup
            if 'response' in locals():
                del response
            if 'prod_resp' in locals():
                del prod_resp
            if 'r' in locals():  # если r — ответ requests
                try:
                    r.close()  # закрываем соединение
                except:
                    pass
                del r
            if 'result_one' in locals():
                del result_one
            if 'cat_soup' in locals():
                del cat_soup
            if 'cat_resp' in locals():
                del cat_resp
            # Запускаем сборщик мусора
            gc.collect()
            # ------------------------
            continue
    return None


# категории поставщика промповер для товара
def get_category_delta(supplier, vendor, category, group,group_all):
    from apps.supplier.models import (
        SupplierCategoryProduct,
        SupplierCategoryProductAll,
        SupplierGroupProduct,
    )

    groupe = None
    categ = None
    category_all = None
    try:
        category_all = SupplierCategoryProductAll.objects.get(
            supplier=supplier, vendor=vendor, name=group_all
        )

        groupe = category_all.group_supplier
        categ = category_all.category_supplier
        print(" try:1", category_all, groupe, categ)
    except SupplierCategoryProductAll.DoesNotExist:
        try:
            groupe = SupplierGroupProduct.objects.get(
                supplier=supplier, vendor=vendor, name=group
            )
            category_all = None

            categ = groupe.category_supplier
        except SupplierGroupProduct.DoesNotExist:
            try:
                categ = SupplierCategoryProduct.objects.get(
                    supplier=supplier, vendor=vendor, name=category
                )
                category_all = None
                groupe = None

            except SupplierCategoryProduct.DoesNotExist:
                category_all = None
                groupe = None
                categ = None



    return (category_all, groupe, categ)


def parse_drives_ru_category():
    import requests
    from bs4 import BeautifulSoup
    import re
    from django.db import transaction # Для атомарности операций
    """
    Парсит категории, группы и подгруппы VEDA с drives.ru.
    Категории ищутся внутри блока .cat-prod-grid.
    Категория: .cat-prod-name
    Группы: .cat-prod-sub-name (в том числе внутри .hidden-list)
    Подгруппы: заходим на страницу группы и ищем .h-menu__img-ul-1 .h-menu__img-link-1
    """
    try:
        supplier = Supplier.objects.get(slug="veda-mc")
    except Supplier.DoesNotExist:
        print("Поставщик с slug 'veda-mc' не найден.")
        return []
        
    try:
        vendor = Vendor.objects.get(slug="veda")
    except Vendor.DoesNotExist:
        print("Производитель с slug 'veda' не найден.")
        return []

    base_url = "https://drives.ru"
    url = f"{base_url}/products/"
    
    session = requests.Session()
    # session.headers.update({'User-Agent': 'Mozilla/5.0 ...'})

    try:
        response = session.get(url)
        response.raise_for_status()
    except requests.RequestException as e:
        print(f"Ошибка запроса к {url}: {e}")
        return []

    soup = BeautifulSoup(response.text, "html.parser")

    # Проверяем наличие нужного заголовка (упрощено)
    h3 = soup.find("h3", class_="main__title-2")
    if not h3 or "Приводная техника и средства автоматизации VEDA" not in h3.get_text():
        print("Заголовок 'Приводная техника и средства автоматизации VEDA' не найден или не совпадает")
        return []

    cat_grid = soup.find("div", class_="cat-prod-grid")
    if not cat_grid:
        print("cat-prod-grid не найден")
        return []

    categories_data = []
    # --- Парсинг структуры ---
    for cat_prod in cat_grid.find_all("div", class_="cat-prod", recursive=False):
        cat_name_div = cat_prod.find("div", class_="cat-prod-name")
        if not cat_name_div:
            continue
        cat_a = cat_name_div.find("a")
        if not cat_a:
            continue
        category_name = re.sub(r"\s+", " ", cat_a.get_text(strip=True))
        category_url_raw = cat_a["href"]
        category_url = category_url_raw
        if not category_url_raw.startswith("http"):
            category_url = f"{base_url}{category_url_raw}" if not category_url_raw.startswith('/') else f"{base_url}{category_url_raw}"

        groups = []
        # Группы (основные и скрытые)
        all_group_divs = cat_prod.find_all("div", class_="cat-prod-sub-name", recursive=True)
        for group_div in all_group_divs:
            group_a = group_div.find("a")
            if group_a:
                group_name = re.sub(r"\s+", " ", group_a.get_text(strip=True))
                group_url_raw = group_a["href"]
                group_url = group_url_raw
                if not group_url_raw.startswith("http"):
                    group_url = f"{base_url}{group_url_raw}" if not group_url_raw.startswith('/') else f"{base_url}{group_url_raw}"

                # --- Парсинг подгрупп ДЛЯ КОНКРЕТНОЙ ГРУППЫ ---
                subgroups = []
                try:
                    print(f"Запрос к странице группы: {group_url}")
                    group_response = session.get(group_url)
                    group_response.raise_for_status()
                    group_soup = BeautifulSoup(group_response.text, "html.parser")
                    
                    # --- ИСПРАВЛЕННЫЙ ПОДХОД ---
                    # Вариант 1: Ищем контейнер, относящийся к текущей группе
                    # (Требует анализа реального HTML страницы группы)
                    # subgroup_container = group_soup.find('div', {'data-current-group': 'some_identifier'})
                    # if subgroup_container:
                    #     potential_subgroup_links = subgroup_container.select('a.h-menu__img-link-1')
                    # else:
                    #     print(f"  Контейнер подгрупп для {group_url} не найден.")
                    #     continue # Или продолжить с альтернативным методом

                    # Вариант 2: Ищем все ссылки и фильтруем по URL
                    # Сначала находим все потенциальные ссылки на подгруппы
                    all_potential_links = group_soup.select('a.h-menu__img-link-1')
                    
                    if all_potential_links:
                        print(f"Найдено {len(all_potential_links)} потенциальных подгрупп на {group_url}. Фильтруем...")
                        for sub_link in all_potential_links:
                            subgroup_href = sub_link.get('href', '')
                            # Формируем полный URL для подгруппы
                            if subgroup_href:
                                if not subgroup_href.startswith("http"):
                                    subgroup_url_full = f"{base_url}{subgroup_href}" if not subgroup_href.startswith('/') else f"{base_url}{subgroup_href}"
                                else:
                                    subgroup_url_full = subgroup_href
                                
                                # --- КЛЮЧЕВАЯ ПРОВЕРКА ---
                                # Проверяем, принадлежит ли найденная подгруппа текущей группе
                                # URL подгруппы должен начинаться с URL группы
                                if subgroup_url_full.startswith(group_url.rstrip('/')):
                                    # --- Извлекаем данные подгруппы ---
                                    subgroup_name_raw = sub_link.get_text(strip=True)
                                    if not subgroup_name_raw:
                                        subgroup_name_span = sub_link.find('span', class_='h-menu__name')
                                        if subgroup_name_span:
                                            subgroup_name_raw = subgroup_name_span.get_text(strip=True)
                                    
                                    subgroup_name = re.sub(r"\s+", " ", subgroup_name_raw) if subgroup_name_raw else "Без названия"
                                    
                                    subgroups.append({
                                        "subgroup_name": subgroup_name,
                                        "subgroup_url": subgroup_url_full
                                    })
                                    print(f"  Найдена и добавлена подгруппа: {subgroup_name} - {subgroup_url_full}")
                                else:
                                    print(f"  Пропущена подгруппа (не принадлежит группе): {subgroup_href} (Группа: {group_url})")
                        print(f"  Всего добавлено подгрупп для {group_name}: {len(subgroups)}")
                    else:
                        print(f"  Потенциальные подгруппы по h-menu__img-link-1 на {group_url} НЕ НАЙДЕНЫ")

                except requests.RequestException as e:
                    print(f"Ошибка при запросе группы {group_url}: {e}")
                except Exception as e:
                    print(f"Ошибка при парсинге подгрупп для {group_url}: {e}")
                    # import traceback
                    # traceback.print_exc() # Для отладки

                # Добавляем данные группы с ЕЁ (отфильтрованными) подгруппами
                groups.append({
                    "group_name": group_name,
                    "group_url": group_url,
                    "subgroups": subgroups # Только подгруппы этой конкретной группы, благодаря фильтрации
                })

        # -------------------------

        categories_data.append({
            "category_name": category_name,
            "category_url": category_url,
            "groups": groups
        })
    # --- Конец парсинга ---

    # --- Сохранение в БД ---
    # Используем транзакцию для обеспечения целостности данных
    try:
        with transaction.atomic():
            for category_data in categories_data:
                # Получаем или создаем категорию поставщика
                category_obj, created = SupplierCategoryProduct.objects.get_or_create(
                    supplier=supplier,
                    vendor=vendor,
                    name=category_data["category_name"],
                   
                )
               

                for group_item in category_data["groups"]:
                    # Получаем или создаем группу поставщика
                    group_obj, group_created = SupplierGroupProduct.objects.get_or_create(
                        supplier=supplier,
                        vendor=vendor,
                        category_supplier=category_obj,
                        name=group_item["group_name"],
                       
                    )
                    
                    # --- Сохранение подгрупп ---
                    for subgroup_item in group_item.get("subgroups", []):
                        # Создаем или получаем запись подгруппы, связанную с КОНКРЕТНОЙ Группой
                        subgroup_obj, subgroup_created = SupplierCategoryProductAll.objects.get_or_create(
                            supplier=supplier,
                            vendor=vendor,
                            category_supplier=category_obj, # Опционально, если нужно прямое связывание
                            group_supplier=group_obj, # Ключевая связь - с группой
                            name=subgroup_item["subgroup_name"],
                            
                        )
                        

    except Exception as e:
        print(f"Ошибка при сохранении данных в БД: {e}")
        # Транзакция будет откачена автоматически
        return [] # Возвращаем пустой список, если сохранение не удалось

    # -----------------------

    return categories_data


def save_categories_to_excel(parsed_data, filename="veda_categories.xlsx"):
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    """
    Сохраняет данные категорий, групп и подгрупп в Excel файл.
    Структура:
    - Столбец A: Категории
    - Столбец B: Группы (относящиеся к категории)
    - Столбец C: Подгруппы (относящиеся к группе)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Категории VEDA"

    # Заголовки
    headers = ["Категория", "Группа", "Подгруппа"]
    ws.append(headers)
    
    # Стилизация заголовков (опционально)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    for col_num, header in enumerate(headers, 1):
        cell = ws[f"{get_column_letter(col_num)}1"]
        cell.font = header_font
        cell.fill = header_fill

    # Заполнение данными
    row_num = 2 # Начинаем с второй строки, так как первая - заголовки
    for category_data in parsed_data:
        category_name = category_data["category_name"]
        category_written = False # Флаг, чтобы записать категорию только один раз

        if not category_data["groups"]:
            # Если у категории нет групп, всё равно запишем её
            ws[f"A{row_num}"] = category_name
            row_num += 1
        else:
            for group_item in category_data["groups"]:
                group_name = group_item["group_name"]
                group_written = False # Флаг для записи группы

                if not group_item["subgroups"]:
                    # Если у группы нет подгрупп
                    if not category_written:
                        ws[f"A{row_num}"] = category_name
                        category_written = True
                    ws[f"B{row_num}"] = group_name
                    # URL можно тоже добавить, например, в комментарии к ячейке или в соседних столбцах
                    row_num += 1
                else:
                    # Если есть подгруппы
                    for subgroup_item in group_item["subgroups"]:
                        subgroup_name = subgroup_item["subgroup_name"]
                        
                        if not category_written:
                            ws[f"A{row_num}"] = category_name
                            category_written = True
                        if not group_written:
                            ws[f"B{row_num}"] = group_name
                            group_written = True
                        
                        ws[f"C{row_num}"] = subgroup_name
                        # URL подгруппы тоже можно добавить
                        row_num += 1

    # Автонастройка ширины столбцов
    for column in ws.columns:
         max_length = 0
         column_letter = get_column_letter(column[0].column)
         for cell in column:
             try:
                 if len(str(cell.value)) > max_length:
                     max_length = len(str(cell.value))
             except:
                 pass
         adjusted_width = (max_length + 2) # Немного отступа
         if adjusted_width > 100: # Ограничим максимальную ширину
             adjusted_width = 100
         elif adjusted_width < 10: # Минимальная ширина
             adjusted_width = 10
         ws.column_dimensions[column_letter].width = adjusted_width

    try:
        wb.save(filename)
        print(f"Данные успешно сохранены в файл: {filename}")
    except Exception as e:
        print(f"Ошибка при сохранении файла {filename}: {e}")

