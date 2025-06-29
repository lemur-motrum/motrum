import openpyxl
import os

def analyze_excel():
    file_path = r"media\documents\Счет на оплату № НФ-168 от 17.04.2025.xlsx"
    
    print(f"Проверяем файл: {file_path}")
    print(f"Файл существует: {os.path.exists(file_path)}")
    
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        print(f"Лист: {ws.title}")
        print(f"Строк: {ws.max_row}, Столбцов: {ws.max_column}")
        
        # Показываем первые 20 строк
        print("\nПервые 20 строк:")
        for row in range(1, min(21, ws.max_row + 1)):
            row_values = []
            for col in range(1, min(11, ws.max_column + 1)):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    row_values.append(str(cell_value)[:30])  # Обрезаем длинные значения
            if row_values:
                print(f"Строка {row}: {row_values}")
        
        # Показываем объединенные ячейки
        print("\nОбъединенные ячейки:")
        for merged_range in ws.merged_cells.ranges:
            top_left = ws.cell(merged_range.min_row, merged_range.min_col)
            print(f"{merged_range}: {top_left.value}")
        
        wb.close()
        
    except Exception as e:
        print(f"Ошибка: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    analyze_excel() 