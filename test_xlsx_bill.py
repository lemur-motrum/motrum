#!/usr/bin/env python
"""
Тестовый модуль для создания XLSX файла счета
"""
import os
import sys
import django
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Добавляем путь к проекту
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

# Настройка Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'project.settings')
django.setup()

from apps.client.models import Order
from apps.client.utils import create_xlsx_bill
from django.test import RequestFactory

def test_create_xlsx_bill():
    """
    Создает тестовый XLSX файл счета для проверки конвертации
    """
    try:
        # Создаем временный файл
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            xlsx_path = tmp_file.name
        
        # Создаем новую книгу Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Счет"
        
        # Настройка стилей
        header_font = Font(name='Arial', size=12, bold=True)
        normal_font = Font(name='Arial', size=10)
        title_font = Font(name='Arial', size=14, bold=True)
        
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заголовок
        ws['A1'] = "ТЕСТОВЫЙ СЧЕТ"
        ws['A1'].font = title_font
        ws['A1'].alignment = center_alignment
        ws.merge_cells('A1:H1')
        
        # Информация о поставщике
        ws['A3'] = "Поставщик:"
        ws['A3'].font = normal_font
        ws['A3'].alignment = left_alignment
        
        ws['B3'] = "ООО 'Тестовая Компания'"
        ws['B3'].font = normal_font
        ws['B3'].alignment = left_alignment
        ws.merge_cells('B3:H3')
        
        # Информация о покупателе
        ws['A5'] = "Покупатель:"
        ws['A5'].font = normal_font
        ws['A5'].alignment = left_alignment
        
        ws['B5'] = "ООО 'Тестовый Клиент'"
        ws['B5'].font = normal_font
        ws['B5'].alignment = left_alignment
        ws.merge_cells('B5:H5')
        
        # Заголовки таблицы
        headers = ['№', 'Товар', 'Код', 'Кол-во', 'Ед.', 'Цена', 'Сумма', 'Срок поставки']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # Тестовые данные
        test_data = [
            [1, 'Товар 1', 'T001', 2, 'шт', '1000,00', '2000,00', '01.12.2024'],
            [2, 'Товар 2', 'T002', 1, 'шт', '1500,00', '1500,00', '01.12.2024'],
            [3, 'Товар 3', 'T003', 3, 'шт', '500,00', '1500,00', '01.12.2024'],
        ]
        
        for row_idx, row_data in enumerate(test_data, 9):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = normal_font
                cell.alignment = center_alignment
                cell.border = thin_border
        
        # Итоги
        ws['A12'] = "Итого:"
        ws['A12'].font = normal_font
        ws['A12'].alignment = left_alignment
        
        ws['G12'] = "5000,00"
        ws['G12'].font = normal_font
        ws['G12'].alignment = center_alignment
        ws['G12'].border = thin_border
        
        # Настройка ширины столбцов
        column_widths = [3, 20, 8, 6, 4, 10, 10, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Сохраняем файл
        wb.save(xlsx_path)
        print(f"Тестовый XLSX файл создан: {xlsx_path}")
        
        return xlsx_path
        
    except Exception as e:
        print(f"Ошибка создания тестового XLSX файла: {e}")
        return None

if __name__ == "__main__":
    test_create_xlsx_bill() 